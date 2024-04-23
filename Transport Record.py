import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QStackedWidget, QVBoxLayout, QWidget,QDesktopWidget,QSizePolicy,QGraphicsDropShadowEffect,QTableView,QStyledItemDelegate,QMessageBox,QDialog,QCheckBox,QHeaderView
from PyQt5.QtGui import QPixmap,QIcon,QFont
from PyQt5.uic import loadUi
from PyQt5.QtCore import Qt,QSize,QDate,QSortFilterProxyModel,QSize
import os,re
from PyQt5.QtSql import QSqlDatabase, QSqlQuery,QSqlQueryModel  
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import range_boundaries
from utils import *
from sql import open_database
import datetime


class BasePage(QWidget):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)

    def __init__(self, ui_file):
        super().__init__()
        loadUi(ui_file, self)

        self.close_button.clicked.connect(self.close_main_window)   
        self.minimize_button.clicked.connect(self.minimize_main_window)
        self.option_button.clicked.connect(self.pop_side_bar)

        file_pixmap = QPixmap("photo\\file.png")
        invoice_pixmap = QPixmap("photo\\invoice_icon.png")
        
        self.set_triangle(self.triangle1)
        self.set_triangle(self.triangle2)
        self.set_triangle(self.triangle3)
        self.set_triangle(self.triangle4)
        self.set_triangle(self.triangle5)
        self.set_triangle(self.triangle6)
        self.set_triangle(self.triangle7)

        coin_pixmap = QPixmap("photo\\coin.png")
        salary_pixmap = QPixmap("photo\\salary_icon.png")


        self.file_pic.resize(120,60)
        self.file_pic.setPixmap(file_pixmap)
        self.file_pic.setScaledContents(True)

        self.file_pic_2.resize(140,80)
        self.file_pic_2.setPixmap(invoice_pixmap)
        self.file_pic_2.setScaledContents(True)

        self.coin_pic.resize(180,100)
        self.coin_pic.setPixmap(coin_pixmap)
        self.coin_pic.setScaledContents(True)

        self.coin_pic_2.resize(150,80)
        self.coin_pic_2.setPixmap(salary_pixmap)
        self.coin_pic_2.setScaledContents(True)

        shadow = QGraphicsDropShadowEffect(self.sidebar)
        shadow.setBlurRadius(30)  # 控制阴影模糊程度
        shadow.setOffset(3, 0)    # 控制阴影偏移
        self.sidebar.setGraphicsEffect(shadow)

        base_frame_shadow = QGraphicsDropShadowEffect(self.sidebar)
        base_frame_shadow.setBlurRadius(30)  # 控制阴影模糊程度
        base_frame_shadow.setOffset(3, 5)    # 控制阴影偏移
    
        option_button_shadow = QGraphicsDropShadowEffect(self.sidebar)
        option_button_shadow.setBlurRadius(10)  # 控制阴影模糊程度
        option_button_shadow.setOffset(1, 1)  

        close_button_shadow = QGraphicsDropShadowEffect(self.sidebar)
        close_button_shadow.setBlurRadius(20)  # 控制阴影模糊程度
        close_button_shadow.setOffset(1, 1)  

        minimize_button_shadow = QGraphicsDropShadowEffect(self.sidebar)
        minimize_button_shadow.setBlurRadius(20)  # 控制阴影模糊程度
        minimize_button_shadow.setOffset(1, 1)  

        self.sidebar.setVisible(False)

        self.option_button.enterEvent = self.option_onEnter
        self.option_button.leaveEvent = self.option_onLeave

        self.close_button.enterEvent = self.close_onEnter
        self.close_button.leaveEvent = self.close_onLeave

        self.minimize_button.enterEvent = self.minimize_onEnter
        self.minimize_button.leaveEvent = self.minimize_onLeave

        self.add_record_button.enterEvent = lambda event, btn=self.add_record_button: self.button_onEnter(event, btn)
        self.triangle1.enterEvent = lambda event, btn=self.add_record_button: self.button_onEnter(event, btn) 
        self.check_record_button.enterEvent = lambda event, btn=self.check_record_button: self.button_onEnter(event, btn)
        self.triangle2.enterEvent = lambda event, btn=self.check_record_button: self.button_onEnter(event, btn)
        self.add_expenses_button.enterEvent = lambda event, btn=self.add_expenses_button: self.button_onEnter(event, btn)
        self.triangle3.enterEvent = lambda event, btn=self.add_expenses_button: self.button_onEnter(event, btn)
        self.check_expenses_button.enterEvent = lambda event, btn=self.check_expenses_button: self.button_onEnter(event, btn)
        self.triangle4.enterEvent = lambda event, btn=self.check_expenses_button: self.button_onEnter(event, btn)
        self.check_invoice_button.enterEvent = lambda event, btn=self.check_invoice_button: self.button_onEnter(event, btn)
        self.triangle5.enterEvent = lambda event, btn=self.check_invoice_button: self.button_onEnter(event, btn)
        self.check_salary_button.enterEvent = lambda event, btn=self.check_salary_button: self.button_onEnter(event, btn)
        self.triangle7.enterEvent = lambda event, btn=self.check_salary_button: self.button_onEnter(event, btn)
        self.add_salary_button.enterEvent = lambda event, btn=self.add_salary_button: self.button_onEnter(event, btn)
        self.triangle6.enterEvent = lambda event, btn=self.add_salary_button: self.button_onEnter(event, btn)

        if ui_file != 'ui\\add_record_page.ui' :
            self.add_record_button.leaveEvent = lambda event, btn=self.add_record_button: self.button_onLeave(event, btn)
            self.triangle1.leaveEvent = lambda event, btn=self.add_record_button: self.button_onLeave(event, btn)
        if ui_file != 'ui\\check_record_page.ui' :
            self.check_record_button.leaveEvent = lambda event, btn=self.check_record_button: self.button_onLeave(event, btn)
            self.triangle2.leaveEvent = lambda event, btn=self.check_record_button: self.button_onLeave(event, btn)
        if ui_file != 'ui\\add_expenses_page.ui' :
            self.add_expenses_button.leaveEvent = lambda event, btn=self.add_expenses_button: self.button_onLeave(event, btn)
            self.triangle3.leaveEvent = lambda event, btn=self.add_expenses_button: self.button_onLeave(event, btn)
        if ui_file != 'ui\\check_expenses_page.ui' :
            self.check_expenses_button.leaveEvent = lambda event, btn=self.check_expenses_button: self.button_onLeave(event, btn)
            self.triangle4.leaveEvent = lambda event, btn=self.check_expenses_button: self.button_onLeave(event, btn)
        if ui_file != 'ui\\check_invoice_page.ui' :
            self.check_invoice_button.leaveEvent = lambda event, btn=self.check_invoice_button: self.button_onLeave(event, btn)
            self.triangle5.leaveEvent = lambda event, btn=self.check_invoice_button: self.button_onLeave(event, btn)
        if ui_file != 'ui\\add_salary_page.ui' :
            self.add_salary_button.leaveEvent = lambda event, btn=self.add_salary_button: self.button_onLeave(event, btn)
            self.triangle6.leaveEvent = lambda event, btn=self.add_salary_button: self.button_onLeave(event, btn)
        if ui_file != 'ui\\check_salary_page.ui' :
            self.check_salary_button.leaveEvent = lambda event, btn=self.check_salary_button: self.button_onLeave(event, btn)
            self.triangle7.leaveEvent = lambda event, btn=self.check_salary_button: self.button_onLeave(event, btn)

        # 计算屏幕大小
        screen = QDesktopWidget().screenGeometry()
        window_width = screen.width()
        window_height = screen.height()

        # sidebar 尺寸位置
        self.sidebar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.sidebar.setMinimumHeight(window_height-150)
        self.sidebar.setMinimumWidth(150)
        self.sidebar.move(10,15)

        # sidebar 四个 button 和 图片
        button_height = 80
        button_width = 350
        label_x = 20
        label_y = 120
        start_x = label_x + 100  # 120
        start_y = label_y + 100 # 250

        self.add_record_button.setFixedSize(button_width, button_height)
        self.add_record_button.move(start_x,start_y)
        self.triangle1.move(start_x + 200,start_y + 25)

        self.check_record_button.setFixedSize(button_width, button_height)
        self.check_record_button.move(start_x,start_y+button_height)
        self.triangle2.move(start_x + 200,start_y + button_height + 25)

        self.add_expenses_button.setFixedSize(button_width, button_height)
        self.add_expenses_button.move(start_x,start_y + button_height*2 + 150)
        self.triangle3.move(start_x + 200,start_y + button_height*2 + 175)

        self.check_expenses_button.setFixedSize(button_width, button_height)
        self.check_expenses_button.move(start_x,start_y+button_height*3 + 150)
        self.triangle4.move(start_x + 200,start_y+button_height*3 +175)

        self.check_invoice_button.setFixedSize(button_width, button_height)
        self.check_invoice_button.move(start_x,start_y+button_height*5 + 220)
        self.triangle5.move(start_x + 200,start_y+button_height*5 +245)

        self.add_salary_button.setFixedSize(button_width, button_height)
        self.add_salary_button.move(start_x,start_y + button_height*7 + 280)
        self.triangle6.move(start_x + 200,start_y + button_height*7 + 305)

        self.check_salary_button.setFixedSize(button_width, button_height)
        self.check_salary_button.move(start_x,start_y+button_height*8 + 280)
        self.triangle7.move(start_x + 200,start_y+button_height*8 + 305)
  
        self.file_pic.move(label_x,label_y)
        self.record_text.move(label_x + 125 , label_y + 13)

        self.coin_pic.move(label_x - 30 ,label_y + button_height * 2 + 150 )
        self.expenses_text.move(label_x + 125 , label_y + button_height * 2 + 180 )

        self.file_pic_2.move(label_x - 10,label_y + button_height * 5 + 230)
        self.invoice_text.move(label_x + 125 , label_y + button_height * 5 + 250 )

        self.coin_pic_2.move(label_x - 15 ,label_y + button_height * 7 + 284 )
        self.salary_text.move(label_x + 125 , label_y + button_height * 7 + 310 )

        # base_frame 尺寸位置
        self.base_frame.setMinimumHeight(window_height-130)
        self.base_frame.setMinimumWidth(window_width-30)
        self.base_frame.move(0,5)
        self.base_frame.setGraphicsEffect(base_frame_shadow)

        # workspace.resize(window_width,window_height-200)
        # workspace.move(0,100)

        button_x = window_width -  95 
        button_y = 30  
        self.option_button.move(20,20)
        self.option_button.setGraphicsEffect(option_button_shadow)
        self.close_button.move(button_x,button_y)
        self.close_button.setGraphicsEffect(close_button_shadow)
        self.minimize_button.setGraphicsEffect(minimize_button_shadow)
        self.minimize_button.move(button_x - 50,button_y)
    
    def set_triangle(self,triangle):
        triangle_pixmap = QPixmap("photo\\triangle.png")

        triangle.setPixmap(triangle_pixmap)
        triangle.resize(35,35)
        triangle.setScaledContents(True)
    
    def show_message_box(self,title,text):
        # 创建一个警告框弹窗
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle(title)
        text += "       "
        message_box.setText(text)
        message_box.setStandardButtons(QMessageBox.Ok)
        # message_box.setStyleSheet(
        #     "QLabel { color: white; }"
        #     "QMessageBox { background-color: black; color: white; font-family:Times New Roman; font-size:26px;font-color:white}"
        #     "QPushButton { background-color: white; color: black;}"
        # )
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()

    def button_onLeave(self,event,btn):
        style = """
                border-radius:0px;
                border:0px;
                background-color:transparent;
                """
        object_name = btn.objectName()
        # 根据按钮来设置样式
        if object_name == "add_record_button":
            self.add_record_button.setStyleSheet(style)
        elif object_name == "check_record_button":
            self.check_record_button.setStyleSheet(style)
        elif object_name == "add_expenses_button":
            self.add_expenses_button.setStyleSheet(style)
        elif object_name == "check_expenses_button":
            self.check_expenses_button.setStyleSheet(style)
        elif object_name == "check_invoice_button":
            self.check_invoice_button.setStyleSheet(style)
        elif object_name == "add_salary_button":
            self.add_salary_button.setStyleSheet(style)
        elif object_name == "check_salary_button":
            self.check_salary_button.setStyleSheet(style)
    
    def button_onEnter(self,event,btn):
        
        style = """
                border-radius:0px;
                border:0px;
                background-color:white;
                """
        
        object_name = btn.objectName()  # 获取按钮的objectName

        # 根据按钮来设置样式
        if object_name == "add_record_button":
            self.add_record_button.setStyleSheet(style)
        elif object_name == "check_record_button":
            self.check_record_button.setStyleSheet(style)
        elif object_name == "add_expenses_button":
            self.add_expenses_button.setStyleSheet(style)
        elif object_name == "check_expenses_button":
            self.check_expenses_button.setStyleSheet(style)
        elif object_name == "check_invoice_button":
            self.check_invoice_button.setStyleSheet(style)
        elif object_name == "add_salary_button":
            self.add_salary_button.setStyleSheet(style)
        elif object_name == "check_salary_button":
            self.check_salary_button.setStyleSheet(style)
    
    def option_onEnter(self, event):
        self.option_button.setStyleSheet("""background-color: rgba(216,216,216,100);
                    font-family: "Helvetica"; 
                    font-size: 60px; 
                    font-weight: bold; 
                    border-radius: 10px;
                    padding-top: 0px;
                    padding-right: 0px;
                    padding-bottom: 10px;
                    padding-left: 0px;""")
        event.accept()

    def option_onLeave(self, event):
        self.option_button.setStyleSheet("""background-color: transparent;
                    font-family: "Helvetica"; 
                    font-size: 60px; 
                    font-weight: bold; 
                    padding-top: 0px;
                    padding-right: 0px;
                    padding-bottom: 10px;
                    padding-left: 0px;""")
        event.accept()

    def close_onEnter(self, event):
        self.close_button.setStyleSheet("""
                                        background-color:rgba(216,216,216,200);
                                        color:black;
                                        font:Bold;
                                        border-radius: 5px;
                                        font-size:25px;
                                        padding:0px"""
                                        )
        event.accept()

    def close_onLeave(self, event):
        self.close_button.setStyleSheet("""
                                        background-color:transparent;
                                        color:black;
                                        font:Bold;
                                        font-size:25px;
                                        padding:0px""")
        event.accept()

    def minimize_onEnter(self, event):
        self.minimize_button.setStyleSheet("""
                                        background-color:rgba(216,216,216,200);
                                        color:black;
                                        font:Bold;
                                        border-radius: 5px;
                                        font-size:20px;
                                        padding-bottom:2px"""
                                        )
        event.accept()

    def minimize_onLeave(self, event):
        self.minimize_button.setStyleSheet("""
                                        border: 0;
                                        background-color:transparent;
                                        color:black;
                                        font:Bold;
                                        font-size:20px;
                                        padding-bottom:2px""")
        event.accept()

    def close_main_window(self):
        main_window.close()

    def minimize_main_window(self):
        main_window.showMinimized()

    def pop_side_bar(self):
        self.sidebar.setVisible(not self.sidebar.isVisible())

class HomePage(QWidget):
    def __init__(self):
        super().__init__()
        loadUi('ui\\home_page.ui', self)

        bg_pixmap = QPixmap('photo\\BG2.png')
        record_pixmap = QPixmap('photo\\RECORDS.png')
        expenses_pixmap = QPixmap('photo\\EXPENSES.png')
        next_page_pixmap = QPixmap('photo\\arrow.png')
        self.close_button.clicked.connect(BasePage.close_main_window)   
        self.minimize_button.clicked.connect(BasePage.minimize_main_window)

        close_button_shadow = QGraphicsDropShadowEffect(self.close_button)
        close_button_shadow.setBlurRadius(20)  # 控制阴影模糊程度
        close_button_shadow.setOffset(1, 1)  

        minimize_button_shadow = QGraphicsDropShadowEffect(self.close_button)
        minimize_button_shadow.setBlurRadius(20)  # 控制阴影模糊程度
        minimize_button_shadow.setOffset(1, 1)  

        self.close_button.enterEvent = self.close_onEnter
        self.close_button.leaveEvent = self.close_onLeave

        self.minimize_button.enterEvent = self.minimize_onEnter
        self.minimize_button.leaveEvent = self.minimize_onLeave

        new_button_list = []
        history_button_list = []

        new_icon = QIcon("photo\\add_file.png")
        self.add_record_button.setIcon(new_icon)
        new_button_list.append(self.add_record_button)

        history_icon = QIcon("photo\\history.png")
        self.check_record_button.setIcon(history_icon)
        history_button_list.append(self.check_record_button)

        self.add_expenses_button.setIcon(new_icon)
        new_button_list.append(self.add_expenses_button)

        self.check_expenses_button.setIcon(history_icon)
        history_button_list.append(self.check_expenses_button)

        font = QFont("Tw Cen MT Condensed Extra Bold", 40)
        for button in new_button_list:
            shadow = QGraphicsDropShadowEffect(button)
            shadow.setBlurRadius(80)  # 控制阴影模糊程度
            shadow.setOffset(0, 0)    # 控制阴影偏移

            button.setGraphicsEffect(shadow)
            button.setFont(font)
            button.resize(500,150)
            button.setIconSize(QSize(int(button.width() * 1.2), int(button.height() * 1.2)))

    
        for button in history_button_list:
            shadow = QGraphicsDropShadowEffect(button)
            shadow.setBlurRadius(80)  # 控制阴影模糊程度
            shadow.setOffset(0, 0)    # 控制阴影偏移

            button.setGraphicsEffect(shadow)
            button.setFont(font)
            button.resize(550,150)
            button.setIconSize(QSize(int(button.width() * 0.7), int(button.height() * 0.7)))

        self.add_record_button.move(100,1010)
        self.check_record_button.move(650,1010)
        self.add_expenses_button.move(1270,1010)
        self.check_expenses_button.move(1820,1010)

        self.add_record_button.enterEvent = lambda event, btn=self.add_record_button: self.button_onEnter(event, btn)
        self.check_record_button.enterEvent = lambda event, btn=self.check_record_button: self.button_onEnter(event, btn)
        self.add_expenses_button.enterEvent = lambda event, btn=self.add_expenses_button: self.button_onEnter(event, btn)   
        self.check_expenses_button.enterEvent = lambda event, btn=self.check_expenses_button: self.button_onEnter(event, btn)

        self.add_record_button.leaveEvent = lambda event, btn=self.add_record_button: self.button_onLeave(event, btn)
        self.check_record_button.leaveEvent = lambda event, btn=self.check_record_button: self.button_onLeave(event, btn)
        self.add_expenses_button.leaveEvent = lambda event, btn=self.add_expenses_button: self.button_onLeave(event, btn)
        self.check_expenses_button.leaveEvent = lambda event, btn=self.check_expenses_button: self.button_onLeave(event, btn)
            
        self.home_bg.setPixmap(bg_pixmap)
        self.home_bg.setScaledContents(True)

        self.home_bg.setPixmap(bg_pixmap)
        self.home_bg.setScaledContents(True)

        self.record_bg.setPixmap(record_pixmap)
        self.record_bg.setScaledContents(True)

        self.expenses_bg.setPixmap(expenses_pixmap)
        self.expenses_bg.setScaledContents(True)
        

        screen = QDesktopWidget().screenGeometry()
        window_width = screen.width()
        window_height = screen.height()

        # self.next_page_button.resize(1000,1000)
        self.next_page_button.move(window_width - 200,window_height // 2 - 100)
        self.next_page_button.enterEvent = lambda event, btn=self.next_page_button: self.page_button_onEnter(event, btn)
        self.next_page_button.leaveEvent = lambda event, btn=self.next_page_button: self.page_button_onLeave(event, btn)

        self.home_bg.setGeometry(0,0,window_width,window_height)

        # x -  250 ,y + 20
        self.record_bg.setGeometry(10,65,1180,1220)
        self.expenses_bg.setGeometry(1000,85,1430,1200)

        button_x = window_width -  95 
        button_y = 30  
        self.close_button.move(button_x,button_y)
        self.close_button.setGraphicsEffect(close_button_shadow)
        self.minimize_button.setGraphicsEffect(minimize_button_shadow)
        self.minimize_button.move(button_x - 50,button_y)
    
        self.add_record_button.clicked.connect(lambda: main_window.switch_page(1))
        self.check_record_button.clicked.connect(lambda: main_window.switch_page(2))
        self.add_expenses_button.clicked.connect(lambda: main_window.switch_page(3))
        self.check_expenses_button.clicked.connect(lambda: main_window.switch_page(4))
        self.next_page_button.clicked.connect(lambda: main_window.switch_page(6))

    def button_onEnter(self,event,button):
        shadow = QGraphicsDropShadowEffect(button)
        shadow.setBlurRadius(5)  # 控制阴影模糊程度
        shadow.setOffset(10, 10) 

        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "add_record_button":
            self.add_record_button.setGraphicsEffect(shadow)
        elif object_name == "check_record_button":
            self.check_record_button.setGraphicsEffect(shadow)
        elif object_name == "add_expenses_button":
            self.add_expenses_button.setGraphicsEffect(shadow)
        elif object_name == "check_expenses_button":
            self.check_expenses_button.setGraphicsEffect(shadow)

    def button_onLeave(self,event,button):
        shadow = QGraphicsDropShadowEffect(button)
        shadow.setBlurRadius(80)  # 控制阴影模糊程度
        shadow.setOffset(0, 0)    # 控制阴影偏移
    
        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "add_record_button":
            self.add_record_button.setGraphicsEffect(shadow)
        elif object_name == "check_record_button":
            self.check_record_button.setGraphicsEffect(shadow)
        elif object_name == "add_expenses_button":
            self.add_expenses_button.setGraphicsEffect(shadow)
        elif object_name == "check_expenses_button":
            self.check_expenses_button.setGraphicsEffect(shadow)

    def close_onEnter(self, event):
        self.close_button.setStyleSheet("""
                                        background-color:rgba(216,216,216,200);
                                        color:black;
                                        font:Bold;
                                        border-radius: 5px;
                                        font-size:25px;
                                        padding:0px"""
                                        )
        event.accept()

    def close_onLeave(self, event):
        self.close_button.setStyleSheet("""
                                        background-color:transparent;
                                        color:black;
                                        font:Bold;
                                        font-size:25px;
                                        padding:0px""")
        event.accept()

    def minimize_onEnter(self, event):
        self.minimize_button.setStyleSheet("""
                                        background-color:rgba(216,216,216,200);
                                        color:black;
                                        font:Bold;
                                        border-radius: 5px;
                                        font-size:20px;
                                        padding-bottom:2px"""
                                        )
        event.accept()

    def minimize_onLeave(self, event):
        self.minimize_button.setStyleSheet("""
                                        border: 0;
                                        background-color:transparent;
                                        color:black;
                                        font:Bold;
                                        font-size:20px;
                                        padding-bottom:2px""")
        event.accept()
      
    def page_button_onEnter(self,event,button):
        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "next_page_button":
            self.next_page_button.setStyleSheet("border-radius:60px;"
                                                "color:black;"
                                                "border:10px solid black;"
                                                "background-color:transparent;"
                                                )
        elif object_name == "prev_page_button":
            self.prev_page_button.setStyleSheet("border-radius:60px;"
                                                "color:black;"
                                                "border:10px solid black;"
                                                "background-color:transparent;"
                                                )
            
    def page_button_onLeave(self,event,button):
        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "next_page_button":
            self.next_page_button.setStyleSheet("border-radius:60px;"
                                                "color:white;"
                                                "border:10px solid white;"
                                                "background-color:transparent;"
                                                )
        elif object_name == "prev_page_button":
            self.prev_page_button.setStyleSheet("border-radius:60px;"
                                                "color:white;"
                                                "border:10px solid white;"
                                                "background-color:transparent;"
                                                )

class HomePage2(QWidget):
    def __init__(self):
        super().__init__()
        loadUi('ui\\home_page2.ui', self)

        bg_pixmap = QPixmap('photo\\BG2.png')
        invoice_pixmap = QPixmap('photo\\INVOICE.png')
        salary_pixmap = QPixmap('photo\\SALARY.png')
        next_page_pixmap = QPixmap('photo\\arrow.png')
        self.close_button.clicked.connect(BasePage.close_main_window)   
        self.minimize_button.clicked.connect(BasePage.minimize_main_window)

        close_button_shadow = QGraphicsDropShadowEffect(self.close_button)
        close_button_shadow.setBlurRadius(20)  # 控制阴影模糊程度
        close_button_shadow.setOffset(1, 1)  

        minimize_button_shadow = QGraphicsDropShadowEffect(self.close_button)
        minimize_button_shadow.setBlurRadius(20)  # 控制阴影模糊程度
        minimize_button_shadow.setOffset(1, 1)  

        self.close_button.enterEvent = self.close_onEnter
        self.close_button.leaveEvent = self.close_onLeave

        self.minimize_button.enterEvent = self.minimize_onEnter
        self.minimize_button.leaveEvent = self.minimize_onLeave

        new_button_list = []
        history_button_list = []

        new_icon = QIcon("photo\\add_file.png")
        self.add_salary_button.setIcon(new_icon)
        new_button_list.append(self.add_salary_button)

        history_icon = QIcon("photo\\history.png")
        self.check_invoice_button.setIcon(history_icon)
        history_button_list.append(self.check_invoice_button)

        self.check_salary_button.setIcon(history_icon)
        history_button_list.append(self.check_salary_button)

        font = QFont("Tw Cen MT Condensed Extra Bold", 40)
        for button in new_button_list:
            shadow = QGraphicsDropShadowEffect(button)
            shadow.setBlurRadius(80)  # 控制阴影模糊程度
            shadow.setOffset(0, 0)    # 控制阴影偏移

            button.setGraphicsEffect(shadow)
            button.setFont(font)
            button.resize(500,150)
            button.setIconSize(QSize(int(button.width() * 1.2), int(button.height() * 1.2)))

    
        for button in history_button_list:
            shadow = QGraphicsDropShadowEffect(button)
            shadow.setBlurRadius(80)  # 控制阴影模糊程度
            shadow.setOffset(0, 0)    # 控制阴影偏移

            button.setGraphicsEffect(shadow)
            button.setFont(font)
            button.resize(550,150)
            button.setIconSize(QSize(int(button.width() * 0.7), int(button.height() * 0.7)))

        self.check_invoice_button.move(460,1060)
        self.add_salary_button.move(1350,1060)
        self.check_salary_button.move(1900,1060)

        self.check_invoice_button.enterEvent = lambda event, btn=self.check_invoice_button: self.button_onEnter(event, btn)
        self.add_salary_button.enterEvent = lambda event, btn=self.add_salary_button: self.button_onEnter(event, btn)   
        self.check_salary_button.enterEvent = lambda event, btn=self.check_salary_button: self.button_onEnter(event, btn)

        self.check_invoice_button.leaveEvent = lambda event, btn=self.check_invoice_button: self.button_onLeave(event, btn)
        self.add_salary_button.leaveEvent = lambda event, btn=self.add_salary_button: self.button_onLeave(event, btn)
        self.check_salary_button.leaveEvent = lambda event, btn=self.check_salary_button: self.button_onLeave(event, btn)
            
        self.home_bg.setPixmap(bg_pixmap)
        self.home_bg.setScaledContents(True)

        self.home_bg.setPixmap(bg_pixmap)
        self.home_bg.setScaledContents(True)

        self.invoice_bg.setPixmap(invoice_pixmap)
        self.invoice_bg.setScaledContents(True)

        self.salary_bg.setPixmap(salary_pixmap)
        self.salary_bg.setScaledContents(True)
        

        screen = QDesktopWidget().screenGeometry()
        window_width = screen.width()
        window_height = screen.height()

        # self.next_page_button.resize(1000,1000)
        self.prev_page_button.move(20,window_height // 2 - 100)
        self.prev_page_button.enterEvent = lambda event, btn=self.prev_page_button: self.page_button_onEnter(event, btn)
        self.prev_page_button.leaveEvent = lambda event, btn=self.prev_page_button: self.page_button_onLeave(event, btn)

        self.home_bg.setGeometry(0,0,window_width,window_height)

        # x -  250 ,y + 20
        self.invoice_bg.setGeometry(140,65,1160,1220)
        self.salary_bg.setGeometry(1220,65,1200,1220)

        button_x = window_width -  95 
        button_y = 30  
        self.close_button.move(button_x,button_y)
        self.close_button.setGraphicsEffect(close_button_shadow)
        self.minimize_button.setGraphicsEffect(minimize_button_shadow)
        self.minimize_button.move(button_x - 50,button_y)
    
        self.prev_page_button.clicked.connect(lambda: main_window.switch_page(0))
        self.check_invoice_button.clicked.connect(lambda: main_window.switch_page(5))
        self.add_salary_button.clicked.connect(lambda: main_window.switch_page(7))
        self.check_salary_button.clicked.connect(lambda: main_window.switch_page(8))

    def button_onEnter(self,event,button):
        shadow = QGraphicsDropShadowEffect(button)
        shadow.setBlurRadius(5)  # 控制阴影模糊程度
        shadow.setOffset(10, 10) 

        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "check_invoice_button":
            self.check_invoice_button.setGraphicsEffect(shadow)
        elif object_name == "add_salary_button":
            self.add_salary_button.setGraphicsEffect(shadow)
        elif object_name == "check_salary_button":
            self.check_salary_button.setGraphicsEffect(shadow)

    def button_onLeave(self,event,button):
        shadow = QGraphicsDropShadowEffect(button)
        shadow.setBlurRadius(80)  # 控制阴影模糊程度
        shadow.setOffset(0, 0)    # 控制阴影偏移
    
        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "check_invoice_button":
            self.check_invoice_button.setGraphicsEffect(shadow)
        elif object_name == "add_salary_button":
            self.add_salary_button.setGraphicsEffect(shadow)
        elif object_name == "check_salary_button":
            self.check_salary_button.setGraphicsEffect(shadow)

    def close_onEnter(self, event):
        self.close_button.setStyleSheet("""
                                        background-color:rgba(216,216,216,200);
                                        color:black;
                                        font:Bold;
                                        border-radius: 5px;
                                        font-size:25px;
                                        padding:0px"""
                                        )
        event.accept()

    def close_onLeave(self, event):
        self.close_button.setStyleSheet("""
                                        background-color:transparent;
                                        color:black;
                                        font:Bold;
                                        font-size:25px;
                                        padding:0px""")
        event.accept()

    def minimize_onEnter(self, event):
        self.minimize_button.setStyleSheet("""
                                        background-color:rgba(216,216,216,200);
                                        color:black;
                                        font:Bold;
                                        border-radius: 5px;
                                        font-size:20px;
                                        padding-bottom:2px"""
                                        )
        event.accept()

    def minimize_onLeave(self, event):
        self.minimize_button.setStyleSheet("""
                                        border: 0;
                                        background-color:transparent;
                                        color:black;
                                        font:Bold;
                                        font-size:20px;
                                        padding-bottom:2px""")
        event.accept()
      
    def page_button_onEnter(self,event,button):
        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "next_page_button":
            self.next_page_button.setStyleSheet("border-radius:60px;"
                                                "color:black;"
                                                "border:10px solid black;"
                                                "background-color:transparent;"
                                                )
        elif object_name == "prev_page_button":
            self.prev_page_button.setStyleSheet("border-radius:60px;"
                                                "color:black;"
                                                "border:10px solid black;"
                                                "background-color:transparent;"
                                                )
            
    def page_button_onLeave(self,event,button):
        object_name = button.objectName()
        # 根据按钮来设置样式
        if object_name == "next_page_button":
            self.next_page_button.setStyleSheet("border-radius:60px;"
                                                "color:white;"
                                                "border:10px solid white;"
                                                "background-color:transparent;"
                                                )
        elif object_name == "prev_page_button":
            self.prev_page_button.setStyleSheet("border-radius:60px;"
                                                "color:white;"
                                                "border:10px solid white;"
                                                "background-color:transparent;"
                                                )

class add_record_page(BasePage):
    def __init__(self):
        super().__init__('ui\\add_record_page.ui')
        self.add_record_button.setStyleSheet("""
                border-radius:0px;
                border:0px;
                background-color:white;
                """)
        
        self.init_ui()
        self.setup_lorry_frame()
        self.setup_company_frame()

        self.check_record_button.clicked.connect(lambda: main_window.switch_page(2))
        self.add_expenses_button.clicked.connect(lambda: main_window.switch_page(3))
        self.check_expenses_button.clicked.connect(lambda: main_window.switch_page(4))
        self.check_invoice_button.clicked.connect(lambda: main_window.switch_page(5))
        self.add_salary_button.clicked.connect(lambda: main_window.switch_page(7))
        self.check_salary_button.clicked.connect(lambda: main_window.switch_page(8))

        self.add_record_submit_button.clicked.connect(self.record_submit_button_clicked)
        self.add_lorry_submit_button.clicked.connect(self.lorry_submit_button_clicked)
        self.add_company_submit_button.clicked.connect(self.company_submit_button_clicked)

    def record_submit_button_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
        current_date = datetime.date.today()
        formatted_date = current_date.strftime("%d/%m/%Y")

        query = QSqlQuery()
        query.prepare("INSERT INTO Record(date,lorry,reference,company,description,quantity,unit_price,amount,created_at) VALUES (:date,:lorry,:reference,:company,:description,:quantity,:unit_price,:amount,:created_at)")
        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":lorry", set_not_chosen(self.lorry_line.currentText()))
        query.bindValue(":reference", self.reference_line.text())
        query.bindValue(":company", set_not_chosen(self.company_line.currentText()))
        query.bindValue(":description", self.description_line.text())
        query.bindValue(":quantity", self.quantity_line.text())
        query.bindValue(":unit_price", self.unitprice_line.text())
        query.bindValue(":amount", self.amount_line.text())
        query.bindValue(":created_at", formatted_date)

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Submit Record","New Record Added")

    def lorry_submit_button_clicked(self):
        text = self.lorry_line_add.text()
        
        if len(text) == 0:
            return 
        
        str=""

        space = False
        for char in text:
            if char.isdigit() and space is False:
                str += " "
                space = True
            if char == ' ':
                space = True
            str += char
        
        query = QSqlQuery(f"SELECT * FROM Lorry WHERE lorry_number = '{str}'")
        if query.next():
            print("Value already exists in the database.")
            return

        query = QSqlQuery()
        query.prepare("INSERT INTO Lorry(lorry_number) VALUES (:lorry)")
        query.bindValue(":lorry", str)

        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.lorry_data(self.lorry_line)
        self.lorry_data(self.delete_lorry_combobox)

        self.show_message_box(f"Add Lorry",f"Save Lorry: {str} ")

    def company_submit_button_clicked(self):
        add_dialog = AddDebtorDialog(self)
        add_dialog.exec_()

        self.company_data(self.company_line)
        self.company_data(self.company_combobox)

    def update_company_clicked(self):
        if self.company_combobox.currentText() != 'Not Chosen':
            update_dialog = UpdateDebtorDialog(self.company_combobox.currentText())
            update_dialog.exec_()
            self.company_data(self.company_line)
            self.company_data(self.company_combobox)

        else:
            self.show_message_box("Warning","No Company is Chosen.")

    def init_ui(self):
        
        self.workspace_frame.resize(700,1400)
        self.workspace_frame.move(600,60)
        
        self.add_record_label.resize(self.label_width,self.label_height)
        self.add_record_label.setFont(self.l_font)
        self.add_record_label.move(0,0)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(0,80)
        move_y = 80

        self.date_line.setDate(QDate.currentDate())
        self.date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.date_line.setDisplayFormat("dd/MM/yyyy")
        self.date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.date_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.date_line.move(0,move_y)
        self.date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:10px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 28px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white }"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:20px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 35px; height: 35px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 3px solid black; font-family: Constantia; font-size: 28px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:50px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:50px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_left.png); }"
        )
    
        self.lorry_label.resize(self.label_width,self.label_height)
        self.lorry_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.lorry_label.move(0,move_y)        

        self.lorry_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.lorry_line.move(0,move_y)
        self.setup_combo_box(self.lorry_line)
        self.lorry_data(self.lorry_line)

        self.reference_label.resize(self.label_width,self.label_height)
        self.reference_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.reference_label.move(0,move_y)     

        self.reference_line.resize(self.line_width,self.line_height)
        self.reference_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.reference_line.move(0,move_y)
        self.reference_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        
        self.company_label.resize(self.label_width + 100,self.label_height)
        self.company_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.company_label.move(0,move_y)

        self.company_line.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.company_line.move(0,move_y)
        self.setup_combo_box(self.company_line)
        self.company_data(self.company_line)

        self.description_label.resize(self.label_width + 80,self.label_height)
        self.description_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.description_label.move(0,move_y)

        self.description_line.resize(self.line_width + 80,self.line_height)
        self.description_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.description_line.move(0,move_y)
        self.description_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   

        self.quantity_label.resize(self.label_width,self.label_height)
        self.quantity_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.quantity_label.move(0,move_y)

        self.quantity_line.resize(self.line_width,self.line_height)
        self.quantity_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.quantity_line.move(0,move_y)
        self.quantity_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        
        self.unitprice_label.resize(self.label_width,self.label_height)
        self.unitprice_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.unitprice_label.move(0,move_y)

        self.unitprice_line.resize(self.line_width,self.line_height)
        self.unitprice_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.unitprice_line.move(0,move_y)
        self.unitprice_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        
        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        self.add_record_submit_button.resize(250,70)
        self.add_record_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_record_submit_button.move(30,move_y + 90)
        self.add_record_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.submit_creditor_button.resize(300,70)
        self.submit_creditor_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.submit_creditor_button.move(320,move_y + 90)
        self.submit_creditor_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.submit_creditor_button.clicked.connect(self.submit_creditor_button_clicked)
    
    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)

    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def submit_creditor_button_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return

        
        company = set_not_chosen(self.company_line.currentText())
        query = QSqlQuery()
        query.prepare("INSERT INTO Expenses_Record(date,lorry,reference,expenses,company,amount) VALUES (:date,:lorry,:reference,:expenses,:company,:amount)")
        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":lorry", set_not_chosen(self.lorry_line.currentText()))
        query.bindValue(":reference", self.reference_line.text())
        query.bindValue(":expenses", "Subfreige")
        query.bindValue(":company", company)
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        
        if company != '-':
            query_str = f"SELECT name FROM Expenses_Company WHERE name='{company}'"
            query.exec_(query_str)

            if query.next():
                print("Value already exists in the database.")
            else:
                query = QSqlQuery()
                query.prepare("INSERT INTO Expenses_Company(name) VALUES (:name)")
                query.bindValue(":name", company)

                if query.exec():
                    print("Data inserted successfully.")
                else:
                    print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Submit Creditor","Creditor/Subfreige Record Added")

    def delete_lorry_clicked(self):
        lorry = self.delete_lorry_combobox.currentText()

        query = QSqlQuery(f"DELETE FROM Lorry WHERE lorry_number = '{lorry}'")

        if query.exec():
            print("Data deleted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Delete Lorry",f"Lorry {lorry} Deleted")
        self.lorry_data(self.lorry_line)
        self.lorry_data(self.delete_lorry_combobox)

    def delete_company_clicked(self):
        company = self.company_combobox.currentText()

        query = QSqlQuery(f"DELETE FROM Company WHERE name = '{company}'")

        if query.exec():
            print("Data deleted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Delete Company",f"Company {company} Deleted")
        self.company_data(self.company_line)
        self.company_data(self.company_combobox)  
        
    def setup_lorry_frame(self):
        screen = QDesktopWidget().screenGeometry()
        self.add_lorry_frame.resize(800,600)
        self.add_lorry_frame.move(screen.width() - 1200,110)

        font = QFont("Tw Cen MT Condensed Extra Bold",20)
        font.setUnderline(True)
        self.add_lorry_label.resize(self.label_width,self.label_height)
        self.add_lorry_label.setFont(font)
        self.add_lorry_label.move(0,0)

        self.lorry_label_add.resize(self.label_width,self.label_height)
        self.lorry_label_add.setFont(self.s_font)
        self.lorry_label_add.move(0,80)

        self.lorry_label_add_2.resize(self.label_width,self.label_height)
        self.lorry_label_add_2.setFont(self.s_font)
        self.lorry_label_add_2.move(self.line_width + 50,80)

        move_y = 80

        self.lorry_line_add.resize(self.line_width,self.line_height)
        self.lorry_line_add.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.lorry_line_add.move(0,move_y)
        self.lorry_line_add.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        
        self.add_lorry_submit_button.resize(150,70)
        self.add_lorry_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_lorry_submit_button.move(70,move_y + 90)
        self.add_lorry_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.setup_combo_box(self.delete_lorry_combobox)
        self.delete_lorry_combobox.resize(self.line_width,self.line_height)
        self.delete_lorry_combobox.move(self.line_width + 50,move_y)

        self.delete_lorry_button.resize(160,70)
        self.delete_lorry_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_lorry_button.move(self.line_width + 120,move_y + 90)
        self.delete_lorry_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.lorry_data(self.delete_lorry_combobox)
        self.delete_lorry_button.clicked.connect(self.delete_lorry_clicked)

    def setup_company_frame(self):
        self.add_company_frame.resize(850,600)
        screen = QDesktopWidget().screenGeometry()
        self.add_company_frame.move(screen.width() - 1200,600)

        font = QFont("Tw Cen MT Condensed Extra Bold",20)
        font.setUnderline(True)
        self.add_company_label.resize(self.label_width + 180,self.label_height)
        self.add_company_label.setFont(font)
        self.add_company_label.move(0,0)

        self.company_label_add_2.resize(self.label_width + 180,self.label_height)
        self.company_label_add_2.setFont(self.s_font)
        self.company_label_add_2.move(self.label_width + 230,80)

        move_y = self.label_height + 65

        self.add_company_submit_button.resize(350,70)
        self.add_company_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_company_submit_button.move(0,move_y - 10)
        self.add_company_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        
        self.setup_combo_box(self.company_combobox)
        self.company_combobox.resize(self.line_width + 80,self.line_height)
        self.company_combobox.move(self.line_width + 130,move_y)

        self.update_company_button.resize(160,70)
        self.update_company_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_company_button.move(self.line_width + 240,move_y + 110)
        self.update_company_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.delete_company_button.resize(160,70)
        self.delete_company_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_company_button.move(self.line_width + 240,move_y + 200)
        self.delete_company_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.company_data(self.company_combobox)
        self.delete_company_button.clicked.connect(self.delete_company_clicked)
        self.update_company_button.clicked.connect(self.update_company_clicked)

class check_record_page(BasePage):
    current_sum = 0
    query_str = ""
    def __init__(self):
        super().__init__('ui\\check_record_page.ui')
        self.check_record_button.setStyleSheet("""
                border-radius:0px;
                border:0px;
                background-color:white;
                """)

        # self.init_database("Record")
        self.init_ui()
        self.setup_database()
        self.current_page = 0
        self.query_str = "ORDER BY id DESC "
        self.update_model()

        # self.export_to_excel()


        self.add_record_button.clicked.connect(lambda: main_window.switch_page(1))
        self.add_expenses_button.clicked.connect(lambda: main_window.switch_page(3))
        self.check_expenses_button.clicked.connect(lambda: main_window.switch_page(4))
        self.check_invoice_button.clicked.connect(lambda: main_window.switch_page(5))
        self.add_salary_button.clicked.connect(lambda: main_window.switch_page(7))
        self.check_salary_button.clicked.connect(lambda: main_window.switch_page(8))

    def setup_database(self):
        self.model = RecordQueryModel()
        self.data_table.setModel(self.model)
        # self.update_model()
        self.proxy_model = HideFirstColumnProxyModel(self)
        self.proxy_model.setSourceModel(self.model)
        self.data_table.setModel(self.proxy_model)

    def update_model(self):
        if len(self.query_str) == 0:
            self.query_str = "ORDER BY id DESC "

        sum_text = "SELECT amount FROM Record " + self.query_str
        query = QSqlQuery()
        query.exec(sum_text)

        self.current_sum = 0
        
        while query.next():
            if contain_comma(str(query.value(0))):
                self.current_sum += locale.atof(query.value(0))
            else:
                self.current_sum += query.value(0)


        if is_3_decimal_places(self.current_sum):
            self.current_sum_str = locale.format_string("%.3f", self.current_sum, grouping=True)
        else:
            self.current_sum_str = locale.format_string("%.2f", self.current_sum, grouping=True)
                    
                    
        self.amount_line.setText(str(self.current_sum_str))

        offset = self.current_page * ROWS_PER_PAGE 
    
        query = "SELECT * FROM Record " + self.query_str + f"LIMIT {ROWS_PER_PAGE} OFFSET {offset}"
        self.model.setQuery(query)
        self.page_number.setText(str(self.current_page+1))

        self.data_table.setColumnWidth(0, 160)
        self.data_table.setColumnWidth(1, 160)
        self.data_table.setColumnWidth(2, 280)
        self.data_table.setColumnWidth(3, 460)
        self.data_table.setColumnWidth(4, 450)
        self.data_table.setColumnWidth(5, 150)
        self.data_table.setColumnWidth(6, 145)
        self.data_table.setColumnWidth(7, 145)

    def clearCheckboxes(self,scroll_widget):
        scroll_layout = scroll_widget.layout()
        for i in reversed(range(scroll_layout.count())):
            item = scroll_layout.itemAt(i)
            if isinstance(item.widget(), QCheckBox):
                item.widget().deleteLater()

    def lorry_data(self):
        self.clearCheckboxes(self.lorry_scroll_widget)
        self.lorry_checkboxes = []
        scroll_layout = self.lorry_scroll_widget.layout()

        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            self.lorry_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中

    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 38px; height: 38px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def date_data(self,combobox):
        combobox.clear()
        i = 0
        while i<32:
            combobox.addItem(str(i))
            i += 1

        combobox.setCurrentIndex(0)

    def company_data(self):
        self.clearCheckboxes(self.company_scroll_widget)
        self.company_checkboxes = []
        scroll_layout = self.company_scroll_widget.layout()

        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            self.company_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中

    def show_record_date(self):
        self.record_date_label.setVisible(not self.record_date_label.isVisible())
        self.record_date_line.setVisible(not self.record_date_line.isVisible())
        self.record_date_button.setVisible(not self.record_date_button.isVisible())

    def init_ui(self):
        screen = QDesktopWidget().screenGeometry()

        self.check_history_label.resize(self.label_width + 50,self.label_height)
        self.check_history_label.setFont(self.l_font)
        self.check_history_label.move(0,0)

        self.data_frame.resize(screen.width() - 570,screen.height() - 230)
        self.data_frame.move(110,50)

        self.start_date_label.resize(self.label_width,self.label_height)
        self.start_date_label.setFont(self.s_font)
        self.start_date_label.move(950,-20)     

        self.start_date_line.resize(self.line_width - 200,self.line_height-5)
        self.start_date_line.move(1100,0)
        self.setup_combo_box(self.start_date_line)
        self.date_data(self.start_date_line)

        self.end_date_label.resize(self.label_width,self.label_height)
        self.end_date_label.setFont(self.s_font)
        self.end_date_label.move(1240,-20)

        self.end_date_line.resize(self.line_width - 200,self.line_height-5)
        self.end_date_line.move(1380,0)
        self.setup_combo_box(self.end_date_line)
        self.date_data(self.end_date_line)

        self.data_month.resize(self.line_width,self.line_height)
        self.data_month.move(0,self.label_height)
        self.data_month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.data_month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.data_month.addItem("Not Chosen")
        self.data_month.addItems(months)
        self.data_month.setCurrentIndex(0)

        self.data_year.resize(self.line_width,self.line_height)
        self.data_year.move(self.line_width + 20 , self.label_height)
        self.data_year.setMinimum(2022)
        self.data_year.setMaximum(2030)
        self.data_year.setValue(2023)
        self.data_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.data_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

        self.record_date_label.resize(self.label_width,self.label_height)
        self.record_date_label.setFont(self.s_font)
        self.record_date_label.move(650,5)
        self.record_date_label.setVisible(False)


        self.record_date_line.setDate(QDate.currentDate())
        self.record_date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.record_date_line.setDisplayFormat("dd/MM/yyyy")
        self.record_date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.record_date_line.resize(self.line_width,self.line_height)
        self.record_date_line.move(650,self.label_height)
        self.record_date_line.setVisible(False)
        self.record_date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:5px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 20px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white;padding-bottom:5px}"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:10px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 25px; height: 25px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 2px solid black; font-family: Constantia; font-size: 20px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:25px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:25px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 25px; height: 25px; icon-size: 25px, 25px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 25px; height: 25px; icon-size: 25px, 25Spx; qproperty-icon: url(photo/arrow_left.png); }"
        )
        
        self.record_date_button.resize(280,60)
        self.record_date_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.record_date_button.move(680,60)
        self.record_date_button.setStyleSheet(f"background-color:black;color:white;border-radius:30px;")
        self.record_date_button.clicked.connect(self.show_record_date)

        self.data_table.resize(self.data_frame.width(),self.data_frame.height()-220)
        self.data_table.move(0,self.label_height + 60)
        self.data_table.setFont(QFont("Times New Roman",13))
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.data_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.data_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        delegate = CenteredNumberDelegate()
        self.data_table.setItemDelegate(delegate)
        self.data_table.setSelectionBehavior(QTableView.SelectRows)
        self.data_table.hideColumn(0)
        self.data_table.setStyleSheet(
            "QTableView::item:selected{ background-color:black;color:white;}"
            "QTableView { color: black; gridline-color: black; border: 2px solid black; background-color:white ;gridline-width: 2px;}"
            "QHeaderView { background-color: rgba(216,216,216,200); color:black ;font-family: Rockwell ; font-size: 27px; }"

            "QHeaderView::section { background-color: rgba(216,216,216,200) ;color: black;height: 50px;}"
        )


        
        self.prev_button.resize(100,50)
        self.prev_button.move(self.data_table.width()//2 - 150 ,self.data_frame.height()-50)
        self.prev_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.prev_button.setStyleSheet("QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
                                       "QPushButton:hover{ border: 3px solid black }")
        
        self.page_number.resize(self.label_width,self.label_height)
        self.page_number.move(self.data_table.width()//2,self.data_frame.height()-65 )
        self.page_number.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        

        self.next_button.resize(100,50)
        self.next_button.move(self.data_frame.width()//2 + 60,self.data_frame.height()-50)
        self.next_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.next_button.setStyleSheet(
            "QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
            "QPushButton:hover{ border: 3px solid black }"
            )
        
        self.filter_frame.resize(500,1000)
        self.filter_frame.move(screen.width() - 400,100)
        
        self.check_record_lorry_label.resize(self.label_width,self.label_height)
        self.check_record_lorry_label.setFont(self.s_font)
        self.check_record_lorry_label.move(0,0)        

        move_y = self.label_height - 15
        
        self.lorry_scroll.setFixedSize(self.line_width,220)
        self.lorry_scroll.move(0,move_y)
        self.lorry_scroll.setWidgetResizable(True)
        self.lorry_scroll.setStyleSheet(
            "QScrollArea{"
            "    border: 2px solid black; color: black;"
            "}"
        )
        scroll_layout = QVBoxLayout()
        self.lorry_scroll_widget.setLayout(scroll_layout)
        self.lorry_data()

        self.check_record_company_label.resize(self.label_width,self.label_height)
        self.check_record_company_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 170
        self.check_record_company_label.move(0,move_y)

        self.company_scroll.setFixedSize(self.line_width,220)
        self.company_scroll.move(0,move_y + self.label_height - 15)
        self.company_scroll.setWidgetResizable(True)
        self.company_scroll.setStyleSheet(
            "QScrollArea{"
            "    border: 2px solid black; color: black;"
            "}"
        )   
        company_scroll_layout = QVBoxLayout()
        self.company_scroll_widget.setLayout(company_scroll_layout)
        self.company_data()

        # 连接按钮的点击事件
        self.prev_button.clicked.connect(self.prev_page)
        self.next_button.clicked.connect(self.next_page)


        # 显示第一页数据
        self.current_page = 0
        # self.update_table_data()

        
        self.update_record_button.resize(160,70)
        self.update_record_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_record_button.move(screen.width() - 320,screen.height() - 550)
        self.update_record_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_record_button.clicked.connect(self.update_record_clicked)

        self.search_button.resize(160,70)
        self.search_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.search_button.move(screen.width() - 750,90)
        self.search_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.search_button.clicked.connect(self.search_button_clicked)

        self.refresh_button.resize(180,70)
        self.refresh_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.refresh_button.move(screen.width() - 960,90)
        self.refresh_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.refresh_button.clicked.connect(self.refresh_button_clicked)

        self.search_label.resize(self.label_width,self.label_height)
        self.search_label.setFont(self.s_font)
        self.search_label.move(1110,95)   

        self.search_box.resize(self.line_width,self.line_height)
        self.search_box.setFont(self.s_font)
        self.search_box.move(1240,110)
        self.search_box.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }")


        self.export_button.resize(180,70)
        self.export_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_button.move(screen.width() - 325,screen.height() - 300)
        self.export_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_button.clicked.connect(self.export_to_excel)

        self.delete_record_button.resize(160,70)
        self.delete_record_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_record_button.move(screen.width() - 320,screen.height() - 450)
        self.delete_record_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.delete_record_button.clicked.connect(self.delete_record_clicked)

        
        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        self.amount_label.move(screen.width() - 350,700)   

        self.amount_line.resize(self.label_width+ 100,self.label_height)
        self.amount_line.setFont(self.l_font)
        self.amount_line.move(screen.width() - 350,750)   

    def refresh_button_clicked(self):
        self.query_str = "ORDER BY id DESC "
        self.current_page = 0
        if self.record_date_button.isVisible() == False:
            self.show_record_date()
        self.update_model()

    def search_button_clicked(self):
        str = "WHERE "
        temp = ""
        condition_num = 0

        # month filter
        month = self.data_month.currentIndex()
        if month > 0:
            if month < 10:
                temp = f"date like '%/0{month}/%'"
            else:
                temp = f"date like '%/{month}/%'"

            str += temp
            condition_num += 1

        if condition_num > 0:
            str += " AND "
        
        # year filter
        year = self.data_year.value()
        temp = f"date like '%/%/{year}'"
        str += temp 
        condition_num += 1

        # lorry filter
        checked_boxes = []
        for checkbox in self.lorry_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())

        if len(checked_boxes) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "
            
            temp += " ( "
            first = True
            for lorry in checked_boxes:
                if not first:
                    temp += " OR "

                first = False
                temp += f"lorry = '{lorry}'"
            
            temp += " ) "
            str += temp 
            condition_num += 1
        
        # company filter
        checked_boxes = []
        for checkbox in self.company_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())

        if len(checked_boxes) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "
            
            temp += " ( "
            first = True
            for company in checked_boxes:
                if not first:
                    temp += " OR "

                first = False
                temp += f"company = '{company}'"
                
                
            temp += " ) "
            str += temp 

            condition_num += 1

        # keyword search
        keyword = self.search_box.text()
        if len(keyword) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "

            temp += f"( LOWER(lorry) LIKE LOWER('%{keyword}%') OR LOWER(reference) LIKE LOWER('%{keyword}%') OR LOWER(company) LIKE LOWER('%{keyword}%') OR LOWER(date) LIKE LOWER('%{keyword}%'))"

            str += temp 
            condition_num += 1

        #record_date search
        if self.record_date_line.isVisible() == True:
            temp=""
            if condition_num > 0:
                str += " AND "

            date = self.record_date_line.date().toString("dd/MM/yyyy")
            temp = f"created_at like '{date}'"
            str += temp 
            condition_num += 1

        # start date & end date
        start_date = self.start_date_line.currentIndex()

        if start_date > 0:
            if condition_num > 0:
                str += " AND "
            temp = f"(date - DATE('0000-01-02'))  >= {start_date}"
            str += temp
            condition_num += 1
        #WHERE DATEPART(DAY, your_date_column) > 15;

        end_date = self.end_date_line.currentIndex()
        if end_date > 0:
            if condition_num > 0:
                str += " AND "
            temp = f"(date - DATE('0000-01-02'))  <= {end_date}"
            str += temp
            condition_num += 1

        if len(str) == 0:
            self.query_str = ""
        else:
            self.query_str = str + " "

        if len(str) == 0:
            self.query_str = ""
        else:
            self.query_str = str + " "

        
        self.current_page = 0
        self.update_model()
        
    def update_record_clicked(self):
        row_data = self.get_table_record()

        if row_data != -1:
            # 实例化 UpdateRecordDialog，并显示弹窗
            update_dialog = UpdateRecordDialog(row_data,self)
            update_dialog.exec_()

            # self.current_page = 0
            self.update_model()
            
    def delete_record_clicked(self):
        row_data = self.get_table_record()
        if row_data != -1:
            query = QSqlQuery() 
            query.prepare(f"DELETE FROM Record WHERE id = {row_data[0]}")
            

            if query.exec():
                print("Data daleted successfully.")
            else:
                print("Error while deleting data:", query.lastError().text())

            self.show_message_box("Delete Record","Record Deleted")
            # self.current_page = 0
            self.update_model()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_model()
        self.page_number.setText(str(self.current_page+1))

    def next_page(self):
        offset = (self.current_page + 1) * ROWS_PER_PAGE
        query = QSqlQuery()
        # query_str = f"SELECT * FROM Record LIMIT 1 OFFSET {offset}"
        query_str = "SELECT * FROM Record " + self.query_str + f"LIMIT 1 OFFSET {offset}"
        query.exec(query_str)
        if query.next():
            self.model.setQuery(query_str)
            self.current_page += 1
            self.update_model()
            self.page_number.setText(str(self.current_page+1))

    def get_table_record(self):
        # 获取选中模型
        selection_model = self.data_table.selectionModel()
        
        # 获取选中的行索引
        selected_indexes = selection_model.selectedRows()
     
        if not selected_indexes:
            # 如果没有选中的行，弹出消息框提示
            self.show_message_box("Warning","Please select a row.")
            return -1

        # 获取第一个选中的行索引
        row = selected_indexes[0].row()

        row_data = [self.model.data(self.model.index(row, column), Qt.DisplayRole) for column in range(self.model.columnCount())]

        return row_data

    def export_subfreige(self,export_dialog):
        filename = export_dialog.get_file()
        company = export_dialog.get_company()

        model = self.data_table.model()
        if not model:
            return

        workbook = Workbook()
        sheet = workbook.active

        page_margins = PageMargins(
            left=0.4,  # 左边距
            right=0.4,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距
        )
        sheet.page_margins = page_margins

        font = Font(name='Times New Roman', size=11, bold=False, color='000000')
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        sheet.merge_cells('A1:G4')
        merged_cell = sheet['A1']
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = font
        merged_cell.border = border

        # 在合并的单元格中写入文本
        merged_cell.value = company

        empty_row = []
        for i in range(3):
            sheet.append(empty_row)
        
        headers = ["Date","Lorry no","Ref. no","Description","Quantity(Ton)","Unit Price","Amount (RM)"]    
        sheet.append(headers)

        current_row = 5

        if self.query_str != "ORDER BY id DESC ":
            text = "SELECT * FROM Record " + self.query_str + " ORDER BY date ASC"
        else:
            text = "SELECT * FROM Record ORDER BY date ASC"

        query = QSqlQuery()
        query.exec_(text)    

        while query.next():
            current_row += 1
            row_data = []
            value = None
            for i in range(1,query.record().count()):
                value = query.value(i)
                if i in [6,7,8] and isinstance(value, (float,int)):
                    if is_3_decimal_places(value):
                        value =  "{:.3f}".format(query.value(i))  # Display numeric data with two decimal places
                    else:
                        value = "{:.2f}".format(query.value(i))
                if i in [6,7,8] and isinstance(value, str) and is_numeric_string(value) and not contain_comma(str(value)):
                    value = float(query.value(i))
                    if is_3_decimal_places(value):
                        value =  "{:.3f}".format(value)  # Display numeric data with two decimal places
                    else:
                        value =  "{:.2f}".format(value)
                if i == 5 or i == 9:
                    continue    

                row_data.append(value)
            
            sheet.append(row_data)
        
        sheet.merge_cells('A1:G4')
        
        # 合并下一行的单元格并居中
        next_row = current_row + 1

        merge_range = f"A{next_row}:F{next_row}"
        sheet.merge_cells(merge_range)
        merged_cell = sheet[f"A{next_row}"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.value = "Total"

        sheet[f"G{next_row}"] = self.current_sum

        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
            for cell in row:
                cell.font = font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2   # 加上一些额外宽度
            column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
            if column_letter == 'A':
                adjusted_width = 12
            sheet.column_dimensions[column_letter].width = adjusted_width


        workbook.save(filename)

    def export_invoice(self,export_dialog):
        filename = export_dialog.get_file()
        invoice_info = export_dialog.get_invoice_info()
        debtor_company = export_dialog.get_company()

        model = self.data_table.model()
        if not model:
            return
    

        workbook = Workbook()
        sheet = workbook.active

        page_margins = PageMargins(
            left=0.4,  # 左边距
            right=0.4,  # 右边距
            top=0.5,  # 上边距
            bottom=0.5,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距
        )
        sheet.page_margins = page_margins

        normal_font = Font(name='Times New Roman', size=10, bold=False, color='000000')
        normal_bold_font = Font(name='Times New Roman', size=10, bold=True, color='000000')
        title_bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
        invoice_underline_font = Font(name='Times New Roman', size=11,underline="single", bold=True, color='000000')
        
        top_border = Border(top=Side(style='thin'))
        bottom_border = Border(bottom=Side(style='thin'))
        bottom_right_border = Border(bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
        top_right_border = Border(top=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
        left_right_border = Border(left=Side(style='thin'),right=Side(style='thin'))
        dotted_border = Border(bottom=Side(border_style='dotted', color='000000'))

        title_text = ["BOEY LAI PIN ENTERPRISE","( IP0280569-U )","Lot 196831, Selasar Rokam 20, Kg. Sri Ampang,"
                    ,"31350 Ipoh, Perak Darul Ridzuan.","Tel : 012-5757988","Email: boeylaipin@outlook.com","Service Tax ID No. A10-2403-32000006"]
        
        current_row = 1
        for text in title_text:
            sheet.merge_cells(f'A{current_row}:G{current_row}')
            merged_cell = sheet[f'A{current_row}']
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            merged_cell.value = text
            if current_row == 1:
                merged_cell.font = title_bold_font
            else:
                merged_cell.font = normal_bold_font
            current_row += 1

        current_row += 1
        sheet.merge_cells(f'A{current_row}:G{current_row}')
        merged_cell = sheet[f'A{current_row}']
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.value = "INVOICE"
        merged_cell.font = invoice_underline_font
        current_row += 1

        sheet[f'F{current_row}'] = "INVOICE NO.  :"
        sheet[f'F{current_row}'].font = normal_font
        sheet[f'G{current_row}'] = invoice_info['invoice']
        sheet[f'G{current_row}'].font = normal_font

        current_row += 1
        sheet.merge_cells(f'F{current_row}:G{current_row}')
        merged_cell = sheet[f'F{current_row}']
        merged_cell.alignment = Alignment(horizontal='left', vertical='center')
        merged_cell.value = "DATE  :"
        merged_cell.font = normal_font

        current_row += 1
        sheet[f'A{current_row}'] = "BILL TO :"
        sheet[f'A{current_row}'].font = normal_bold_font
        sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

        company_data = []
        company_query = QSqlQuery()

        if not company_query.exec(f"SELECT * FROM Company where name = '{debtor_company}'"):
            print("Error executing query:", company_query.lastError().text())

        if company_query.next():
            for i in range(1,8):
                company_data.append(company_query.value(i))


        for i in range(7):
            if i< 6 and len(company_data[i]) == 0:
                continue
                
            sheet.merge_cells(f'B{current_row}:G{current_row}')
            merged_cell = sheet[f'B{current_row}']
            merged_cell.alignment = Alignment(horizontal='left', vertical='center')
            if i == 0:
                merged_cell.font = normal_bold_font
            else:
                merged_cell.font = normal_font
            current_row += 1

            if i == 6:
                merged_cell.value = "TEL: " + company_data[i].upper()
            else:
                merged_cell.value = company_data[i].upper()
    
            
        for cell in sheet[current_row]:
            cell.border = top_right_border
            cell.font = normal_bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[current_row+1]:
            cell.border = bottom_right_border
            cell.font = normal_bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # DATE
        sheet.merge_cells(f'A{current_row}:A{current_row + 1}')
        merged_cell = sheet[f'A{current_row}']
        merged_cell.value = "DATE"

        # LORRY NO
        sheet[f'B{current_row}'] = "LORRY"
        sheet[f'B{current_row + 1}'] = "NO."

        # REFERENCE
        sheet[f'C{current_row}'] = "REFERENCE"
        sheet[f'C{current_row + 1}'] = "NO."

        # DESCRIPTION
        sheet.merge_cells(f'D{current_row}:D{current_row + 1}')
        merged_cell = sheet[f'D{current_row}']
        merged_cell.value = "DESCRIPTION"

        # QUANTITY
        sheet[f'E{current_row}'] = "QUANTITY"
        sheet[f'E{current_row + 1}'] = "(TON)"

        # UNIT PRICE
        sheet[f'F{current_row}'] = "UNIT PRICE"
        sheet[f'F{current_row + 1}'] = "(RM)"

        # AMOUNT
        sheet[f'G{current_row}'] = "AMOUNT"
        sheet[f'G{current_row + 1}'] = "(RM)"

        current_row += 2

        if self.query_str != "ORDER BY id DESC ":
            text = "SELECT * FROM Record " + self.query_str + " ORDER BY date ASC"
        else:
            text = "SELECT * FROM Record ORDER BY date ASC"

        query = QSqlQuery()
        query.exec_(text)    

        start_data_row = current_row


        while query.next():
            row_data = []
            value = None
            for i in range(1,query.record().count()-1):
                value = query.value(i)
                if i in [6,7,8] and isinstance(value, (float,int)):
                    if is_3_decimal_places(value):
                        value =  locale.format_string("%.3f", value, grouping=True)
                    else:
                        value = locale.format_string("%.2f", value, grouping=True)

                    
                if i in [6,7,8] and isinstance(value, str) and is_numeric_string(value) and not contain_comma(str(value)):
                    value = float(query.value(i))
                    if is_3_decimal_places(value):
                        value =  locale.format_string("%.3f", value, grouping=True) 
                    else:
                        value =  locale.format_string("%.2f", value, grouping=True)

                if i == 8:
                    value += "    "
                
                if i == 9:
                    continue

                if i != 4:
                    row_data.append(value)

            print(row_data)
            sheet.append(row_data)
            
            current_row += 1


        empty_row = []
        while current_row <= 36:
            sheet.append(empty_row)
            current_row += 1

        for row in sheet.iter_rows(min_row=start_data_row, max_row=current_row-1):
            i = 0
            for cell in row:
                i += 1
                cell.border = left_right_border
                cell.font = normal_font
                if i == 7:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        # BANK A/C
        sheet.merge_cells(f'A{current_row}:B{current_row}')
        merged_cell = sheet[f'A{current_row}']
        merged_cell.value = " BANK A/C NO. : "
        merged_cell.font = normal_bold_font
        merged_cell.alignment = Alignment(horizontal='left', vertical='center')

        # AMOUNT
        sheet.merge_cells(f'D{current_row}:F{current_row}')
        merged_cell = sheet[f'D{current_row}']
        merged_cell.value = " SUBTOTAL (EXLCUDING TAX)      :"
        merged_cell.font = normal_font
        merged_cell.alignment = Alignment(horizontal='right', vertical='center')

        # sheet[f'F{current_row}'] = "AMOUNT (EXLCUDING TAX): "
        # sheet[f'F{current_row}'].font = normal_bold_font

        # SUM
        sheet[f'G{current_row}'] = self.current_sum_str + "    "
        sheet[f'G{current_row}'].font = normal_font
        sheet[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='center')

        for cell in sheet[current_row]:
            cell.border = top_border
        
        current_row += 1
        
        # for bank acount
        sheet.merge_cells(f'A{current_row}:C{current_row}')
        merged_cell = sheet[f'A{current_row}']
        merged_cell.value = " PBB   3985876021"
        merged_cell.font = normal_bold_font
        merged_cell.alignment = Alignment(horizontal='left', vertical='center')

        sheet.merge_cells(f'D{current_row}:F{current_row}')
        merged_cell = sheet[f'D{current_row}']
        merged_cell.value = " SST AMOUNT @ 6%      :"
        merged_cell.font = normal_font
        merged_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        sst = self.current_sum * 0.06

        if is_3_decimal_places(sst):
            sst_str = locale.format_string("%.3f", sst, grouping=True)
        else:
            sst_str = locale.format_string("%.2f", sst, grouping=True)

        print(f"sst = {sst}")
        sheet[f'G{current_row}'] = sst_str + "    "
        sheet[f'G{current_row}'].font = normal_font
        sheet[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='center')

        current_row += 1

        print("third row")
        sheet.merge_cells(f'D{current_row}:F{current_row}')
        merged_cell = sheet[f'D{current_row}']
        merged_cell.value = " TOTAL      :"
        merged_cell.font = normal_bold_font
        merged_cell.alignment = Alignment(horizontal='right', vertical='center')

        total = self.current_sum + sst
        if is_3_decimal_places(sst):
            total_str = locale.format_string("%.3f",total, grouping=True)
        else:
            total_str = locale.format_string("%.2f", total, grouping=True)

        sheet[f'G{current_row}'] = total_str + "    "
        sheet[f'G{current_row}'].font = normal_font
        sheet[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='center')



        for cell in sheet[current_row]:
            cell.border = bottom_border
            cell.font = normal_font
            
        current_row += 1
        sheet.merge_cells(f'A{current_row}:G{current_row}')
        merged_cell = sheet[f'A{current_row}']
        merged_cell.value = " RINGGIT MALAYSIA : "
        merged_cell.font = normal_font
        merged_cell.alignment = Alignment(horizontal='left', vertical='center')

        current_row += 2
        sheet.merge_cells(f'A{current_row}:G{current_row}')
        merged_cell = sheet[f'A{current_row}']
        merged_cell.value = "NOTE : Goods conveying by our lorries are not responsible for any loss or damage, unavoidable delays due to hold-ups robbery, "

        sheet.merge_cells(f'A{current_row+1}:G{current_row+1}')
        merged_cell = sheet[f'A{current_row+1}']
        merged_cell.value = "accident, hijacking and any unforeseen circumstances. Consignors are requested to notify us within 7 days if there is any complaint, "

        sheet.merge_cells(f'A{current_row+2}:G{current_row+2}')
        merged_cell = sheet[f'A{current_row+2}']
        merged_cell.value = "otherwise we take no responsibility."

        for row in sheet.iter_rows(min_row=current_row, max_row=current_row + 3):
            for cell in row:
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='left', vertical='center')

        current_row += 6

        sheet[f'A{current_row}'].border = dotted_border
        
        current_row += 1
        sheet[f'A{current_row}'] = "ISSUED BY"
        sheet[f'A{current_row}'].font = normal_font

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 17
        sheet.column_dimensions['D'].width = 23
        sheet.column_dimensions['E'].width = 11.4
        sheet.column_dimensions['F'].width = 13.3
        sheet.column_dimensions['G'].width = 11.4

        workbook.save(filename)


        #Save invoice
        query = QSqlQuery()
        
        query.exec_(f"DELETE FROM Invoice WHERE no={invoice_info['invoice']}")
        query = QSqlQuery()
        query.prepare("INSERT INTO Invoice(no,month,year,company,amount) VALUES (:no,:month,:year,:company,:amount)")
        query.bindValue(":no", invoice_info['invoice'])
        query.bindValue(":month", invoice_info['month'])
        query.bindValue(":year", invoice_info['year'])
        query.bindValue(":company", invoice_info['company'])
        query.bindValue(":amount", self.current_sum)

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

    def export_to_excel(self):
        export_dialog = ExportRecordDialog(self)
        export_dialog.exec_()

        if export_dialog.get_exported() < 0:
            return

        if export_dialog.get_exported() == 0:
            self.export_subfreige(export_dialog)

        if export_dialog.get_exported() == 1:
            print("here")
            self.export_invoice(export_dialog)
            print("here")
            self.show_message_box("Export","Data Exported")
 
class check_invoice_page(BasePage):
    current_sum = 0
    query_str = ""
    def __init__(self):
        super().__init__('ui\\check_invoice_page.ui')
        self.check_invoice_button.setStyleSheet("""
                border-radius:0px;
                border:0px;
                background-color:white;
                """)

        self.init_ui()
        self.setup_database()
        self.current_page = 0
        self.query_str = "ORDER BY no DESC "
        self.update_model()

        self.add_record_button.clicked.connect(lambda: main_window.switch_page(1))
        self.check_record_button.clicked.connect(lambda: main_window.switch_page(2))
        self.add_expenses_button.clicked.connect(lambda: main_window.switch_page(3))
        self.check_expenses_button.clicked.connect(lambda: main_window.switch_page(4))
        self.add_salary_button.clicked.connect(lambda: main_window.switch_page(7))
        self.check_salary_button.clicked.connect(lambda: main_window.switch_page(8))

    def setup_database(self):
        self.model = InvoiceQueryModel()
        self.data_table.setModel(self.model)
        self.proxy_model = HideFirstColumnProxyModel(self)
        self.proxy_model.setSourceModel(self.model)
        self.data_table.setModel(self.proxy_model)

    def update_model(self):
        if len(self.query_str) == 0:
            self.query_str = "ORDER BY no DESC "

        offset = self.current_page * ROWS_PER_PAGE
    
        query = "SELECT * FROM Invoice " + self.query_str + f"LIMIT {ROWS_PER_PAGE} OFFSET {offset}"
        self.model.setQuery(query)
        self.page_number.setText(str(self.current_page+1))

        for i in range(6):
            if i == 1:
                self.data_table.setColumnWidth(i, 600)
            else:
                self.data_table.setColumnWidth(i, 300)
    
    def clearCheckboxes(self,scroll_widget):
        scroll_layout = scroll_widget.layout()
        for i in reversed(range(scroll_layout.count())):
            item = scroll_layout.itemAt(i)
            if isinstance(item.widget(), QCheckBox):
                item.widget().deleteLater()

    def company_data(self):
        self.clearCheckboxes(self.company_scroll_widget)
        self.company_checkboxes = []
        scroll_layout = self.company_scroll_widget.layout()

        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            self.company_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中

    def init_ui(self):
        screen = QDesktopWidget().screenGeometry()

        self.check_history_label.resize(self.label_width + 50,self.label_height)
        self.check_history_label.setFont(self.l_font)
        self.check_history_label.move(0,0)

        self.data_frame.resize(screen.width() - 700,screen.height() - 230)
        self.data_frame.move(110,50)

        self.data_month.resize(self.line_width,self.line_height)
        self.data_month.move(0,self.label_height)
        self.data_month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.data_month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.data_month.addItem("Not Chosen")
        self.data_month.addItems(months)
        self.data_month.setCurrentIndex(0)

        self.data_year.resize(self.line_width,self.line_height)
        self.data_year.move(self.line_width + 20 , self.label_height)
        self.data_year.setMinimum(2022)
        self.data_year.setMaximum(2030)
        self.data_year.setValue(2023)
        self.data_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.data_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

        self.data_table.resize(self.data_frame.width(),self.data_frame.height()-220)
        self.data_table.move(0,self.label_height + 60)
        self.data_table.setFont(QFont("Times New Roman",13))
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.data_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        delegate = CenteredNumberDelegate()
        self.data_table.setItemDelegate(delegate)
        self.data_table.setSelectionBehavior(QTableView.SelectRows)
        self.data_table.hideColumn(0)
        self.data_table.setStyleSheet(
            "QTableView::item:selected{ background-color:black;color:white;}"
            "QTableView { color: black; gridline-color: black; border: 2px solid black; background-color:white ;gridline-width: 2px;}"
            "QHeaderView { background-color: rgba(216,216,216,200); color:black ;font-family: Rockwell ; font-size: 27px; }"

            "QHeaderView::section { background-color: rgba(216,216,216,200) ;color: black;height: 50px;}"
        )
        
        
        self.prev_button.resize(100,50)
        self.prev_button.move(self.data_table.width()//2 - 150 ,self.data_frame.height()-50)
        self.prev_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.prev_button.setStyleSheet("QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
                                       "QPushButton:hover{ border: 3px solid black }")
        
        self.page_number.resize(self.label_width,self.label_height)
        self.page_number.move(self.data_table.width()//2,self.data_frame.height()-65 )
        self.page_number.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        
        self.next_button.resize(100,50)
        self.next_button.move(self.data_frame.width()//2 + 60,self.data_frame.height()-50)
        self.next_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.next_button.setStyleSheet(
            "QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
            "QPushButton:hover{ border: 3px solid black }"
            )
        
        self.filter_frame.resize(500,1000)
        self.filter_frame.move(screen.width()-500,100)
        
        self.check_record_company_label.resize(self.label_width,self.label_height)
        self.check_record_company_label.setFont(self.s_font)
        self.check_record_company_label.move(0,0) 
        move_y = self.label_height - 15

        self.company_scroll.setFixedSize(self.line_width + 100,220)
        self.company_scroll.move(0,move_y)
        self.company_scroll.setWidgetResizable(True)
        self.company_scroll.setStyleSheet(
            "QScrollArea{"
            "    border: 2px solid black; color: black;"
            "}"
        )   
        scroll_layout = QVBoxLayout()
        self.company_scroll_widget.setLayout(scroll_layout)
        self.company_data()
    
        # 连接按钮的点击事件
        self.prev_button.clicked.connect(self.prev_page)
        self.next_button.clicked.connect(self.next_page)

        self.add_button.resize(140,70)
        self.add_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_button.move(screen.width() -395,screen.height() - 660)
        self.add_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.add_button.clicked.connect(self.add_button_clicked)

        self.update_record_button.resize(160,70)
        self.update_record_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_record_button.move(screen.width() - 400,screen.height() - 550)
        self.update_record_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_record_button.clicked.connect(self.update_record_clicked)

        self.search_button.resize(160,70)
        self.search_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.search_button.move(screen.width() - 750,90)
        self.search_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.search_button.clicked.connect(self.search_button_clicked)

        self.refresh_button.resize(180,70)
        self.refresh_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.refresh_button.move(screen.width() - 960,90)
        self.refresh_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.refresh_button.clicked.connect(self.refresh_button_clicked)

        self.search_label.resize(self.label_width,self.label_height)
        self.search_label.setFont(self.s_font)
        self.search_label.move(970,95)   

        self.search_box.resize(self.line_width,self.line_height)
        self.search_box.setFont(self.s_font)
        self.search_box.move(1100,110)
        self.search_box.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }")

        self.export_button.resize(320,70)
        self.export_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_button.move(screen.width() - 470,screen.height() - 300)
        self.export_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_button.clicked.connect(self.export_button_clicked)

        self.delete_record_button.resize(160,70)
        self.delete_record_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_record_button.move(screen.width() - 400,screen.height() - 450)
        self.delete_record_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.delete_record_button.clicked.connect(self.delete_record_clicked)

    def refresh_button_clicked(self):
        self.query_str = "ORDER BY id DESC "
        self.current_page = 0
        self.update_model()

    def search_button_clicked(self):
        str = "WHERE "
        temp = ""
        condition_num = 0

        # month filter
        month = self.data_month.currentIndex()
        if month > 0:
            month = self.data_month.currentText()
            temp = f"month = '{month}'"
            
            str += temp
            condition_num += 1

        if condition_num > 0:
            str += " AND "
        
        # year filter
        year = self.data_year.value()
        temp = f"year = '{year}'"
        str += temp 
        condition_num += 1

        # company filter
        checked_boxes = []
        for checkbox in self.company_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())

        if len(checked_boxes) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "
            
            temp += " ( "
            first = True
            for company in checked_boxes:
                if not first:
                    temp += " OR "

                first = False
                temp += f"company = '{company}'"
                
                
            temp += " ) "
            str += temp 

            condition_num += 1

        # keyword search
        keyword = self.search_box.text()
        if len(keyword) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "

            temp += f"( LOWER(no) LIKE LOWER('%{keyword}%') OR LOWER(company) LIKE LOWER('%{keyword}%'))"

            str += temp 
            condition_num += 1


        if len(str) == 0:
            self.query_str = ""
        else:
            self.query_str = str + " "

        self.current_page = 0
        self.update_model()

    def update_record_clicked(self):
        row_data = self.get_table_record()

        if row_data != -1:
            # 实例化 UpdateRecordDialog，并显示弹窗
            update_dialog = UpdateInvoiceDialog(row_data,self)
            update_dialog.exec_()

            # self.current_page = 0
            self.update_model()
            
    def delete_record_clicked(self):
        row_data = self.get_table_record()
        if row_data != -1:
            query = QSqlQuery() 
            print("Selected Row Data:", row_data)
            query.prepare(f"DELETE FROM Invoice WHERE id = {row_data[0]}")
            
            if query.exec():
                print("Data daleted successfully.")
            else:
                print("Error while deleting data:", query.lastError().text())

            self.show_message_box("Delete Invoice","Invoice Deleted")
            # self.current_page = 0
            self.update_model()

    def add_button_clicked(self):
        add_dialog = AddInvoiceDialog(self)
        add_dialog.exec_()
        
    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_model()
        self.page_number.setText(str(self.current_page+1))

    def next_page(self):
        offset = (self.current_page + 1) * ROWS_PER_PAGE
        query = QSqlQuery()
        # query_str = f"SELECT * FROM Invoice LIMIT 1 OFFSET {offset}"
        query_str = "SELECT * FROM Invoice " + self.query_str + f"LIMIT 1 OFFSET {offset}"
        query.exec(query_str)
        if query.next():
            self.model.setQuery(query_str)
            self.current_page += 1
            self.update_model()
            self.page_number.setText(str(self.current_page+1))

    def get_table_record(self):
        # 获取选中模型
        selection_model = self.data_table.selectionModel()
        
        # 获取选中的行索引
        selected_indexes = selection_model.selectedRows()
     
        if not selected_indexes:
            # 如果没有选中的行，弹出消息框提示
            self.show_message_box("Warning","Please select a row.")
            return -1

        # 获取第一个选中的行索引
        row = selected_indexes[0].row()

        row_data = [self.model.data(self.model.index(row, column), Qt.DisplayRole) for column in range(self.model.columnCount())]

        return row_data


    def export_button_clicked(self):
        export_dialog = ExportDebtorDialog(self)
        export_dialog.exec_()

class add_expenses_page(BasePage):
    def __init__(self):
        super().__init__('ui\\add_expenses_page.ui')

        self.init_ui()
        self.setup_expenses_frame()
        self.setup_company_frame()

        self.add_expenses_button.setStyleSheet("""
                border-radius:0px;
                border:0px;
                background-color:white;
                """)
        self.add_record_button.clicked.connect(lambda: main_window.switch_page(1))
        self.check_record_button.clicked.connect(lambda: main_window.switch_page(2))
        self.check_expenses_button.clicked.connect(lambda: main_window.switch_page(4))
        self.check_invoice_button.clicked.connect(lambda: main_window.switch_page(5))
        self.add_salary_button.clicked.connect(lambda: main_window.switch_page(7))
        self.check_salary_button.clicked.connect(lambda: main_window.switch_page(8))

    def init_ui(self):
        self.workspace_frame.resize(600,1200)
        self.workspace_frame.move(600,110)
        
        self.add_expenses_label.resize(self.label_width + 200,self.label_height) 
        self.add_expenses_label.setFont(self.l_font)
        self.add_expenses_label.move(0,0)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(0,80)
        move_y = 80

        self.date_line.setDate(QDate.currentDate())
        self.date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.date_line.setDisplayFormat("dd/MM/yyyy")
        self.date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.date_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.date_line.move(0,move_y)
        self.date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:10px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 28px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white }"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:20px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 35px; height: 35px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 3px solid black; font-family: Constantia; font-size: 28px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:50px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:50px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_left.png); }"
        )
    
        self.lorry_label.resize(self.label_width,self.label_height)
        self.lorry_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.lorry_label.move(0,move_y)        

        self.lorry_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.lorry_line.move(0,move_y)
        self.setup_combo_box(self.lorry_line)
        self.lorry_data(self.lorry_line)

        self.reference_label.resize(self.label_width,self.label_height)
        self.reference_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.reference_label.move(0,move_y)

        self.reference_line.resize(self.line_width,self.line_height)
        self.reference_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.reference_line.move(0,move_y)
        self.reference_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        
        self.expenses_label.resize(self.label_width,self.label_height)
        self.expenses_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.expenses_label.move(0,move_y)

        self.expenses_line.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.expenses_line.move(0,move_y)
        self.setup_combo_box(self.expenses_line)
        self.expenses_data(self.expenses_line)

        self.company_label.resize(self.label_width,self.label_height)
        self.company_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.company_label.move(0,move_y)

        self.company_line.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.company_line.move(0,move_y)
        self.setup_combo_box(self.company_line)
        self.expenses_company_data(self.company_line)

        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        self.submit_lorry_expenses.resize(330,70)
        self.submit_lorry_expenses.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.submit_lorry_expenses.move(10,move_y + 110)
        self.submit_lorry_expenses.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.submit_lorry_expenses.clicked.connect(self.submit_lorry_expenses_clicked)

        self.submit_creditor_expenses.resize(370,70)
        self.submit_creditor_expenses.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.submit_creditor_expenses.move(0,move_y + 210)
        self.submit_creditor_expenses.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.submit_creditor_expenses.clicked.connect(self.submit_creditor_expenses_clicked)
        
       
    def setup_company_frame(self):
        self.add_company_frame.resize(850,600)
        screen = QDesktopWidget().screenGeometry()
        self.add_company_frame.move(screen.width() - 1200,600)

        font = QFont("Tw Cen MT Condensed Extra Bold",20)
        font.setUnderline(True)
        self.add_company_label.resize(self.label_width + 80,self.label_height)
        self.add_company_label.setFont(font)
        self.add_company_label.move(0,0)

        self.company_label_add.resize(self.label_width + 80,self.label_height)
        self.company_label_add.setFont(self.s_font)
        self.company_label_add.move(0,80)

        self.company_label_add_2.resize(self.label_width + 80,self.label_height)
        self.company_label_add_2.setFont(self.s_font)
        self.company_label_add_2.move(self.label_width + 230,80)

        move_y = 80

        self.company_line_add.resize(self.line_width+80,self.line_height)
        self.company_line_add.setFont(self.s_font) 
        move_y = move_y + self.label_height - 15
        self.company_line_add.move(0,move_y)
        self.company_line_add.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        
        self.add_company_submit_button.resize(150,70)
        self.add_company_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_company_submit_button.move(100,move_y + 110)
        self.add_company_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.add_company_submit_button.clicked.connect(self.company_add_button_clicked)
         
        self.setup_combo_box(self.delete_company_combobox)
        self.delete_company_combobox.resize(self.line_width + 80,self.line_height)
        self.delete_company_combobox.move(self.line_width + 130,move_y)

        self.delete_company_button.resize(160,70)
        self.delete_company_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_company_button.move(self.line_width + 250,move_y + 110)
        self.delete_company_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.expenses_company_data(self.delete_company_combobox)
        self.delete_company_button.clicked.connect(self.delete_company_clicked)

    def expenses_company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Expenses_Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)

    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def expenses_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT description FROM Expenses")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
  
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
    
    def submit_lorry_expenses_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return

        query = QSqlQuery()
        query.prepare("INSERT INTO Expenses_Record(date,lorry,reference,expenses,company,amount) VALUES (:date,:lorry,:reference,:expenses,:company,:amount)")
        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":lorry", set_not_chosen(self.lorry_line.currentText()))
        query.bindValue(":reference", "-")
        query.bindValue(":expenses", set_not_chosen(self.expenses_line.currentText()))
        query.bindValue(":company", "-")
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Submit Lorry Expenses","New Lorry Expenses Added")

    def submit_creditor_expenses_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
            
        
        query = QSqlQuery()
        query.prepare("INSERT INTO Expenses_Record(date,lorry,reference,expenses,company,amount) VALUES (:date,:lorry,:reference,:expenses,:company,:amount)")
        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":lorry", set_not_chosen(self.lorry_line.currentText()))
        query.bindValue(":reference", self.reference_line.text())
        query.bindValue(":expenses", set_not_chosen(self.expenses_line.currentText()))
        query.bindValue(":company", set_not_chosen(self.company_line.currentText()))
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Submit Creditor Expenses Record","New Creditor Expenses Added")

    def expenses_add_button_clicked(self):
        text = self.new_expenses_line.text()
        
        if len(text) == 0:
            return 
        
        query = QSqlQuery(f"SELECT * FROM Expenses WHERE description = '{text}'")
        if query.next():
            print("Value already exists in the database.")
            return

        query = QSqlQuery()
        query.prepare("INSERT INTO Expenses(description) VALUES (:description)")
        query.bindValue(":description", text)

        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.expenses_data(self.expenses_line)
        self.expenses_data(self.expenses_combobox)

        self.show_message_box(f"Add Expenses Type",f"Save Expenses Type: {text}")

    def company_add_button_clicked(self):
        text = self.company_line_add.text()

        if len(text) == 0:
            return

        query = QSqlQuery(f"SELECT * FROM Expenses Company WHERE name = '{text}'")
        if query.next():
            print("Value already exists in the database.")
            return
    
        query = QSqlQuery()
        query.prepare("INSERT INTO Expenses_Company(name) VALUES (:name)")
        query.bindValue(":name", text)

        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.expenses_company_data(self.company_line)
        self.expenses_company_data(self.delete_company_combobox)  

        self.show_message_box("Add Expenses Company",f"Save Company: {text}")

    def delete_expenses_clicked(self):
        type = self.expenses_combobox.currentText()

        query = QSqlQuery(f"DELETE FROM Expenses WHERE description = '{type}'")

        if query.exec():
            print("Data deleted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box(f"Delete Expenses Type",f"Expenses Type: {type} Deleted")
        self.expenses_data(self.expenses_line)
        self.expenses_data(self.expenses_combobox)

    def delete_company_clicked(self):
        company = self.delete_company_combobox.currentText()

        query = QSqlQuery(f"DELETE FROM Expenses_Company WHERE name = '{company}'")

        if query.exec():
            print("Data deleted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Delete Expenses Company",f"Expenses Company {company} Deleted")
        self.expenses_company_data(self.company_line)
        self.expenses_company_data(self.delete_company_combobox)  
        
    def setup_expenses_frame(self):
        screen = QDesktopWidget().screenGeometry()
        self.add_expenses_frame.resize(800,600)
        self.add_expenses_frame.move(screen.width() - 1200,210)

        font = QFont("Tw Cen MT Condensed Extra Bold",20)
        font.setUnderline(True)
        self.expenses_type_label.resize(self.label_width + 100,self.label_height)
        self.expenses_type_label.setFont(font)
        self.expenses_type_label.move(0,0)

        self.new_expenses_name.resize(self.label_width + 50,self.label_height)
        self.new_expenses_name.setFont(self.s_font)
        self.new_expenses_name.move(0,80)

        self.delete_expenses_name.resize(self.label_width + 50,self.label_height)
        self.delete_expenses_name.setFont(self.s_font)
        self.delete_expenses_name.move(self.line_width + 50,80)

        move_y = 80

        self.new_expenses_line.resize(self.line_width,self.line_height)
        self.new_expenses_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.new_expenses_line.move(0,move_y)
        self.new_expenses_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        
        self.add_button.resize(150,70)
        self.add_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_button.move(70,move_y + 90)
        self.add_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.setup_combo_box(self.expenses_combobox)
        self.expenses_combobox.resize(self.line_width,self.line_height)
        self.expenses_combobox.move(self.line_width + 50,move_y)

        self.delete_button.resize(160,70)
        self.delete_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_button.move(self.line_width + 120,move_y + 90)
        self.delete_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.expenses_data(self.expenses_combobox)
        self.delete_button.clicked.connect(self.delete_expenses_clicked)

        self.add_button.clicked.connect(self.expenses_add_button_clicked)
        
class check_expenses_page(BasePage):
    current_sum = 0
    query_str = ""
    def __init__(self):
        super().__init__('ui\\check_expenses_page.ui')
        self.init_ui()
        self.setup_database()
        self.current_page = 0
        self.query_str = "ORDER BY id DESC "
        self.update_model()

        self.check_expenses_button.setStyleSheet("""
                border-radius:0px;
                border:0px;
                background-color:white;
                """)
        self.add_record_button.clicked.connect(lambda: main_window.switch_page(1))
        self.check_record_button.clicked.connect(lambda: main_window.switch_page(2))
        self.add_expenses_button.clicked.connect(lambda: main_window.switch_page(3))
        self.check_invoice_button.clicked.connect(lambda: main_window.switch_page(5))
        self.add_salary_button.clicked.connect(lambda: main_window.switch_page(7))
        self.check_salary_button.clicked.connect(lambda: main_window.switch_page(8))

    def setup_database(self):   
        self.model = ExpensesQueryModel()
        self.data_table.setModel(self.model)
        # self.update_model()
        self.proxy_model = HideFirstColumnProxyModel(self)
        self.proxy_model.setSourceModel(self.model)
        self.data_table.setModel(self.proxy_model)

    def update_model(self):
        print("in update_model")
        if len(self.query_str) == 0:
            self.query_str = "ORDER BY id DESC "

        sum_text = "SELECT amount FROM Expenses_Record " + self.query_str
        query = QSqlQuery()
        query.exec(sum_text)        

        self.current_sum = 0
        while query.next():
            if contain_comma(str(query.value(0))):
                self.current_sum += locale.atof(query.value(0))
            else:
                self.current_sum += query.value(0)

        if is_3_decimal_places(self.current_sum):
            self.current_sum = locale.format_string("%.3f", self.current_sum, grouping=True)
        else:
            self.current_sum = locale.format_string("%.2f", self.current_sum, grouping=True)        
        
        
        self.amount_line.setText(str(self.current_sum))

        offset = self.current_page * ROWS_PER_PAGE
        query = "SELECT * FROM Expenses_Record " + self.query_str + f"LIMIT {ROWS_PER_PAGE} OFFSET {offset}"
        self.model.setQuery(query)
        self.page_number.setText(str(self.current_page+1))
        
        self.data_table.setColumnWidth(0, 200)
        self.data_table.setColumnWidth(1, 200)
        self.data_table.setColumnWidth(2, 250)
        self.data_table.setColumnWidth(3, 325)
        self.data_table.setColumnWidth(4, 425)
        self.data_table.setColumnWidth(5, 200)

        # for i in range(7):
        #     if i == 2:
        #         self.data_table.setColumnWidth(i, 300)
        #     elif i == 3:
        #         self.data_table.setColumnWidth(i, 500)
        #     elif i == 4:
        #         self.data_table.setColumnWidth(i, 300)
        #     else:
        #         self.data_table.setColumnWidth(i, 200)

    def clearCheckboxes(self,scroll_widget):
        scroll_layout = scroll_widget.layout()
        for i in reversed(range(scroll_layout.count())):
            item = scroll_layout.itemAt(i)
            if isinstance(item.widget(), QCheckBox):
                item.widget().deleteLater()

    def lorry_data(self):
        self.clearCheckboxes(self.lorry_scroll_widget)
        self.lorry_checkboxes = []
        scroll_layout = self.lorry_scroll_widget.layout()

        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            self.lorry_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中
        
    def expenses_data(self):
        self.clearCheckboxes(self.expenses_scroll_widget)
        self.expenses_checkboxes = []
        scroll_layout = self.expenses_scroll_widget.layout()

        query = QSqlQuery("SELECT description FROM Expenses")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            self.expenses_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中

    def company_data(self):
        self.clearCheckboxes(self.company_scroll_widget)
        self.company_checkboxes = []
        scroll_layout = self.company_scroll_widget.layout()

        query = QSqlQuery("SELECT name FROM Expenses_Company")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            self.company_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中

    def init_ui(self):
        screen = QDesktopWidget().screenGeometry()

        self.check_history_label.resize(self.label_width + 50,self.label_height)
        self.check_history_label.setFont(self.l_font)
        self.check_history_label.move(0,0)

        self.data_frame.resize(screen.width() - 900,screen.height() - 230)
        self.data_frame.move(110,50)

        self.data_month.resize(self.line_width,self.line_height)
        self.data_month.move(0,self.label_height)
        self.data_month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.data_month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.data_month.addItem("Not Chosen")
        self.data_month.addItems(months)
        self.data_month.setCurrentIndex(0)

        self.data_year.resize(self.line_width,self.line_height)
        self.data_year.move(self.line_width + 20 , self.label_height)
        self.data_year.setMinimum(2022)
        self.data_year.setMaximum(2030)
        self.data_year.setValue(2023)
        self.data_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.data_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

        self.data_table.resize(self.data_frame.width(),self.data_frame.height()-220)
        self.data_table.move(0,self.label_height + 60)
        self.data_table.setFont(QFont("Times New Roman",13))
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.data_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.data_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        delegate = CenteredNumberDelegate()
        self.data_table.setItemDelegate(delegate)
        self.data_table.setSelectionBehavior(QTableView.SelectRows)
        self.data_table.hideColumn(0)
        self.data_table.setStyleSheet(
            "QTableView::item:selected{ background-color:black;color:white;}"
            "QTableView { color: black; gridline-color: black; border: 2px solid black; background-color:white ;gridline-width: 2px;}"
            "QHeaderView { background-color: rgba(216,216,216,200); color:black ;font-family: Rockwell ; font-size: 27px; }"

            "QHeaderView::section { background-color: rgba(216,216,216,200) ;color: black;height: 50px;}"
        )
        
        
        self.prev_button.resize(100,50)
        self.prev_button.move(self.data_table.width()//2 - 150 ,self.data_frame.height()-50)
        self.prev_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.prev_button.setStyleSheet("QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
                                       "QPushButton:hover{ border: 3px solid black }")
        
        self.page_number.resize(self.label_width,self.label_height)
        self.page_number.move(self.data_table.width()//2,self.data_frame.height()-65 )
        self.page_number.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        

        self.next_button.resize(100,50)
        self.next_button.move(self.data_frame.width()//2 + 60,self.data_frame.height()-50)
        self.next_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.next_button.setStyleSheet(
            "QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
            "QPushButton:hover{ border: 3px solid black }"
            )
        
        self.filter_frame.resize(800,1000)
        self.filter_frame.move(screen.width()-750,150)
        
        self.check_record_lorry_label.resize(self.label_width,self.label_height)
        self.check_record_lorry_label.setFont(self.s_font)
        self.check_record_lorry_label.move(0,0)        

    
        self.check_record_company_label.resize(self.label_width,self.label_height)
        self.check_record_company_label.setFont(self.s_font)
        self.check_record_company_label.move(self.line_width + 30,0)
        move_y = self.label_height - 15

        
        self.lorry_scroll.setFixedSize(self.line_width,220)
        self.lorry_scroll.move(0,move_y)
        self.lorry_scroll.setWidgetResizable(True)
        self.lorry_scroll.setStyleSheet(
            "QScrollArea{"
            "    border: 2px solid black; color: black;"
            "}"
        )
        scroll_layout = QVBoxLayout()
        scroll_layout = self.lorry_scroll_widget.setLayout(scroll_layout)
        self.lorry_data()

        
        self.company_scroll.setFixedSize(self.line_width,220)
        self.company_scroll.move(self.line_width + 30,move_y)
        self.company_scroll.setWidgetResizable(True)
        self.company_scroll.setStyleSheet(
            "QScrollArea{"
            "    border: 2px solid black; color: black;"
            "}"
        )   
        scroll_layout = QVBoxLayout()
        self.company_scroll_widget.setLayout(scroll_layout)
        self.company_data()
        
        self.expenses_label.resize(self.label_width,self.label_height)
        self.expenses_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 170
        self.expenses_label.move(0,move_y)

        self.expenses_scroll.setFixedSize(self.line_width,220)
        self.expenses_scroll.move(0,move_y + self.label_height - 15)
        self.expenses_scroll.setWidgetResizable(True)
        self.expenses_scroll.setStyleSheet(
            "QScrollArea{"
            "    border: 2px solid black; color: black;"
            "}"
        )   
        scroll_layout = QVBoxLayout()
        self.expenses_scroll_widget.setLayout(scroll_layout)
        self.expenses_data()

        # 连接按钮的点击事件
        self.prev_button.clicked.connect(self.prev_page)
        self.next_button.clicked.connect(self.next_page)

        # 显示第一页数据
        self.current_page = 0
        
        self.search_button.resize(160,70)
        self.search_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.search_button.move(screen.width() - 950,90)
        self.search_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.search_button.clicked.connect(self.search_button_clicked)

        self.refresh_button.resize(180,70)
        self.refresh_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.refresh_button.move(screen.width() - 1160,90)
        self.refresh_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        
        self.refresh_button.clicked.connect(self.refresh_button_clicked)

        self.search_label.resize(self.label_width,self.label_height)
        self.search_label.setFont(self.s_font)
        self.search_label.move(870,95)   

        self.search_box.resize(self.line_width,self.line_height)
        self.search_box.setFont(self.s_font)
        self.search_box.move(1000,110)
        self.search_box.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }")

        self.update_button.resize(160,70)
        self.update_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_button.move(screen.width() - 350,screen.height() - 550)
        self.update_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_button.clicked.connect(self.update_record_clicked)

        self.export_button.resize(180,70)
        self.export_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_button.move(screen.width() - 355,screen.height() - 300)
        self.export_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_button.clicked.connect(self.export_button_clicked)
        
        self.delete_record_button.resize(160,70)
        self.delete_record_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_record_button.move(screen.width() - 350,screen.height() - 450)
        self.delete_record_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.delete_record_button.clicked.connect(self.delete_record_clicked)

        
        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        self.amount_label.move(screen.width() - 700,screen.height() - 550)   

        self.amount_line.resize(self.label_width,self.label_height)
        self.amount_line.setFont(self.l_font)
        self.amount_line.move(screen.width() - 700,screen.height() - 490)   

    def refresh_button_clicked(self):
        self.query_str = "ORDER BY id DESC "
        self.current_page = 0
        self.update_model()

    def search_button_clicked(self):
        str = "WHERE "
        temp = ""
        condition_num = 0

        # month filter
        month = self.data_month.currentIndex()
        if month > 0:
            if month < 10:
                temp = f"date like '%/0{month}/%'"
            else:
                temp = f"date like '%/{month}/%'"

            str += temp
            condition_num += 1

        if condition_num > 0:
            str += " AND "
        
        # year filter
        year = self.data_year.value()
        temp = f"date like '%/%/{year}'"
        str += temp 
        condition_num += 1

        # lorry filter
        checked_boxes = []
        for checkbox in self.lorry_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())

        if len(checked_boxes) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "
            
            temp += " ( "
            first = True
            for lorry in checked_boxes:
                if not first:
                    temp += " OR "

                first = False
                temp += f"lorry = '{lorry}'"
            
            temp += " ) "
            str += temp 
            condition_num += 1
        
        # expenses filter
        checked_boxes = []
        for checkbox in self.expenses_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())

        if len(checked_boxes) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "
            
            temp += " ( "
            first = True
            for expenses in checked_boxes:
                if not first:
                    temp += " OR "

                first = False
                temp += f"expenses = '{expenses}'"
                
                
            temp += " ) "
            str += temp 

            condition_num += 1

        # company filter
        checked_boxes = []
        for checkbox in self.company_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())

        if len(checked_boxes) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "
            
            temp += " ( "
            first = True
            for company in checked_boxes:
                if not first:
                    temp += " OR "

                first = False
                temp += f"company = '{company}'"
                
                
            temp += " ) "
            str += temp 

            condition_num += 1

        # keyword search
        keyword = self.search_box.text()
        if len(keyword) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "

            temp += f"( LOWER(lorry) LIKE LOWER('%{keyword}%') OR LOWER(reference) LIKE LOWER('%{keyword}%') OR LOWER(expenses) LIKE LOWER('%{keyword}%') OR LOWER(company) LIKE LOWER('%{keyword}%') OR LOWER(date) LIKE LOWER('%{keyword}%'))"

            str += temp 
            condition_num += 1

        if len(str) == 0:
            self.query_str = ""
        else:
            self.query_str = str + " "

        self.current_page = 0
        self.update_model()

    def export_button_clicked(self):
        export_dialog = ExportExpensesDialog(self)
        export_dialog.exec_()

    def update_record_clicked(self):
        row_data = self.get_table_record()

        if row_data != -1:
            if row_data[5] == '-' :
                update_dialog = UpdateLorryExpensesDialog(row_data,self)
                update_dialog.exec_()
            else:
                update_dialog = UpdateCreditorExpensesDialog(row_data,self)
                update_dialog.exec_()
            self.current_page = 0
            self.update_model()
            
    def delete_record_clicked(self):
        row_data = self.get_table_record()
        if row_data != -1:
            query = QSqlQuery() 
            print("Selected Row Data:", row_data)
            query.prepare(f"DELETE FROM Expenses_Record WHERE id = {row_data[0]}")
            

            if query.exec():
                print("Data daleted successfully.")
            else:
                print("Error while deleting data:", query.lastError().text())

            self.show_message_box("Delete Expenses","Expenses Record Deleted")
            # self.current_page = 0
            self.update_model()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_model()
        self.page_number.setText(str(self.current_page+1))

    def next_page(self):
        offset = (self.current_page + 1) * ROWS_PER_PAGE
        query = QSqlQuery()
        # query_str = f"SELECT * FROM Expenses_Record LIMIT 1 OFFSET {offset}"
        query_str = "SELECT * FROM Expenses_Record " + self.query_str + f"LIMIT 1 OFFSET {offset}"
        query.exec(query_str)
        if query.next():
            self.model.setQuery(query_str)
            self.current_page += 1
            self.update_model()
            self.page_number.setText(str(self.current_page+1))

    def get_table_record(self):
        # 获取选中模型
        selection_model = self.data_table.selectionModel()
        
        # 获取选中的行索引
        selected_indexes = selection_model.selectedRows()
     
        if not selected_indexes:
            # 如果没有选中的行，弹出消息框提示
            self.show_message_box("Warning","Please select a row.")
            return -1

        # 获取第一个选中的行索引
        row = selected_indexes[0].row()

        row_data = [self.model.data(self.model.index(row, column), Qt.DisplayRole) for column in range(self.model.columnCount())]
        return row_data

class add_salary_page(BasePage):
    def __init__(self):
        super().__init__('ui\\add_salary_page.ui')

        self.init_ui()

        self.add_salary_button.setStyleSheet("""
                border-radius:0px;
                border:0px;
                background-color:white;
                """)
        self.add_record_button.clicked.connect(lambda: main_window.switch_page(1))
        self.check_record_button.clicked.connect(lambda: main_window.switch_page(2))
        self.add_expenses_button.clicked.connect(lambda: main_window.switch_page(3))
        self.check_expenses_button.clicked.connect(lambda: main_window.switch_page(4))
        self.check_invoice_button.clicked.connect(lambda: main_window.switch_page(5))
        self.check_salary_button.clicked.connect(lambda: main_window.switch_page(8))

    def init_ui(self):
        self.workspace_frame.resize(600,1200)
        self.workspace_frame.move(600,210)
        
        self.add_salary_label.resize(self.label_width + 200,self.label_height) 
        self.add_salary_label.setFont(self.l_font)
        self.add_salary_label.move(0,0)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(0,80)
        move_y = 80

        self.date_line.setDate(QDate.currentDate())
        self.date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.date_line.setDisplayFormat("dd/MM/yyyy")
        self.date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.date_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.date_line.move(0,move_y)
        self.date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:10px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 28px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white }"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:20px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 35px; height: 35px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 3px solid black; font-family: Constantia; font-size: 28px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:50px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:50px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_left.png); }"
        )
    
        self.employee_label.resize(self.label_width,self.label_height)
        self.employee_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.employee_label.move(0,move_y)        

        self.employee_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.employee_line.move(0,move_y)
        self.setup_combo_box(self.employee_line)
        self.employee_data(self.employee_line)

        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        
        move_y = move_y + self.label_height - 15
        self.submit_button.resize(180,70)
        self.submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.submit_button.move(50,move_y + 50)
        self.submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.submit_button.clicked.connect(self.salary_submit_button_clicked)

        # add_frame
        screen = QDesktopWidget().screenGeometry()
        self.add_salary_frame.resize(800,1000)
        self.add_salary_frame.move(screen.width() - 1200,210)

        font = QFont("Tw Cen MT Condensed Extra Bold",20)
        font.setUnderline(True)
        self.add_employee_label.resize(self.label_width + 100,self.label_height)
        self.add_employee_label.setFont(font)
        self.add_employee_label.move(0,0)

        self.employee_name.resize(self.label_width + 50,self.label_height)
        self.employee_name.setFont(self.s_font)
        self.employee_name.move(0,80)

        self.delete_employee_name.resize(self.label_width + 50,self.label_height)
        self.delete_employee_name.setFont(self.s_font)
        self.delete_employee_name.move(self.line_width + 50,80)

        move_y = 80

        self.employee_name_line.resize(self.line_width,self.line_height)
        self.employee_name_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.employee_name_line.move(0,move_y)
        self.employee_name_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")  

        self.setup_combo_box(self.employee_combobox)
        self.employee_combobox.resize(self.line_width,self.line_height)
        self.employee_combobox.move(self.line_width + 50,move_y)
        self.employee_data(self.employee_combobox)

        self.delete_button.resize(160,70)
        self.delete_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_button.move(self.line_width + 120,move_y + 90)
        self.delete_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")

        self.update_button.resize(160,70)
        self.update_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_button.move(self.line_width + 120,move_y + 180)
        self.update_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_button.clicked.connect(self.update_button_clicked)

        self.employee_ic.resize(self.label_width + 50,self.label_height)
        self.employee_ic.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.employee_ic.move(0,move_y)

        self.employee_ic_line.resize(self.line_width,self.line_height)
        self.employee_ic_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.employee_ic_line.move(0,move_y)
        self.employee_ic_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")  
        
        self.employee_lorry.resize(self.label_width + 50,self.label_height)
        self.employee_lorry.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.employee_lorry.move(0,move_y)

        self.employee_lorry_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.employee_lorry_line.move(0,move_y)
        self.setup_combo_box(self.employee_lorry_line)
        self.lorry_data(self.employee_lorry_line)

    
        self.add_button.resize(150,70)
        self.add_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_button.move(70,move_y + 90)
        self.add_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        
        
        self.delete_button.clicked.connect(self.delete_employee_clicked)
        self.add_button.clicked.connect(self.add_employee_clicked)
    
    def update_button_clicked(self):
        if self.employee_combobox.currentText() != 'Not Chosen':
            update_dialog = UpdateEmployeeDialog(self.employee_combobox.currentText())
            update_dialog.exec_()
            self.employee_data(self.employee_line)
            self.employee_data(self.employee_combobox)
        else:
            self.show_message_box("Warning","No Employee is Chosen.")    
            
    def employee_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Employee")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
 
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
    
    def salary_submit_button_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
        
        query = QSqlQuery()
        query.prepare("INSERT INTO Salary(employee,date,amount) VALUES (:employee,:date,:amount)")
        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":employee", set_not_chosen(self.employee_line.currentText()))
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.show_message_box("Submit Salary Record","New Salary Record Added")

    def add_employee_clicked(self):
        name = self.employee_name_line.text()
        ic = self.employee_ic_line.text()
        lorry = self.employee_lorry_line.currentText()

        if len(name) == 0 or len(ic) == 0 or lorry == 'Not Chosen':
            return 
        
        query = QSqlQuery(f"SELECT * FROM Employee WHERE name = '{name}'")
        if query.next():
            print("Value already exists in the database.")
            return

        query = QSqlQuery()
        query.prepare("INSERT INTO Employee(name,ic,lorry) VALUES (:name,:ic,:lorry)")
        query.bindValue(":name", name)
        query.bindValue(":ic", ic)
        query.bindValue(":lorry", lorry)


        if query.exec():
            print("Data inserted successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        self.employee_data(self.employee_line)
        self.employee_data(self.employee_combobox)

        self.show_message_box(f"Add Employee",f"Save Employee: {name}")

    def delete_employee_clicked(self):
        name = self.employee_combobox.currentText()

        query = QSqlQuery(f"DELETE FROM Employee WHERE name = '{name}'")

        if query.exec():
            print("Data deleted successfully.")
        else:
            print("Error while deleting data:", query.lastError().text())

        self.show_message_box(f"Delete Employee",f"Employee : {name} Deleted")
        self.employee_data(self.employee_line)
        self.employee_data(self.employee_combobox)

class check_salary_page(BasePage):
    current_sum = 0
    query_str = ""
    def __init__(self):
        super().__init__('ui\\check_salary_page.ui')
        self.init_ui()
        self.setup_database()
        self.current_page = 0
        self.query_str = "ORDER BY id DESC "
        self.update_model()

        self.check_salary_button.setStyleSheet("""
                border-radius:0px;
                border:0px;
                background-color:white;
                """)
        self.add_record_button.clicked.connect(lambda: main_window.switch_page(1))
        self.check_record_button.clicked.connect(lambda: main_window.switch_page(2))
        self.add_expenses_button.clicked.connect(lambda: main_window.switch_page(3))
        self.check_expenses_button.clicked.connect(lambda: main_window.switch_page(4))
        self.check_invoice_button.clicked.connect(lambda: main_window.switch_page(5))
        self.add_salary_button.clicked.connect(lambda: main_window.switch_page(7))

    def setup_database(self):
        self.model = SalaryQueryModel()
        self.data_table.setModel(self.model)
        self.proxy_model = HideFirstColumnProxyModel(self)
        self.proxy_model.setSourceModel(self.model)
        self.data_table.setModel(self.proxy_model)

    def update_model(self):
        if len(self.query_str) == 0:
            self.query_str = "ORDER BY id DESC "

        sum_text = "SELECT amount FROM Salary " + self.query_str
        query = QSqlQuery()
        query.exec(sum_text)

        self.current_sum = 0
        while query.next():
            if contain_comma(str(query.value(0))):
                self.current_sum += locale.atof(query.value(0))
            else:
                self.current_sum += query.value(0)

        if is_3_decimal_places(self.current_sum):
            self.current_sum = locale.format_string("%.3f", self.current_sum, grouping=True)
        else:
            self.current_sum = locale.format_string("%.2f", self.current_sum, grouping=True)
            
        self.amount_line.setText(str(self.current_sum))

               
        

        offset = self.current_page * ROWS_PER_PAGE
    
        query = "SELECT * FROM Salary " + self.query_str + f"LIMIT {ROWS_PER_PAGE} OFFSET {offset}"
        self.model.setQuery(query)
        self.page_number.setText(str(self.current_page+1))

        self.data_table.setColumnWidth(0, 400)
        self.data_table.setColumnWidth(1, 900)
        self.data_table.setColumnWidth(2, 400)

    def clearCheckboxes(self,scroll_widget):
        scroll_layout = scroll_widget.layout()
        for i in reversed(range(scroll_layout.count())):
            item = scroll_layout.itemAt(i)
            if isinstance(item.widget(), QCheckBox):
                item.widget().deleteLater()

    def employee_data(self):
        self.clearCheckboxes(self.employee_scroll_widget)
        self.employee_checkboxes = []
        scroll_layout = self.employee_scroll_widget.layout()

        query = QSqlQuery("SELECT name FROM Employee")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            self.employee_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中
        
    def init_ui(self):
        screen = QDesktopWidget().screenGeometry()
        self.check_history_label.resize(self.label_width + 50,self.label_height)
        self.check_history_label.setFont(self.l_font)
        self.check_history_label.move(0,0)

        self.data_frame.resize(screen.width() - 700,screen.height() - 230)
        self.data_frame.move(110,50)

        self.data_month.resize(self.line_width,self.line_height)
        self.data_month.move(0,self.label_height)
        self.data_month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.data_month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.data_month.addItem("Not Chosen")
        self.data_month.addItems(months)
        self.data_month.setCurrentIndex(0)

        self.data_year.resize(self.line_width,self.line_height)
        self.data_year.move(self.line_width + 20 , self.label_height)
        self.data_year.setMinimum(2022)
        self.data_year.setMaximum(2030)
        self.data_year.setValue(2023)
        self.data_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.data_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

        self.data_table.resize(self.data_frame.width(),self.data_frame.height()-220)
        self.data_table.move(0,self.label_height + 60)
        self.data_table.setFont(QFont("Times New Roman",13))
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.data_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.data_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        delegate = CenteredNumberDelegate()
        self.data_table.setItemDelegate(delegate)
        self.data_table.setSelectionBehavior(QTableView.SelectRows)
        self.data_table.hideColumn(0)
        self.data_table.setStyleSheet(
            "QTableView::item:selected{ background-color:black;color:white;}"
            "QTableView { color: black; gridline-color: black; border: 2px solid black; background-color:white ;gridline-width: 2px;}"
            "QHeaderView { background-color: rgba(216,216,216,200); color:black ;font-family: Rockwell ; font-size: 27px; }"

            "QHeaderView::section { background-color: rgba(216,216,216,200) ;color: black;height: 50px;}"
        )
        
        self.prev_button.resize(100,50)
        self.prev_button.move(self.data_table.width()//2 - 150 ,self.data_frame.height()-50)
        self.prev_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.prev_button.setStyleSheet("QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
                                       "QPushButton:hover{ border: 3px solid black }")
        
        self.page_number.resize(self.label_width,self.label_height)
        self.page_number.move(self.data_table.width()//2,self.data_frame.height()-65 )
        self.page_number.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        
        self.next_button.resize(100,50)
        self.next_button.move(self.data_frame.width()//2 + 60,self.data_frame.height()-50)
        self.next_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",15))
        self.next_button.setStyleSheet(
            "QPushButton { background-color:white;color:black;border-radius:20px;border:2px solid black; padding :5px;}"
            "QPushButton:hover{ border: 3px solid black }"
            )
        
        self.filter_frame.resize(500,1000)
        self.filter_frame.move(screen.width()-500,100)

        self.employee_label.resize(self.label_width,self.label_height)
        self.employee_label.setFont(self.s_font)
        self.employee_label.move(0,0)        

        move_y = self.label_height - 15
        
        self.employee_scroll.setFixedSize(self.line_width + 100,220)
        self.employee_scroll.move(0,move_y)
        self.employee_scroll.setWidgetResizable(True)
        self.employee_scroll.setStyleSheet(
            "QScrollArea{"
            "    border: 2px solid black; color: black;"
            "}"
        )
        scroll_layout = QVBoxLayout()
        self.employee_scroll_widget.setLayout(scroll_layout)
        self.employee_data()

        # 连接按钮的点击事件
        self.prev_button.clicked.connect(self.prev_page)
        self.next_button.clicked.connect(self.next_page)

        self.update_record_button.resize(160,70)
        self.update_record_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_record_button.move(screen.width() - 400,screen.height() - 550)
        self.update_record_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_record_button.clicked.connect(self.update_record_clicked)

        self.search_button.resize(160,70)
        self.search_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.search_button.move(screen.width() - 750,90)
        self.search_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.search_button.clicked.connect(self.search_button_clicked)

        self.refresh_button.resize(180,70)
        self.refresh_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.refresh_button.move(screen.width() - 960,90)
        self.refresh_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.refresh_button.clicked.connect(self.refresh_button_clicked)

        self.search_label.resize(self.label_width,self.label_height)
        self.search_label.setFont(self.s_font)
        self.search_label.move(970,95)   

        self.search_box.resize(self.line_width,self.line_height)
        self.search_box.setFont(self.s_font)
        self.search_box.move(1100,110)
        self.search_box.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }")


        self.export_button.resize(180,70)
        self.export_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_button.move(screen.width() - 405,screen.height() - 300)
        self.export_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_button.clicked.connect(self.export_button_clicked)

        self.delete_record_button.resize(160,70)
        self.delete_record_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.delete_record_button.move(screen.width() - 400,screen.height() - 450)
        self.delete_record_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.delete_record_button.clicked.connect(self.delete_record_clicked)

        
        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        self.amount_label.move(screen.width() - 450,600)   

        self.amount_line.resize(self.label_width+ 100,self.label_height)
        self.amount_line.setFont(self.l_font)
        self.amount_line.move(screen.width() - 450,650)   

    def refresh_button_clicked(self):
        self.query_str = "ORDER BY id DESC "
        self.current_page = 0
        self.update_model()
        # self.clearCheckboxes()
        
    def search_button_clicked(self):
        str = "WHERE "
        temp = ""
        condition_num = 0

        # month filter
        month = self.data_month.currentIndex()
        if month > 0:
            if month < 10:
                temp = f"date like '%/0{month}/%'"
            else:
                temp = f"date like '%/{month}/%'"

            str += temp
            condition_num += 1

        if condition_num > 0:
            str += " AND "
        
        # year filter
        year = self.data_year.value()
        temp = f"date like '%/%/{year}'"
        str += temp 
        condition_num += 1

        # employee filter
        checked_boxes = []
        for checkbox in self.employee_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())

        if len(checked_boxes) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "
            
            temp += " ( "
            first = True
            for employee in checked_boxes:
                if not first:
                    temp += " OR "

                first = False
                temp += f"employee = '{employee}'"
            
            temp += " ) "
            str += temp 
            condition_num += 1
        
        # keyword search
        keyword = self.search_box.text()
        if len(keyword) > 0:
            temp=""
            if condition_num > 0:
                str += " AND "

            temp += f"( LOWER(employee) LIKE LOWER('%{keyword}%') OR LOWER(date) LIKE LOWER('%{keyword}%'))"

            str += temp 
            condition_num += 1


        if len(str) == 0:
            self.query_str = ""
        else:
            self.query_str = str + " "

        self.current_page = 0
        self.update_model()
        
    def update_record_clicked(self):
        row_data = self.get_table_record()

        if row_data != -1:
            # 实例化 UpdateRecordDialog，并显示弹窗
            update_dialog = UpdateSalaryDialog(row_data,self)
            update_dialog.exec_()

            # self.current_page = 0
            self.update_model()
            
    def delete_record_clicked(self):
        row_data = self.get_table_record()
        if row_data != -1:
            query = QSqlQuery() 
            print("Selected Row Data:", row_data)
            query.prepare(f"DELETE FROM Salary WHERE id = {row_data[0]}")
            

            if query.exec():
                print("Data daleted successfully.")
            else:
                print("Error while deleting data:", query.lastError().text())

            self.show_message_box("Delete Salary Record","Salary Record Deleted")
            # self.current_page = 0
            self.update_model()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_model()
        self.page_number.setText(str(self.current_page+1))

    def next_page(self):
        offset = (self.current_page + 1) * ROWS_PER_PAGE
        query = QSqlQuery()
        query_str = "SELECT * FROM Salary " + self.query_str + f"LIMIT 1 OFFSET {offset}"
        query.exec(query_str)
        if query.next():
            self.model.setQuery(query_str)
            self.current_page += 1
            self.update_model()
            self.page_number.setText(str(self.current_page+1))

    def get_table_record(self):
        # 获取选中模型
        selection_model = self.data_table.selectionModel()
        
        # 获取选中的行索引
        selected_indexes = selection_model.selectedRows()
     
        if not selected_indexes:
            # 如果没有选中的行，弹出消息框提示
            self.show_message_box("Warning","Please select a row.")
            return -1

        # 获取第一个选中的行索引
        row = selected_indexes[0].row()

        row_data = [self.model.data(self.model.index(row, column), Qt.DisplayRole) for column in range(self.model.columnCount())]

        return row_data

    def export_button_clicked(self):
        export_dialog = ExportSalaryDialog(self)
        export_dialog.exec_()
  
class MainWindow(QMainWindow):
    page_list = {}

    def __init__(self):
        super().__init__()
        
        icon = QIcon('photo\\icon.ico')
        self.setWindowIcon(icon)
        
        self.showMaximized()  # 设置窗口最大化
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.central_widget = QWidget()
        
        # self.central_widget.setStyleSheet("background-color: #FFAA66;")
        # self.central_widget.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0 rgba(255, 148, 90, 150), stop:1 rgba(255, 217, 189, 100));")
        self.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0 rgba(240, 158, 90, 255), stop:1 rgba(240, 227, 189, 255));")

        self.layout = QVBoxLayout(self.central_widget)

        self.stacked_widget = QStackedWidget()
        self.layout.addWidget(self.stacked_widget)

        #先加载home_page放入stackWidget
        self.home_page = HomePage()
        self.stacked_widget.addWidget(self.home_page)
        self.setCentralWidget(self.central_widget)

        #记录home_page已加载，并且记录它在stackWidget里的index
        self.current_page_index = 0
        self.page_list[self.current_page_index] = 0

        self.init_database()

    def init_database(self):

        self.db = QSqlDatabase.addDatabase("QSQLITE")
        self.db.setDatabaseName(DB_NAME)
        if not self.db.open():
            print("Failed to connect to the database.")
            return

        # 查询数据总条数
        # self.cur.execute(f"SELECT COUNT(*) FROM {table_name}")
        # self.total_rows = self.cur.fetchone()[0]

    def create_page(self, page_index):
        if page_index == 1:
            self.add_record_page = add_record_page()
            self.stacked_widget.addWidget(self.add_record_page)
        elif page_index == 2:
            self.check_record_page = check_record_page()
            self.stacked_widget.addWidget(self.check_record_page)
        elif page_index == 3:
            self.add_expenses_page = add_expenses_page()
            self.stacked_widget.addWidget(self.add_expenses_page)
        elif page_index == 4:
            self.check_expenses_page = check_expenses_page()
            self.stacked_widget.addWidget(self.check_expenses_page)
        elif page_index == 5:
            self.check_invoice_page = check_invoice_page()
            self.stacked_widget.addWidget(self.check_invoice_page)
        elif page_index == 6:
            self.home_page2 = HomePage2()
            self.stacked_widget.addWidget(self.home_page2)
        elif page_index == 7:
            self.add_salary_page = add_salary_page()
            self.stacked_widget.addWidget(self.add_salary_page)
        elif page_index == 8:
            self.check_salary_page = check_salary_page()
            self.stacked_widget.addWidget(self.check_salary_page)

    def switch_page(self, page_index):
        self.current_page_index = page_index

        if page_index not in self.page_list:
            self.page_list[page_index] = self.stacked_widget.count() 
            stack_index = self.page_list[page_index]  
            self.create_page(page_index)
        elif page_index in [2,4,5,8]:
            if page_index == 2:
                self.check_record_page.lorry_data()
                self.check_record_page.company_data()
                self.check_record_page.update_model()
            elif page_index == 4:
                self.check_expenses_page.lorry_data()
                self.check_expenses_page.expenses_data()
                self.check_expenses_page.company_data()
                self.check_expenses_page.update_model()
            elif page_index == 5:
                self.check_invoice_page.company_data()
                self.check_invoice_page.update_model()
            elif page_index == 8:
                self.check_salary_page.employee_data()
                self.check_salary_page.update_model()
            stack_index = self.page_list[page_index]
        else:
            stack_index = self.page_list[page_index]
        self.stacked_widget.setCurrentIndex(stack_index)

    
if __name__ == '__main__':
    app = QApplication(sys.argv)
    open_database()
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())

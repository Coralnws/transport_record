import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QStackedWidget, QVBoxLayout, QWidget,QDesktopWidget,QSizePolicy,QGraphicsDropShadowEffect,QTableView,QStyledItemDelegate,QMessageBox,QDialog,QCheckBox,QHeaderView
from PyQt5.QtGui import QPixmap,QIcon,QFont
from PyQt5.uic import loadUi
from PyQt5.QtCore import Qt,QSize,QDate,QSortFilterProxyModel,QSize
import os,re
from PyQt5.QtSql import QSqlDatabase, QSqlQuery,QSqlQueryModel  
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import range_boundaries
import locale

# 设置地区格式（这里使用美国地区的格式）
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
# 数据库名称，请将你的数据库文件名填入这里
DB_NAME = "database.db"
 
# 每页显示的数据条数
ROWS_PER_PAGE = 18

script_path = os.path.abspath(__file__)
currrent_folder = os.path.dirname(script_path)
parent_folder = os.path.dirname(currrent_folder)
output_folder = os.path.join(currrent_folder, 'output_folder')

def expenses_company_data(self,combobox):
    combobox.clear()
    combobox.addItem("Not Chosen")
    query = QSqlQuery("SELECT name FROM Expenses_Company")
    while query.next():
        name = query.value(0)
        combobox.addItem(name)
    combobox.setCurrentIndex(0)

def is_numeric_string(s):
    # 使用正则表达式判断字符串是否只包含数字和小数点
    pattern = r'^[0-9.,]+$'
    if re.match(pattern, s) or (re.match(pattern, s) and contain_comma(s)):
        return True
    
def set_not_chosen(str):
    print(str)
    if str == 'Not Chosen':
        return '-'
    else:
        return str

def contain_comma(str):
    if ',' in str:
        return True
    
    return False
    

def rebuild_numeric_string(value):
    str = ""
    for c in value:
        if c == ',':
            continue
            
        str += c

    return str

def is_3_decimal_places(value):
    if not is_numeric_string:
        return
    # 将float转换为字符串
    if contain_comma(str(value)):
        value = float(rebuild_numeric_string(str(value)))

    value_str = "{:.3f}".format(value)
    # 检测字符串是否为3位小数
    return value_str[-1] != '0'

class CenteredNumberDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        option.displayAlignment = Qt.AlignCenter
        super().paint(painter, option, index)


    def displayText(self, value, locale):
        if isinstance(value, (float,int)):
            if is_3_decimal_places(value):
                return "{:.3f}".format(value) 
            return "{:.2f}".format(value)  # 设置数字显示为两位小数
        return super().displayText(value, locale)

class RecordQueryModel(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and index.isValid():
            
            value = super().data(index, role)
            column = index.column()
            if column in [6,7,8] and isinstance(value, (float,int)):
                if contain_comma (str(value)):
                    value = float(rebuild_numeric_string(value))
                else:
                    value = float(value)
                if is_3_decimal_places(value):
                    return locale.format_string("%.3f", value, grouping=True)
                else:
                    return locale.format_string("%.2f", value, grouping=True)
            
            if column in [6,7,8] and isinstance(value, str) and is_numeric_string(value):
                if contain_comma (str(value)):
                    value = float(rebuild_numeric_string(value))
                else:
                    value = float(value)
                if is_3_decimal_places(value):
                    return locale.format_string("%.3f", value, grouping=True)
                else:
                    return locale.format_string("%.2f", value, grouping=True)
            
        return super().data(index, role)


    def headerData(self, section, orientation, role=Qt.DisplayRole):
        COLUMN_HEADERS = ["ID","Date", "Lorry No.", "Reference No.", "Debtor","Description","Quantity","Unit Price","Amount","Record Date"]
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return COLUMN_HEADERS[section] if 0 <= section < len(COLUMN_HEADERS) else ""
        return super().headerData(section, orientation, role)

class SalaryQueryModel(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and index.isValid():
            
            value = super().data(index, role)
            column = index.column()
            if column in [3] and isinstance(value, (float,int)):
                if contain_comma (str(value)):
                    value = float(rebuild_numeric_string(value))
                else:
                    value = float(value)
                if is_3_decimal_places(value):
                    print(locale.format_string("%.3f", value, grouping=True))
                    return locale.format_string("%.3f", value, grouping=True)
                else:
                    return locale.format_string("%.2f", value, grouping=True)

        return super().data(index, role)


    def headerData(self, section, orientation, role=Qt.DisplayRole):
        COLUMN_HEADERS = ["ID","Date", "Employee", "Amount"]
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return COLUMN_HEADERS[section] if 0 <= section < len(COLUMN_HEADERS) else ""
        return super().headerData(section, orientation, role)

class ExpensesQueryModel(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and index.isValid():
            
            value = super().data(index, role)
            column = index.column()
            if column in [6] and isinstance(value, (float,int)):
                if contain_comma (str(value)):
                    value = float(rebuild_numeric_string(value))
                else:
                    value = float(value)
                if is_3_decimal_places(value):
                    print(locale.format_string("%.3f", value, grouping=True))
                    return locale.format_string("%.3f", value, grouping=True)
                else:
                    return locale.format_string("%.2f", value, grouping=True)
        return super().data(index, role)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        COLUMN_HEADERS = ["ID","Date", "Lorry No.", "Ref. No","Expenses ","Company","Amount"]
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return COLUMN_HEADERS[section] if 0 <= section < len(COLUMN_HEADERS) else ""
        return super().headerData(section, orientation, role)

class InvoiceQueryModel(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and index.isValid():
            
            value = super().data(index, role)
            column = index.column()
            if column in [5] and isinstance(value, (float,int)):
                if contain_comma (str(value)):
                    value = float(rebuild_numeric_string(value))
                else:
                    value = float(value)
                if is_3_decimal_places(value):
                    return locale.format_string("%.3f", value, grouping=True)
                else:
                    return locale.format_string("%.2f", value, grouping=True)
        return super().data(index, role)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        COLUMN_HEADERS = ["ID","Invoice No.", "Company", "Month","Year","Amount"]
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return COLUMN_HEADERS[section] if 0 <= section < len(COLUMN_HEADERS) else ""
        return super().headerData(section, orientation, role)

class HideFirstColumnProxyModel(QSortFilterProxyModel):
    def filterAcceptsColumn(self, source_column, source_parent):
        # Hide the first column (index 0)
        return source_column != 0

class ExportRecordDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    export_type = -1  #0 - subfreige 1-invoice

    def __init__(self,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\export_record_dialog.ui', self)
        self.setWindowTitle("Export Excel")
        self.setMinimumSize(1400, 600)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()
    
    def init_ui(self):
        self.title_label.resize(self.label_width,self.label_height)
        self.title_label.setFont(self.l_font)
        self.title_label.move(80,40)

        self.filename_label.resize(self.label_width,self.label_height)
        self.filename_label.setFont(self.s_font)
        self.filename_label.move(425,40)

        self.filename_line.resize(self.line_width,self.line_height)
        self.filename_line.setFont(self.s_font)
        self.filename_line.move(600,60)
        self.filename_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
 
        self.company_label.resize(self.label_width + 50,self.label_height)
        self.company_label.setFont(self.s_font)
        self.company_label.move(90,130 + self.label_height - 15)

        self.company_line.resize(self.line_width,self.line_height)
        self.company_line.setFont(self.s_font)
        self.company_line.move(90,200 + self.label_height - 15)
        self.company_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        
        self.invoice_label.resize(self.label_width + 50,self.label_height)
        self.invoice_label.setFont(self.s_font)
        self.invoice_label.move(580,self.label_height * 3  + 50)

        self.invoice_company_label.resize(self.label_width,self.label_height)
        self.invoice_company_label.setFont(self.s_font)
        self.invoice_company_label.move(940,self.label_height * 3  + 50)

        self.invoice_company.resize(self.line_width+80,self.line_height)
        self.invoice_company.move(940,self.label_height * 4  + 45)
        self.setup_combo_box(self.invoice_company)
        self.company_data(self.invoice_company)

        self.invoice_line.resize(self.line_width,self.line_height)
        self.invoice_line.setFont(self.s_font)
        self.invoice_line.move(580,self.label_height * 4  + 45)
        self.invoice_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        self.export_subfreige_button.resize(340,70)
        self.export_subfreige_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_subfreige_button.move(80,self.label_height * 4 + 25)
        self.export_subfreige_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_subfreige_button.clicked.connect(self.export_subfreige_button_clicked)

        self.export_invoice_button.resize(320,70)
        self.export_invoice_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_invoice_button.move(800,self.label_height * 5 + 65)
        self.export_invoice_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_invoice_button.clicked.connect(self.export_invoice_button_clicked)

        self.line.resize(10,300)
        self.line.move(500,200)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(580,70 + self.label_height - 15)

        self.invoice_month.resize(self.line_width,self.line_height)
        self.invoice_month.move(580,150 + self.label_height - 15)
        self.invoice_month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.invoice_month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.invoice_month.addItem("Not Chosen")
        self.invoice_month.addItems(months)
        self.invoice_month.setCurrentIndex(0)


        self.invoice_year.resize(self.line_width,self.line_height)
        self.invoice_year.move(940,150 + self.label_height - 15)
        self.invoice_year.setMinimum(2022)
        self.invoice_year.setMaximum(2030)
        self.invoice_year.setValue(2023)
        self.invoice_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.invoice_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

    def company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)

    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def export_subfreige_button_clicked(self):
        self.export_file = self.filename_line.text()
        self.export_company = self.company_line.text()

        if len(self.export_file) == 0 or len(self.export_company) == 0:
            self.show_message_box("Warning","Please fill in all blank")
            return

        subfreige_folder = os.path.join(output_folder, 'Subfreige')
        os.makedirs(subfreige_folder, exist_ok=True)
        output_file = os.path.join(subfreige_folder, self.export_file+".xlsx")

        if os.path.exists(output_file):
            self.show_message_box("Duplicate Name","File exist, please enter another filename")
            print("File already exists. Please choose a different filename.")
            return
        
        self.export_file = output_file

        self.show_message_box("Export","Data Exported")

        self.export_type = 0

        # 显示弹窗
        self.close()

    def export_invoice_button_clicked(self):
        self.export_file = self.filename_line.text()
        self.export_invoice = self.invoice_line.text()
        self.export_month = self.invoice_month.currentText()
        self.export_year = self.invoice_year.value()
        self.export_company = self.invoice_company.currentText()
        
        if len(self.export_file) == 0 or len(self.export_invoice) == 0 or self.export_month=='Not Chosen' or self.export_company == 'Not Chosen':
            self.show_message_box("Warning","Please fill in all blank")
            return
        
        query = QSqlQuery()
        query.exec(f"SELECT * FROM Invoice WHERE no = '{self.export_invoice}'")

        if query.next():
            self.show_message_box("Warning","Invoice No Exist")
            return

        invoice_folder = os.path.join(output_folder, 'Invoice')
        os.makedirs(invoice_folder, exist_ok=True)
        output_file = os.path.join(invoice_folder, self.export_file+".xlsx")

        if os.path.exists(output_file):
            self.show_message_box("Duplicate Name","File exist, please enter another filename")
            print("File already exists. Please choose a different filename.")
            return
        
        self.export_file = output_file

        self.export_type = 1

        # 显示弹窗
        self.close()

    def show_message_box(self,title,content):
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle(title)
        message_box.setText(content + "       ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )
    
        message_box.exec_()
    
    def get_file(self):
        return self.export_file
    
    def get_company(self):
        return self.export_company
    
    def get_invoice_info(self):
        data = {'invoice':self.export_invoice,'month':self.export_month,'year':self.export_year,'company':self.export_company}
        return data

    def get_exported(self):
        return self.export_type
    
class ExportDebtorDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
    file_num = 1

    def __init__(self,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\export_debtor_dialog.ui', self)
        self.setWindowTitle("Export Excel")
        self.setMinimumSize(450, 600)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()
    
    def init_ui(self):
        self.title_label.resize(self.label_width,self.label_height)
        self.title_label.setFont(self.l_font)
        self.title_label.move(80,30)

        self.filename_label.resize(self.label_width,self.label_height)
        self.filename_label.setFont(self.s_font)
        self.filename_label.move(90,100)

        move_y = 100

        self.filename_line.resize(self.line_width,self.line_height)
        self.filename_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.filename_line.move(90,move_y)
        self.filename_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
 
        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        move_y += self.line_height + 30
        self.date_label.move(90,move_y)

        self.invoice_year.resize(self.line_width,self.line_height)
        move_y += self.label_height - 15
        self.invoice_year.move(90,move_y)
        self.invoice_year.setMinimum(2022)
        self.invoice_year.setMaximum(2030)
        self.invoice_year.setValue(2023)
        self.invoice_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.invoice_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

        self.export_button.resize(180,70)
        self.export_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_button.move(110,move_y + 100)
        self.export_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_button.clicked.connect(self.export_button_clicked)

    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def export_button_clicked(self):
        self.export_file = self.filename_line.text()
        self.export_year = self.invoice_year.value()
        
        if len(self.export_file) == 0:
            self.show_message_box("Warning","Please enter filename")
            return

        debtor_folder = os.path.join(output_folder, 'Debtor')
        os.makedirs(debtor_folder, exist_ok=True)
        new_folder = os.path.join(debtor_folder, self.export_file)

        if os.path.exists(new_folder):
            self.show_message_box("Duplicate Name","File exist, please enter another filename")
            print("File already exists. Please choose a different filename.")
            return
        
        os.makedirs(new_folder, exist_ok=True)

        self.export_file = os.path.join(new_folder, self.export_file)

        self.show_message_box("Export","Data Exported")

        self.export_debtor()

        # 显示弹窗
        self.close()

    def export_debtor(self):
        # 第一轮遍历就填
        total_amount = {}  # {'company_name' :  (integer amount) }
        workbook_list = {} # {'company_name' : (company_data[]) }
        workbook_month_count = {} # {'month' : (workbook data_count) }
        company_month_count = {}  # {'month' : (company data_count) }
        company_num = 0

        for month in self.months: 
            workbook_month_count[month] = 0
            company_month_count[month] = 0

        query = QSqlQuery()
        query.exec_(f"SELECT * FROM Invoice WHERE year = {self.export_year} ORDER BY company,no")         
        # query.exec_(f"SELECT * FROM Invoice WHERE year = 'You Hing Transport' ORDER BY company,no")
        
        query.next() 
        row_data =  {'no':query.value(1),'company':query.value(2),'month':query.value(3),'amount':query.value(5)}
        current_company = row_data['company']
        total_amount[current_company] = 0
        company_num += 1

        new_company_data = []
        new_company_data.append(row_data)
        if contain_comma(str(row_data['amount'])):
            total_amount[current_company] += locale.atof(row_data['amount'])
        else:    
            total_amount[current_company] += row_data['amount']

        company_month_count[row_data['month']] += 1
        
        # no , company , month , year , amount
        while query.next():
            row_data =  {'no':query.value(1),'company':query.value(2),'month':query.value(3),'amount':query.value(5)}
            
            if row_data['company'] != current_company: # current row_data is new company,don't save the new row data
                workbook_list[current_company] = new_company_data  # record data
                for key,value in company_month_count.items():
                    if workbook_month_count[key] < value:
                        workbook_month_count[key] = value

                if company_num == 6:
                    self.save_workbook(total_amount,workbook_list,workbook_month_count)
                    total_amount = {}  # {'company_name' :  (integer amount) }
                    workbook_list = {} # {'company_name' : (company_data[]) }
                    workbook_month_count = {} # {'month' : (workbook data_count) }
                    company_month_count = {}  # {'month' : (company data_count) }
                    company_num = 0

                    for month in self.months:
                        workbook_month_count[month] = 0

                current_company = row_data['company']  
                total_amount[current_company] = 0
                company_num += 1

                new_company_data = [] # open new company data
                company_month_count = {}
                for month in self.months:
                    company_month_count[month] = 0
                
            new_company_data.append(row_data)
            if contain_comma(str(row_data['amount'])):
                total_amount[current_company] += locale.atof(row_data['amount'])
            else:    
                total_amount[current_company] += row_data['amount']
            # total_amount[current_company] += row_data['amount']
            company_month_count[row_data['month']] += 1

        workbook_list[current_company] = new_company_data  # record data
        for key,value in company_month_count.items():
            if workbook_month_count[key] < value:
                workbook_month_count[key] = value

        self.save_workbook(total_amount,workbook_list,workbook_month_count)
    
    def save_workbook(self,total_amount,workbook_list,workbook_month_count):
        print(workbook_list)
        page_margins = PageMargins(
            left=0.25,  # 左边距
            right=0.25,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距 
        )

        # for company and invoice no column
        whole_thick_border = Border(left=Side(style='medium'), 
                        right=Side(style='medium'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))

        # for amount_column
        thick_dot_right_border = Border(left=Side(style='dashDot'), 
                        right=Side(style='medium'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        # for invoice column
        thick_dot_left_border = Border(left=Side(style='medium'), 
                        right=Side(style='dashDot'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))

        # for data left column
        normal_middle_left_border = Border(left=Side(style='medium'), 
                        right=Side(style='dashDot'), 
                        bottom=Side(style='thin'))
        
        merge_middle_left_border = Border(left=Side(style='medium'), 
                        right=Side(style='dashDot'))
        
        # for data right column
        normal_middle_right_border = Border(right=Side(style='medium'), 
                        left=Side(style='dashDot'), 
                        bottom=Side(style='thin'))
        
        merge_middle_right_border = Border(right=Side(style='medium'), 
                        left=Side(style='dashDot'))
        
                
        normal_font = Font(name='Times New Roman', size=10, bold=False, color='000000')
        bold_font = Font(name='Times New Roman', size=10, bold=True, color='000000')
        
        workbook_month_index = {}  # {'month' : (row) } 

        workbook = Workbook()
        sheet = workbook.active
        sheet.page_margins = page_margins
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

        current_row = 1
        company_column = 66 #B
        start_column = 65
        total_company_num = len(workbook_list)
        last_column = chr(start_column + (total_company_num * 2))
        company_num = 0
        current_company = ""
        # current company column = company_column + (company_num - 1) * 2 

        if self.file_num == 1:
            sheet.merge_cells(f"A1:{last_column}1")
            merged_cell = sheet['A1']
            merged_cell.font = bold_font
            merged_cell.value = "Debtors"
            for cell in sheet[1]:
                cell.border = whole_thick_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

        company_start_row = current_row

        # month column
        month_row = current_row + 2
        for month,count in workbook_month_count.items():
            if count == 0:
                workbook_month_index[month] = month_row
                sheet[f'A{month_row}'] = month
                month_row += 1
                continue
            end_row = month_row + count - 1
            sheet.merge_cells(f"A{month_row}:A{end_row}")
            merged_cell = sheet[f"A{month_row}"]
            merged_cell.value = month
            workbook_month_index[month] = month_row
            month_row = end_row + 1

        sheet[f'A{month_row}'] = "Total"
        amount_row = month_row

        for row in range(current_row,month_row + 1):
            sheet[f'A{row}'].border = whole_thick_border

        for company,company_data in workbook_list.items():
            company_num += 1
            current_row = company_start_row
            company_month_count = {}
            for month in self.months:
                company_month_count[month] = 0

            # company header
            current_column = company_column + (company_num - 1) * 2
            sheet.merge_cells(f"{chr(current_column)}{current_row}:{chr(current_column+1)}{current_row}")
            merged_cell = sheet[f'{chr(current_column)}{current_row}']
            merged_cell.font = normal_font
            merged_cell.value = company

            for cell in sheet[current_row]:
                cell.border = whole_thick_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

            # 2 - Invocie No & Amount
            invoice_column = chr(current_column)
            amount_column = chr(current_column+1)

            sheet[f'{invoice_column}{current_row}'] = "Invoice No"
            sheet[f'{invoice_column}{current_row}'].border = thick_dot_left_border
            sheet[f'{amount_column}{current_row}'] = "Amount"
            sheet[f'{amount_column}{current_row}'].border = thick_dot_right_border
            current_row += 1

            for month in self.months:
                current_month_count = 1
                pt_row = workbook_month_index[month]
                total_row = workbook_month_count[month]
                while current_month_count < total_row:
                    sheet[f'{invoice_column}{pt_row}'].border = merge_middle_left_border
                    sheet[f'{amount_column}{pt_row}'].border = merge_middle_right_border
                    current_month_count += 1
                    pt_row += 1

                sheet[f'{invoice_column}{pt_row}'].border = normal_middle_left_border
                sheet[f'{amount_column}{pt_row}'].border = normal_middle_right_border

            for data in company_data:
                month_row = workbook_month_index[data['month']] 
                insert_row = month_row + company_month_count[data['month']]
                company_month_count[data['month']] += 1
                
                sheet[f'{invoice_column}{insert_row}'] = data['no']

                value = 0

                if contain_comma(str(data['amount'])):
                    data['amount'] = locale.atof(data['amount'])
    
                if is_3_decimal_places(data['amount']):
                    value =  locale.format_string("%.3f", data['amount'], grouping=True)
                else:
                    value = locale.format_string("%.2f", data['amount'], grouping=True)

                sheet[f'{amount_column}{insert_row}'] = value 
                
                current_row += 1
            sheet.merge_cells(f"{chr(current_column)}{amount_row}:{chr(current_column+1)}{amount_row}")
            merged_cell = sheet[f'{chr(current_column)}{amount_row}']
            merged_cell.font = normal_font
            
            if is_3_decimal_places(total_amount[company]):
                value = locale.format_string("%.3f", total_amount[company], grouping=True)
                # value =  "{:.3f}".format(total_amount[company]) 
            else:
                value = locale.format_string("%.2f", total_amount[company], grouping=True)
                # value = "{:.2f}".format(total_amount[company])
            merged_cell.value = value

        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in sheet[amount_row]:
            cell.border = whole_thick_border


        output_file = self.export_file + f"_{self.file_num}.xlsx"
        workbook.save(output_file)

        self.file_num += 1

    def show_message_box(self,title,content):
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle(title)
        message_box.setText(content + "       ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )
    
        message_box.exec_()
    
class ExportExpensesDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    export_lorry_expenses = False
    export_creditor_expenses = False
    
    def __init__(self,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\export_expenses_dialog.ui', self)
        self.setWindowTitle("Export Excel")
        self.setMinimumSize(1700, 850)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()
    
    def init_ui(self):
        font = QFont("Tw Cen MT Condensed Extra Bold",19)
        font.setUnderline(True)
    
        self.title_label.resize(self.label_width,self.label_height)
        self.title_label.setFont(self.l_font)
        self.title_label.move(100,50)

        self.filename_label.resize(self.label_width,self.label_height)
        self.filename_label.setFont(self.s_font)
        self.filename_label.move(100,self.label_height + 60)

        self.filename_line.resize(self.line_width,self.line_height)
        self.filename_line.setFont(self.s_font)
        self.filename_line.move(100,self.label_height * 2  + 55)
        self.filename_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        
        self.year_label.resize(self.label_width,self.label_height)
        self.year_label.setFont(self.s_font)
        self.year_label.move(500,self.label_height + 60)

        self.year_spinbox.resize(self.line_width,self.line_height)
        self.year_spinbox.move(500,self.label_height*2 + 55)
        self.year_spinbox.setMinimum(2022)
        self.year_spinbox.setMaximum(2030)
        self.year_spinbox.setValue(2023)
        self.year_spinbox.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.year_spinbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

        self.hori_line.resize(1100,10)
        self.hori_line.move(50,self.label_height * 3 + 50)

        self.vertical_line.resize(10,400)
        self.vertical_line.move(850,self.label_height * 3 + 80)
        
        self.lorry_title_label.resize(self.label_width + 50,self.label_height)
        self.lorry_title_label.setFont(font)
        self.lorry_title_label.move(100,self.label_height * 3  + 65)

        self.lorry_no_label.resize(self.label_width,self.label_height)
        self.lorry_no_label.setFont(self.s_font)
        self.lorry_no_label.move(100,self.label_height * 4  + 65)

        self.creditor_label.resize(self.label_width + 100,self.label_height)
        self.creditor_label.setFont(font)
        self.creditor_label.move(900,self.label_height * 3  + 65)

        self.company_label.resize(self.label_width,self.label_height)
        self.company_label.setFont(self.s_font)
        self.company_label.move(900,self.label_height * 4  + 65)

        self.creditor_lorry_label.resize(self.label_width,self.label_height)
        self.creditor_lorry_label.setFont(self.s_font)
        self.creditor_lorry_label.move(1300,self.label_height * 4  + 65)

        self.lorry_combobox.resize(self.line_width,self.line_height)
        self.lorry_combobox.move(100,self.label_height * 5  + 65)
        self.setup_combo_box(self.lorry_combobox)
        self.lorry_data(self.lorry_combobox)
    
        self.company_combobox.resize(self.line_width,self.line_height)
        self.company_combobox.move(900,self.label_height * 5  + 65)
        self.setup_combo_box(self.company_combobox)
        self.expenses_company_data(self.company_combobox)

        self.creditor_lorry_combobox.resize(self.line_width,self.line_height)
        self.creditor_lorry_combobox.move(1300,self.label_height * 5  + 65)
        self.setup_combo_box(self.creditor_lorry_combobox)
        self.lorry_data(self.creditor_lorry_combobox)

        self.expenses_label.resize(self.label_width,self.label_height)
        self.expenses_label.setFont(self.s_font)
        self.expenses_label.move(480,self.label_height * 4  )

        self.expenses_scroll.setFixedSize(self.line_width,220)
        self.expenses_scroll.move(480,self.label_height * 5 )
        self.expenses_scroll.setWidgetResizable(True)
        self.expenses_scroll.setStyleSheet(
            "QScrollArea{"
            "    background-color:white;border: 2px solid black; color: black;"
            "}"
            "QScrollArea QScrollBar::handle:vertical{"
            "    background-color:white;"
            "}"
        )   
        scroll_layout = QVBoxLayout()
        self.expenses_scroll_widget.setLayout(scroll_layout)
        self.expenses_data()
        
        self.lorry_export_button.resize(270,70)
        self.lorry_export_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.lorry_export_button.move(300,self.label_height * 8 + 40)
        self.lorry_export_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.lorry_export_button.clicked.connect(self.export_lorry_button_clicked)

        self.company_export_button.resize(260,70)
        self.company_export_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.company_export_button.move(910,self.label_height * 6 + 65)
        self.company_export_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.company_export_button.clicked.connect(self.creditor_by_company_button_clicked)

        self.by_lorry_button.resize(260,70)
        self.by_lorry_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.by_lorry_button.move(1310,self.label_height * 6 + 65)
        self.by_lorry_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.by_lorry_button.clicked.connect(self.creditor_by_lorry_button_clicked)
    
        self.by_expenses_button.resize(260,70)
        self.by_expenses_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.by_expenses_button.move(1130,self.label_height * 6 + 180)
        self.by_expenses_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.by_expenses_button.clicked.connect(self.creditor_by_expenses_button_clicked)
        
    def check_duplicate(self,folder):
        if len(self.export_file) == 0:
            self.show_message_box("Warning","Please enter filename")
            return False

        expenses_folder = os.path.join(output_folder, f'{folder}')
        os.makedirs(expenses_folder, exist_ok=True)
        self.new_folder = os.path.join(expenses_folder, self.export_file)

        if os.path.exists(self.new_folder):
            self.show_message_box("Duplicate Name","File exist, please enter another filename")
            print("File already exists. Please choose a different filename.")
            return False

        os.makedirs(self.new_folder, exist_ok=True)

        # self.export_file = os.path.join(new_folder, self.export_file)
        # print(self.export_file)
        
        return True

    def export_lorry_button_clicked(self):
        # 一页 一个月
        self.export_file = self.filename_line.text()
        self.export_year = self.year_spinbox.value()
        self.export_lorry = self.lorry_combobox.currentText()

        if self.check_duplicate("Lorry Expenses") == False:
            return

        self.export_file = os.path.join(self.new_folder, self.export_file)

        self.export_lorry_expenses()

        self.show_message_box("Export","Data Exported")

        self.close()

    def creditor_by_company_button_clicked(self):
        self.export_file = self.filename_line.text()
        self.export_year = self.year_spinbox.value()
        self.export_company = self.company_combobox.currentText()

        if self.export_company == 'Not Chosen':
            self.show_message_box("Warning","Please chosoe a company")
            return 

        # self.export_file = "new_file"
        # self.export_year = 2023
        # self.export_company = "Megalink"

        if not self.check_duplicate("Expenses Creditor"):
            return
        
        self.export_by_company()

        self.show_message_box("Export","Data Exported")

        self.close()

    def creditor_by_lorry_button_clicked(self):
        self.export_file = self.filename_line.text()
        self.export_year = self.year_spinbox.value()
        self.export_lorry = self.creditor_lorry_combobox.currentText()

        if self.export_lorry == 'Not Chosen':
            self.show_message_box("Warning","Please chosoe a lorry")
            return 

        # self.export_file = "new_file"
        # self.export_year = 2023
        # self.export_company = "Megalink"

        if not self.check_duplicate("Expenses Creditor"):
            return
        
        self.export_by_lorry()

        self.show_message_box("Export","Data Exported")

        self.close()
    
    def creditor_by_expenses_button_clicked(self):
        self.export_file = self.filename_line.text()
        self.export_year = self.year_spinbox.value()

        # self.export_file = "new_file"
        # self.export_year = 2023
        # self.export_company = "Megalink"

        if not self.check_duplicate("Expenses Creditor"):
            return
        
        self.export_by_expenses()

        self.show_message_box("Export","Data Exported")

        self.close()
     
    def export_by_lorry(self):
        expenses_year_amount_list = {}
        month_list = {1:'Jan',2:'Feb',3:'Mac',4:'Apr',5:'May',6:'June',7:'July',8:'Aug',9:'Sept',10:'Oct',11:'Nov',12:'Dec'}

        page_margins = PageMargins(
            left=0.5,  # 左边距
            right=0.5,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距 
        )
        
        normal_font = Font(name='Times New Roman', size=11, bold=False, color='000000')
        bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
        underline_font = Font(name='Times New Roman', size=11, underline="single",bold=False, color='000000')
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        
        thick_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        
        header = ['Date','Company','Ref. No','Amount']

        # bookmark
        # expenses_type = []
        # query = QSqlQuery()
        # # calculate num of expenses type
        # query.exec_(f"SELECT expenses FROM Expenses_Record WHERE company = '{self.export_company}' GROUP BY expenses") 
        # while query.next():
        #     expenses_type.append(query.value(0))
        
        # for   


        query = QSqlQuery()
        # calculate num of expenses type
        query.exec_(f"SELECT expenses FROM Expenses_Record WHERE lorry = '{self.export_lorry}' AND company NOT LIKE '-' GROUP BY expenses")        
        expenses_type = []

        while query.next():
            expenses_type.append(query.value(0))

        for expenses in expenses_type:
            total_amount = 0

            output_file = os.path.join(self.new_folder, f'{self.export_file}_{expenses}.xlsx') 

            workbook = Workbook()
            sheet = workbook.active
            sheet.page_margins = page_margins

            current_row = 0

            # Company name
            sheet.merge_cells(f'A1:D2')
            merged_cell = sheet['A1']
            merged_cell.font = underline_font
            merged_cell.value = self.export_lorry

            sheet.merge_cells(f'A3:D3')
            merged_cell = sheet['A3']
            merged_cell.font = underline_font
            merged_cell.value = self.export_year

            sheet.merge_cells(f"A4:D4")
            merged_cell = sheet['A4']
            merged_cell.font = bold_font
            merged_cell.value = expenses

            sheet.append(header)

            current_row = 6
            start_row = 6

            selected_row = sheet[5]

            for cell in selected_row:
                cell.font = normal_font
            
            for row in sheet.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for i in range(1,13):
                start_row = current_row
                month_amount = 0
                duplicate_date = {}
                month_data_exist = 1

                if i < 10:
                    month = f"0{i}"
                else:
                    month = f"{i}"

                query_str = f"SELECT * FROM Expenses_Record WHERE date LIKE '%/{month}/{self.export_year}' AND lorry = '{self.export_lorry}' AND expenses = '{expenses}' AND company NOT LIKE '-' ORDER BY date"

                query.exec_(query_str)    
                # id , date , lorry , reference , expenses , company , amount

                while query.next():
                    if month_data_exist:
                        sheet.merge_cells(f"A{current_row}:D{current_row}")
                        merged_cell = sheet[f"A{current_row}"]
                        merged_cell.font = bold_font
                        merged_cell.value = month_list[i]

                        for cell in sheet[current_row]:
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                        current_row += 1 

                        month_data_exist = 0

                    date = query.value(1)
                    amount = 0
                    
                    if contain_comma(str(query.value(6))):
                        amount = locale.atof(str(query.value(6)))
                    else:
                        amount = query.value(6)
                        # print(amount)

                    
                    if date not in duplicate_date:
                        duplicate_date[f'{date}'] = current_row
    

                        row_data = [query.value(1),query.value(5),query.value(3),"{:.2f}".format(amount)]

                        sheet.append(row_data)
                        
                    else:
                        row_data = [query.value(1),query.value(5),query.value(3),"{:.2f}".format(amount)]
                        sheet.append(row_data)
                        if current_row - duplicate_date[f'{date}'] > 1: 
                            sheet.unmerge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row-1}')
                        sheet.merge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row}')   
                        merged_cell = sheet[f'A{duplicate_date[f"{date}"]}']                
                        merged_cell.value = date
                    
                    month_amount += amount
                    current_row += 1
                
                total_amount += month_amount

                if month_data_exist:
                    continue

                sheet.merge_cells(f'A{current_row}:C{current_row}')
                merged_cell = sheet[f'A{current_row}']
                merged_cell.font = bold_font
                merged_cell.value = "Total"

                sheet[f'D{current_row}'] = "{:.2f}".format(month_amount)

                for row in sheet.iter_rows(min_row = start_row + 1, max_row=current_row):
                    for cell in row:
                        cell.border = border
                        cell.font = normal_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                for cell in sheet[current_row]:
                    cell.border = thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += 1

                # for cell in sheet[start_row]:
                #     cell.border = thick_border
                #     cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1

                for idx, column_cells in enumerate(sheet.columns, start=1):
                    max_length = 0
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = max_length + 2   # 加上一些额外宽度
                    column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
                    if column_letter == 'A':
                        adjusted_width = 12
                    sheet.column_dimensions[column_letter].width = adjusted_width

            sheet.merge_cells(f'A{current_row}:C{current_row}')
            merged_cell = sheet[f'A{current_row}']
            merged_cell.font = bold_font
            merged_cell.value = " Year Total"

            sheet[f'D{current_row}'] = "{:.2f}".format(total_amount)

            for cell in sheet[current_row]:
                cell.border = thick_border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            
            expenses_year_amount_list[f'{expenses}'] = total_amount
                
            workbook.save(output_file)
            
        
        workbook = Workbook()
        sheet = workbook.active

        sheet.page_margins = page_margins
        # Company name
        sheet.merge_cells(f'A1:B2')
        merged_cell = sheet['A1']
        merged_cell.font = underline_font
        merged_cell.value = self.export_lorry

        sheet.merge_cells(f'A3:B3')
        merged_cell = sheet['A3']
        merged_cell.font = underline_font
        merged_cell.value = self.export_year

        header = ['Expenses','Amount']
        sheet.append(header)

        current_row = 5
        year_amount = 0
        month = 1
        
        for key,value in expenses_year_amount_list.items():
            row_data = [f'{key}',"{:.2f}".format(value)]
            sheet.append(row_data)
            current_row += 1
            month += 1
            year_amount += value

        sheet[f'A{current_row}'] = "Total"
        sheet[f'B{current_row}'] = "{:.2f}".format(year_amount)
    
        for row in sheet.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[4]:
            cell.border = thick_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
            for cell in row:
                cell.border = border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[current_row]:
            cell.border = thick_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2   # 加上一些额外宽度
            column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
            if column_letter == 'A':
                adjusted_width = 12
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        output_file = os.path.join(self.new_folder, f'{self.export_file}_Year.xlsx') 

        workbook.save(output_file)



        # 

        # for i in range(1,13):
        #     start_row = current_row
        #     month_amount = 0
        #     duplicate_date = {}

        #     if i < 10:
        #         month = f"0{i}"
        #     else:
        #         month = f"{i}"

        #     if i in [1,4,7,10]: # new workbook
        #         output_file = self.export_file + f"_{month_list[i]} - {month_list[i+2]}.xlsx"

        #         workbook = Workbook()
        #         sheet = workbook.active
        #         sheet.page_margins = page_margins

        #         # Company name
        #         sheet.merge_cells(f'A1:D2')
        #         merged_cell = sheet['A1']
        #         merged_cell.font = underline_font
        #         merged_cell.value = self.export_company

        #         sheet.merge_cells(f"A3:D3")
        #         merged_cell = sheet['A3']
        #         merged_cell.font = bold_font
        #         merged_cell.value = expenses_type

        #         sheet.append(header)

        #         current_row = 5
        #         start_row = 5

        #         selected_row = sheet[4]
        #         for cell in selected_row:
        #             cell.font = normal_font
                
        #         for row in sheet.iter_rows(min_row=1, max_row=5):
        #             for cell in row:
        #                 cell.border = border
        #                 cell.alignment = Alignment(horizontal='center', vertical='center')

        #     # month
        #     sheet.merge_cells(f"A{current_row}:D{current_row}")
        #     merged_cell = sheet[f"A{current_row}"]
        #     merged_cell.font = bold_font
        #     merged_cell.value = month_list[i]

        #     for cell in sheet[current_row]:
        #         cell.border = border
        #         cell.alignment = Alignment(horizontal='center', vertical='center')

        #     current_row += 1 

        #     query_str = f"SELECT * FROM Expenses_Record WHERE date LIKE '%/{month}/{self.export_year}' AND company = '{self.export_company}' ORDER BY date"
        #     query.exec_(query_str)    
        #     # id , date , lorry , reference , expenses , company , amount

        #     while query.next():
        #         date = query.value(1)
        #         if date not in duplicate_date:
        #             duplicate_date[f'{date}'] = current_row
        #             row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(query.value(6))]
        #             sheet.append(row_data)
        #         else:
        #             row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(query.value(6))]
        #             sheet.append(row_data)
        #             if current_row - duplicate_date[f'{date}'] > 1: 
        #                 sheet.unmerge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row-1}')
        #             sheet.merge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row}')   
        #             merged_cell = sheet[f'A{duplicate_date[f"{date}"]}']                
        #             merged_cell.value = date
                
        #         month_amount += query.value(6)
        #         current_row += 1

        #     sheet.merge_cells(f'A{current_row}:C{current_row}')
        #     merged_cell = sheet[f'A{current_row}']
        #     merged_cell.font = bold_font
        #     merged_cell.value = "Total"

        #     sheet[f'D{current_row}'] = "{:.2f}".format(month_amount)

        #     for row in sheet.iter_rows(min_row = start_row + 1, max_row=current_row):
        #         for cell in row:
        #             cell.border = border
        #             cell.font = normal_font
        #             cell.alignment = Alignment(horizontal='center', vertical='center')

        #     for cell in sheet[current_row]:
        #         cell.border = thick_border
        #         cell.alignment = Alignment(horizontal='center', vertical='center')
            
        #     current_row += 1

        #     # for cell in sheet[start_row]:
        #     #     cell.border = thick_border
        #     #     cell.alignment = Alignment(horizontal='center', vertical='center')

        #     current_row += 1

        #     for idx, column_cells in enumerate(sheet.columns, start=1):
        #         max_length = 0
        #         for cell in column_cells:
        #             try:
        #                 if len(str(cell.value)) > max_length:
        #                     max_length = len(cell.value)
        #             except:
        #                 pass
        #         adjusted_width = max_length + 2   # 加上一些额外宽度
        #         column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
        #         if column_letter == 'A':
        #             adjusted_width = 12
        #         sheet.column_dimensions[column_letter].width = adjusted_width

        #     year_amount_list[f'{month_list[i]}'] = month_amount

        #     if i % 3 == 0:
        #         workbook.save(output_file)

    def export_by_expenses(self):
        # BOOKMARK
        expenses_year_amount_list = {}  # all month need to recor
        month_list = {1:'Jan',2:'Feb',3:'Mac',4:'Apr',5:'May',6:'June',7:'July',8:'Aug',9:'Sept',10:'Oct',11:'Nov',12:'Dec'}

        page_margins = PageMargins(
            left=0.5,  # 左边距
            right=0.5,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距 
        )
        
        normal_font = Font(name='Times New Roman', size=11, bold=False, color='000000')
        bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
        underline_font = Font(name='Times New Roman', size=11, underline="single",bold=False, color='000000')
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        
        thick_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        
        header = ['Date','Ref. No','Company','Lorry','Amount']

        # bookmark
        # expenses_type = []
        # query = QSqlQuery()
        # # calculate num of expenses type
        # query.exec_(f"SELECT expenses FROM Expenses_Record WHERE company = '{self.export_company}' GROUP BY expenses") 
        # while query.next():
        #     expenses_type.append(query.value(0))
        
        # for   


        query = QSqlQuery()
        # calculate num of expenses type
        query.exec_(f"SELECT expenses FROM Expenses_Record WHERE company NOT LIKE '-' GROUP BY expenses")        
        expenses_type = []

        while query.next():
            expenses_type.append(query.value(0))

        for expenses in expenses_type:
            year_amount_list = {}
            total_amount = 0
            
            output_file = os.path.join(self.new_folder, f'{self.export_file}_{expenses}.xlsx') 

            workbook = Workbook()
            sheet = workbook.active
            sheet.page_margins = page_margins

            current_row = 0

            # expenses name
            sheet.merge_cells(f'A1:E2')
            merged_cell = sheet['A1']
            merged_cell.font = underline_font
            merged_cell.value = expenses

            sheet.merge_cells(f"A3:E3")
            merged_cell = sheet['A3']
            merged_cell.font = bold_font
            merged_cell.value = self.export_year

            sheet.append(header)

            current_row = 5
            start_row = 5

            selected_row = sheet[4]

            for cell in selected_row:
                cell.font = normal_font
            
            for row in sheet.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for i in range(1,13):
                start_row = current_row
                month_amount = 0
                duplicate_date = {}
                month_data_exist = 1

                if i < 10:
                    month = f"0{i}"
                else:
                    month = f"{i}"

                query_str = f"SELECT * FROM Expenses_Record WHERE date LIKE '%/{month}/{self.export_year}' AND company NOT LIKE '-' AND expenses = '{expenses}' ORDER BY date"

                query.exec_(query_str)    
                # id , date , lorry , reference , expenses , company , amount

                while query.next():
                    if month_data_exist:
                        sheet.merge_cells(f"A{current_row}:D{current_row}")
                        merged_cell = sheet[f"A{current_row}"]
                        merged_cell.font = bold_font
                        merged_cell.value = month_list[i]

                        for cell in sheet[current_row]:
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                        current_row += 1 

                        month_data_exist = 0

                    date = query.value(1)
                    amount = 0
                    
                    if contain_comma(str(query.value(6))):
                        amount = locale.atof(str(query.value(6)))
                    else:
                        amount = query.value(6)
                        # print(amount)

                    
                    if date not in duplicate_date:
                        duplicate_date[f'{date}'] = current_row
    

                        row_data = [query.value(1),query.value(3),query.value(5),query.value(2),"{:.2f}".format(amount)]

                        sheet.append(row_data)
                        
                    else:
                        row_data = [query.value(1),query.value(3),query.value(5),query.value(2),"{:.2f}".format(amount)]
                        sheet.append(row_data)
                        if current_row - duplicate_date[f'{date}'] > 1: 
                            sheet.unmerge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row-1}')
                        sheet.merge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row}')   
                        merged_cell = sheet[f'A{duplicate_date[f"{date}"]}']                
                        merged_cell.value = date
                    
                    month_amount += amount
                    current_row += 1
                
                year_amount_list[f'{month_list[i]}'] = month_amount
                total_amount += month_amount

                if month_data_exist:
                    continue

                sheet.merge_cells(f'A{current_row}:D{current_row}')
                merged_cell = sheet[f'A{current_row}']
                merged_cell.font = bold_font
                merged_cell.value = "Total"

                sheet[f'E{current_row}'] = "{:.2f}".format(month_amount)

                for row in sheet.iter_rows(min_row = start_row + 1, max_row=current_row):
                    for cell in row:
                        cell.border = border
                        cell.font = normal_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                for cell in sheet[current_row]:
                    cell.border = thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                

                current_row += 1

                # for cell in sheet[start_row]:
                #     cell.border = thick_border
                #     cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1

                for idx, column_cells in enumerate(sheet.columns, start=1):
                    max_length = 0
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = max_length + 2   # 加上一些额外宽度
                    column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
                    if column_letter == 'A':
                        adjusted_width = 12
                    sheet.column_dimensions[column_letter].width = adjusted_width
            

            sheet.merge_cells(f'A{current_row}:D{current_row}')
            merged_cell = sheet[f'A{current_row}']
            merged_cell.font = bold_font
            merged_cell.value = " Year Total"

            sheet[f'E{current_row}'] = "{:.2f}".format(total_amount)

            for cell in sheet[current_row]:
                cell.border = thick_border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            workbook.save(output_file)

            expenses_year_amount_list[f'{expenses}'] = year_amount_list
        


        workbook = Workbook()
        sheet = workbook.active

        sheet.page_margins = page_margins

        all_expenses_amount = {}

        for expenses in expenses_type:
            all_expenses_amount[f'{expenses}'] = 0

        # header
        max_col = chr(ord('A') + len(expenses_type))
        sheet.merge_cells(f'A1:{max_col}2')
        merged_cell = sheet['A1']
        merged_cell.font = underline_font
        merged_cell.value = self.export_year

        sheet.append([])
        expenses_type.insert(0, "")
        sheet.append(expenses_type)

        selected_row = sheet[3]

        for cell in selected_row:
            cell.font = normal_font
        
        for row in sheet.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')


        current_row = 4
        start_row = current_row
        month = 1
        row_data = []

        for i in range(1,13):
            row_data = []
            row_data.append(f'{month_list[i]}')
            for key,value in expenses_year_amount_list.items():
                amount = value[f'{month_list[i]}']
                row_data.append("{:.2f}".format(amount))
                all_expenses_amount[f'{key}'] += amount
            
            sheet.append(row_data)
            current_row += 1

        row_data = ["Total"]
        
        for value in all_expenses_amount.values():
            row_data.append("{:.2f}".format(value))

        sheet.append(row_data)

        for row in sheet.iter_rows(min_row = start_row, max_row=current_row):
            for cell in row:
                cell.border = border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[current_row]:
            cell.border = thick_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        

        current_row += 1

        # for cell in sheet[start_row]:
        #     cell.border = thick_border
        #     cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        for idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2   # 加上一些额外宽度
            column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
            if column_letter == 'A':
                adjusted_width = 12
            sheet.column_dimensions[column_letter].width = adjusted_width
        

        

        # sheet[f'A{current_row}'] = "Total"
        # sheet[f'B{current_row}'] = "{:.2f}".format(year_amount)
    
        # for row in sheet.iter_rows(min_row=1, max_row=3):
        #     for cell in row:
        #         cell.border = border
        #         cell.alignment = Alignment(horizontal='center', vertical='center')

        # for cell in sheet[4]:
        #     cell.border = thick_border
        #     cell.font = normal_font
        #     cell.alignment = Alignment(horizontal='center', vertical='center')

        # for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
        #     for cell in row:
        #         cell.border = border
        #         cell.font = normal_font
        #         cell.alignment = Alignment(horizontal='center', vertical='center')

        # for cell in sheet[current_row]:
        #     cell.border = thick_border
        #     cell.font = normal_font
        #     cell.alignment = Alignment(horizontal='center', vertical='center')

        # for idx, column_cells in enumerate(sheet.columns, start=1):
        #     max_length = 0
        #     for cell in column_cells:
        #         try:
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = max_length + 2   # 加上一些额外宽度
        #     column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
        #     if column_letter == 'A':
        #         adjusted_width = 12
        #     sheet.column_dimensions[column_letter].width = adjusted_width
        
        output_file = os.path.join(self.new_folder, f'{self.export_file}_All_Expenses.xlsx') 

        workbook.save(output_file)



    

        # 

        # for i in range(1,13):
        #     start_row = current_row
        #     month_amount = 0
        #     duplicate_date = {}

        #     if i < 10:
        #         month = f"0{i}"
        #     else:
        #         month = f"{i}"

        #     if i in [1,4,7,10]: # new workbook
        #         output_file = self.export_file + f"_{month_list[i]} - {month_list[i+2]}.xlsx"

        #         workbook = Workbook()
        #         sheet = workbook.active
        #         sheet.page_margins = page_margins

        #         # Company name
        #         sheet.merge_cells(f'A1:D2')
        #         merged_cell = sheet['A1']
        #         merged_cell.font = underline_font
        #         merged_cell.value = self.export_company

        #         sheet.merge_cells(f"A3:D3")
        #         merged_cell = sheet['A3']
        #         merged_cell.font = bold_font
        #         merged_cell.value = expenses_type

        #         sheet.append(header)

        #         current_row = 5
        #         start_row = 5

        #         selected_row = sheet[4]
        #         for cell in selected_row:
        #             cell.font = normal_font
                
        #         for row in sheet.iter_rows(min_row=1, max_row=5):
        #             for cell in row:
        #                 cell.border = border
        #                 cell.alignment = Alignment(horizontal='center', vertical='center')

        #     # month
        #     sheet.merge_cells(f"A{current_row}:D{current_row}")
        #     merged_cell = sheet[f"A{current_row}"]
        #     merged_cell.font = bold_font
        #     merged_cell.value = month_list[i]

        #     for cell in sheet[current_row]:
        #         cell.border = border
        #         cell.alignment = Alignment(horizontal='center', vertical='center')

        #     current_row += 1 

        #     query_str = f"SELECT * FROM Expenses_Record WHERE date LIKE '%/{month}/{self.export_year}' AND company = '{self.export_company}' ORDER BY date"
        #     query.exec_(query_str)    
        #     # id , date , lorry , reference , expenses , company , amount

        #     while query.next():
        #         date = query.value(1)
        #         if date not in duplicate_date:
        #             duplicate_date[f'{date}'] = current_row
        #             row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(query.value(6))]
        #             sheet.append(row_data)
        #         else:
        #             row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(query.value(6))]
        #             sheet.append(row_data)
        #             if current_row - duplicate_date[f'{date}'] > 1: 
        #                 sheet.unmerge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row-1}')
        #             sheet.merge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row}')   
        #             merged_cell = sheet[f'A{duplicate_date[f"{date}"]}']                
        #             merged_cell.value = date
                
        #         month_amount += query.value(6)
        #         current_row += 1

        #     sheet.merge_cells(f'A{current_row}:C{current_row}')
        #     merged_cell = sheet[f'A{current_row}']
        #     merged_cell.font = bold_font
        #     merged_cell.value = "Total"

        #     sheet[f'D{current_row}'] = "{:.2f}".format(month_amount)

        #     for row in sheet.iter_rows(min_row = start_row + 1, max_row=current_row):
        #         for cell in row:
        #             cell.border = border
        #             cell.font = normal_font
        #             cell.alignment = Alignment(horizontal='center', vertical='center')

        #     for cell in sheet[current_row]:
        #         cell.border = thick_border
        #         cell.alignment = Alignment(horizontal='center', vertical='center')
            
        #     current_row += 1

        #     # for cell in sheet[start_row]:
        #     #     cell.border = thick_border
        #     #     cell.alignment = Alignment(horizontal='center', vertical='center')

        #     current_row += 1

        #     for idx, column_cells in enumerate(sheet.columns, start=1):
        #         max_length = 0
        #         for cell in column_cells:
        #             try:
        #                 if len(str(cell.value)) > max_length:
        #                     max_length = len(cell.value)
        #             except:
        #                 pass
        #         adjusted_width = max_length + 2   # 加上一些额外宽度
        #         column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
        #         if column_letter == 'A':
        #             adjusted_width = 12
        #         sheet.column_dimensions[column_letter].width = adjusted_width

        #     year_amount_list[f'{month_list[i]}'] = month_amount

        #     if i % 3 == 0:
        #         workbook.save(output_file)

    def export_by_company(self):
        month_list = {1:'Jan',2:'Feb',3:'Mac',4:'Apr',5:'May',6:'June',7:'July',8:'Aug',9:'Sept',10:'Oct',11:'Nov',12:'Dec'}

        page_margins = PageMargins(
            left=0.5,  # 左边距
            right=0.5,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距 
        )
        
        normal_font = Font(name='Times New Roman', size=11, bold=False, color='000000')
        bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
        underline_font = Font(name='Times New Roman', size=11, underline="single",bold=False, color='000000')
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        
        thick_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        
        header = ['Date','Ref. No','Lorry','Amount']

        # bookmark
        # expenses_type = []
        # query = QSqlQuery()
        # # calculate num of expenses type
        # query.exec_(f"SELECT expenses FROM Expenses_Record WHERE company = '{self.export_company}' GROUP BY expenses") 
        # while query.next():
        #     expenses_type.append(query.value(0))
        
        # for   


        query = QSqlQuery()
        # calculate num of expenses type
        query.exec_(f"SELECT expenses FROM Expenses_Record WHERE company = '{self.export_company}' GROUP BY expenses")        
        expenses_type = []

        while query.next():
            expenses_type.append(query.value(0))

        for expenses in expenses_type:
            year_amount_list = {}

            query.exec_(f"SELECT * FROM Expenses_Record WHERE company = '{self.export_company}' AND expenses = '{expenses}'")

            
            expenses_type_folder = os.path.join(self.new_folder, f'{self.export_file}_{expenses}') 
            os.makedirs(expenses_type_folder, exist_ok=True)

            output_file = os.path.join(expenses_type_folder, f'{self.export_file}_{expenses}_Month.xlsx') 

            workbook = Workbook()
            sheet = workbook.active
            sheet.page_margins = page_margins

            current_row = 0

            # Company name
            sheet.merge_cells(f'A1:D2')
            merged_cell = sheet['A1']
            merged_cell.font = underline_font
            merged_cell.value = self.export_company

            sheet.merge_cells(f"A3:D3")
            merged_cell = sheet['A3']
            merged_cell.font = bold_font
            merged_cell.value = expenses

            sheet.append(header)

            current_row = 5
            start_row = 5

            selected_row = sheet[4]

            for cell in selected_row:
                cell.font = normal_font
            
            for row in sheet.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for i in range(1,13):
                start_row = current_row
                month_amount = 0
                duplicate_date = {}
                month_data_exist = 1

                if i < 10:
                    month = f"0{i}"
                else:
                    month = f"{i}"

                query_str = f"SELECT * FROM Expenses_Record WHERE date LIKE '%/{month}/{self.export_year}' AND company = '{self.export_company}' AND expenses = '{expenses}' ORDER BY date"

                query.exec_(query_str)    
                # id , date , lorry , reference , expenses , company , amount

                while query.next():
                    if month_data_exist:
                        sheet.merge_cells(f"A{current_row}:D{current_row}")
                        merged_cell = sheet[f"A{current_row}"]
                        merged_cell.font = bold_font
                        merged_cell.value = month_list[i]

                        for cell in sheet[current_row]:
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                        current_row += 1 

                        month_data_exist = 0

                    date = query.value(1)
                    amount = 0
                    
                    if contain_comma(str(query.value(6))):
                        amount = locale.atof(str(query.value(6)))
                    else:
                        amount = query.value(6)
                        # print(amount)

                    
                    if date not in duplicate_date:
                        duplicate_date[f'{date}'] = current_row

                        row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(amount)]

                        sheet.append(row_data)
                        
                    else:
                        row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(amount)]
                        sheet.append(row_data)
                        if current_row - duplicate_date[f'{date}'] > 1: 
                            sheet.unmerge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row-1}')
                        sheet.merge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row}')   
                        merged_cell = sheet[f'A{duplicate_date[f"{date}"]}']                
                        merged_cell.value = date
                    
                    month_amount += amount
                    current_row += 1
                
                year_amount_list[f'{month_list[i]}'] = month_amount

                if month_data_exist:
                    continue

                sheet.merge_cells(f'A{current_row}:C{current_row}')
                merged_cell = sheet[f'A{current_row}']
                merged_cell.font = bold_font
                merged_cell.value = "Total"

                sheet[f'D{current_row}'] = "{:.2f}".format(month_amount)

                for row in sheet.iter_rows(min_row = start_row + 1, max_row=current_row):
                    for cell in row:
                        cell.border = border
                        cell.font = normal_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                for cell in sheet[current_row]:
                    cell.border = thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += 1

                # for cell in sheet[start_row]:
                #     cell.border = thick_border
                #     cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1

                for idx, column_cells in enumerate(sheet.columns, start=1):
                    max_length = 0
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = max_length + 2   # 加上一些额外宽度
                    column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
                    if column_letter == 'A':
                        adjusted_width = 12
                    sheet.column_dimensions[column_letter].width = adjusted_width

                
            workbook.save(output_file)

            workbook = Workbook()
            sheet = workbook.active

            sheet.page_margins = page_margins
            # Company name
            sheet.merge_cells(f'A1:B2')
            merged_cell = sheet['A1']
            merged_cell.font = underline_font
            merged_cell.value = self.export_company

            sheet.merge_cells(f"A3:B3")
            merged_cell = sheet['A3']
            merged_cell.font = bold_font
            merged_cell.value = expenses

            header = ['Month','Amount']
            sheet.append(header)

            current_row = 5
            year_amount = 0
            month = 1
            
            for value in year_amount_list.values():
                row_data = [f'{month_list[month]}',"{:.2f}".format(value)]
                sheet.append(row_data)
                current_row += 1
                month += 1
                year_amount += value

            sheet[f'A{current_row}'] = "Total"
            sheet[f'B{current_row}'] = "{:.2f}".format(year_amount)
        
            for row in sheet.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for cell in sheet[4]:
                cell.border = thick_border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
                for cell in row:
                    cell.border = border
                    cell.font = normal_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for cell in sheet[current_row]:
                cell.border = thick_border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, column_cells in enumerate(sheet.columns, start=1):
                max_length = 0
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2   # 加上一些额外宽度
                column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
                if column_letter == 'A':
                    adjusted_width = 12
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            output_file = os.path.join(expenses_type_folder, f'{self.export_file}_{expenses}_Year.xlsx') 

            workbook.save(output_file)



    

        # 

        # for i in range(1,13):
        #     start_row = current_row
        #     month_amount = 0
        #     duplicate_date = {}

        #     if i < 10:
        #         month = f"0{i}"
        #     else:
        #         month = f"{i}"

        #     if i in [1,4,7,10]: # new workbook
        #         output_file = self.export_file + f"_{month_list[i]} - {month_list[i+2]}.xlsx"

        #         workbook = Workbook()
        #         sheet = workbook.active
        #         sheet.page_margins = page_margins

        #         # Company name
        #         sheet.merge_cells(f'A1:D2')
        #         merged_cell = sheet['A1']
        #         merged_cell.font = underline_font
        #         merged_cell.value = self.export_company

        #         sheet.merge_cells(f"A3:D3")
        #         merged_cell = sheet['A3']
        #         merged_cell.font = bold_font
        #         merged_cell.value = expenses_type

        #         sheet.append(header)

        #         current_row = 5
        #         start_row = 5

        #         selected_row = sheet[4]
        #         for cell in selected_row:
        #             cell.font = normal_font
                
        #         for row in sheet.iter_rows(min_row=1, max_row=5):
        #             for cell in row:
        #                 cell.border = border
        #                 cell.alignment = Alignment(horizontal='center', vertical='center')

        #     # month
        #     sheet.merge_cells(f"A{current_row}:D{current_row}")
        #     merged_cell = sheet[f"A{current_row}"]
        #     merged_cell.font = bold_font
        #     merged_cell.value = month_list[i]

        #     for cell in sheet[current_row]:
        #         cell.border = border
        #         cell.alignment = Alignment(horizontal='center', vertical='center')

        #     current_row += 1 

        #     query_str = f"SELECT * FROM Expenses_Record WHERE date LIKE '%/{month}/{self.export_year}' AND company = '{self.export_company}' ORDER BY date"
        #     query.exec_(query_str)    
        #     # id , date , lorry , reference , expenses , company , amount

        #     while query.next():
        #         date = query.value(1)
        #         if date not in duplicate_date:
        #             duplicate_date[f'{date}'] = current_row
        #             row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(query.value(6))]
        #             sheet.append(row_data)
        #         else:
        #             row_data = [query.value(1),query.value(3),query.value(2),"{:.2f}".format(query.value(6))]
        #             sheet.append(row_data)
        #             if current_row - duplicate_date[f'{date}'] > 1: 
        #                 sheet.unmerge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row-1}')
        #             sheet.merge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row}')   
        #             merged_cell = sheet[f'A{duplicate_date[f"{date}"]}']                
        #             merged_cell.value = date
                
        #         month_amount += query.value(6)
        #         current_row += 1

        #     sheet.merge_cells(f'A{current_row}:C{current_row}')
        #     merged_cell = sheet[f'A{current_row}']
        #     merged_cell.font = bold_font
        #     merged_cell.value = "Total"

        #     sheet[f'D{current_row}'] = "{:.2f}".format(month_amount)

        #     for row in sheet.iter_rows(min_row = start_row + 1, max_row=current_row):
        #         for cell in row:
        #             cell.border = border
        #             cell.font = normal_font
        #             cell.alignment = Alignment(horizontal='center', vertical='center')

        #     for cell in sheet[current_row]:
        #         cell.border = thick_border
        #         cell.alignment = Alignment(horizontal='center', vertical='center')
            
        #     current_row += 1

        #     # for cell in sheet[start_row]:
        #     #     cell.border = thick_border
        #     #     cell.alignment = Alignment(horizontal='center', vertical='center')

        #     current_row += 1

        #     for idx, column_cells in enumerate(sheet.columns, start=1):
        #         max_length = 0
        #         for cell in column_cells:
        #             try:
        #                 if len(str(cell.value)) > max_length:
        #                     max_length = len(cell.value)
        #             except:
        #                 pass
        #         adjusted_width = max_length + 2   # 加上一些额外宽度
        #         column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
        #         if column_letter == 'A':
        #             adjusted_width = 12
        #         sheet.column_dimensions[column_letter].width = adjusted_width

        #     year_amount_list[f'{month_list[i]}'] = month_amount

        #     if i % 3 == 0:
        #         workbook.save(output_file)

    def export_lorry_expenses(self):
        year_amount_list = {}
        total_amount = {}
        checked_boxes = []
        month_list = {1:'Jan',2:'Feb',3:'Mac',4:'Apr',5:'May',6:'June',7:'July',8:'Aug',9:'Sept',10:'Oct',11:'Nov',12:'Dec'}

        page_margins = PageMargins(
            left=0.5,  # 左边距
            right=0.5,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距 
        )
        
        normal_font = Font(name='Times New Roman', size=11, bold=False, color='000000')
        bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
        underline_font = Font(name='Times New Roman', size=11, underline="single",bold=False, color='000000')
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        
        thick_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        

        
        for checkbox in self.expenses_checkboxes:
            if checkbox.isChecked():
                checked_boxes.append(checkbox.text())
        
        tmp = 65 #start from B column
        header = []
        expenses_list = {}
        temp=""

        if len(checked_boxes) > 0:
            temp += " ( "
            first = True
            for expenses in checked_boxes:
                tmp += 1
                header.append(expenses)
                expenses_list[expenses] = tmp
                total_amount[expenses] = 0
                if not first:
                    temp += " OR "

                first = False
                temp += f"expenses = '{expenses}'"
                
                
            temp += " ) AND"
            
        
        column = chr(tmp)

        # header
        header.insert(0,'Date')

        for i in range(1,13):
            month_amount_list = {}
            duplicate_date = {}

            output_file = self.export_file + f"_{month_list[i]}.xlsx"

            if i < 10:
                month = f"0{i}"
            else:
                month = f"{i}"

            for key,value in expenses_list.items():
                month_amount_list[f'{key}'] = 0    

            workbook = Workbook()
            sheet = workbook.active
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            sheet.page_margins = page_margins

            # Lorry No
            sheet.merge_cells(f'A1:{column}2')
            merged_cell = sheet['A1']
            merged_cell.font = underline_font
            merged_cell.value = self.export_lorry

            # month
            sheet.merge_cells(f"A3:{column}3")
            merged_cell = sheet['A3']
            merged_cell.font = bold_font
            merged_cell.value = month_list[i]

            sheet.append(header)

            max_row = 1
            current_row = 4

            # 正式 sql 查询
            query = QSqlQuery()
            query_str = f"SELECT * FROM Expenses_Record WHERE + {temp} + date LIKE '%/{month}/{self.export_year}' AND lorry = '{self.export_lorry}' AND company = '-' AND reference = '-' ORDER BY date"
            query.exec_(query_str)    
            # id , date , lorry , reference , expenses , company , amount

            old_date = None

            while query.next():
                date = query.value(1)         
                if date not in duplicate_date:
                
                    current_row += 1
                    if old_date is not None:
                        current_row = duplicate_date[f'{old_date}'] + max_row
                    duplicate_date[f'{date}'] = current_row
                    current_date = {}
                    for key,value in expenses_list.items():
                        current_date[f'{key}'] = 0   
                    max_row = 1
                    # row_data = [query.value(1)]
                    sheet[f'A{current_row}'] = query.value(1)
                    # sheet.append(row_data)
                else:
                    old_date = date

                if current_date[query.value(4)] >= max_row:
                    current_row += 1
                    if current_row - duplicate_date[f'{date}'] > 1: 
                        sheet.unmerge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row-1}')
                    sheet.merge_cells(f'A{duplicate_date[f"{date}"]}:A{current_row}')   
                    merged_cell = sheet[f'A{duplicate_date[f"{date}"]}']                
                    merged_cell.value = date
                    max_row += 1
                else:
                    current_row = duplicate_date[f'{date}'] + current_date[query.value(4)]
                

                insert_column = expenses_list[f'{query.value(4)}']
                month_amount_list[f'{query.value(4)}'] += query.value(6)
                sheet[f'{chr(insert_column)}{current_row}'] = "{:.2f}".format(query.value(6))
                current_date[query.value(4)] += 1

            
            if old_date is not None:
                current_row = duplicate_date[f'{old_date}'] + max_row


            # current_row += 1    
            sheet[f'A{current_row}'] = "Total"
            sheet[f'A{current_row}'].font = bold_font

            start_column = 66
            for value in month_amount_list.values():
                print(value)
                sheet[f'{chr(start_column)}{current_row}'] = "{:.2f}".format(value)
                start_column += 1    
                     
            
            for row in sheet.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
                for cell in row:
                    cell.border = border
                    cell.font = normal_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for cell in sheet[current_row]:
                cell.border = thick_border        

            for idx, column_cells in enumerate(sheet.columns, start=1):
                max_length = 0
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2   # 加上一些额外宽度
                column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
                if column_letter == 'A':
                    adjusted_width = 12
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(output_file)
            year_amount_list[f'{month_list[i]}'] = month_amount_list

        workbook = Workbook()
        sheet = workbook.active
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

        sheet.page_margins = page_margins

        column = chr(ord(column))

        # Lorry No
        sheet.merge_cells(f'A1:{column}2')
        merged_cell = sheet['A1']
        merged_cell.font = underline_font
        merged_cell.value = self.export_lorry

        # month
        sheet.merge_cells(f"A3:{column}3")
        merged_cell = sheet['A3']
        merged_cell.font = bold_font
        merged_cell.value = self.export_year

        #header
        header.pop(0)
        header.insert(0,"Month")
        sheet.append(header)

        current_row = 5
        start_column = 66

        for key,month_list in year_amount_list.items():
            sheet[f'A{current_row}'] = key
            for key,value in month_list.items():
                insert_column = expenses_list[f'{key}']
                sheet[f'{chr(insert_column)}{current_row}'] = "{:.2f}".format(value)
                total_amount[f'{key}'] += value
            current_row += 1

        sheet[f'A{current_row}'] = "Total"

        for key,value in total_amount.items():
            insert_column = expenses_list[f'{key}']
            sheet[f'{chr(insert_column)}{current_row}'] = "{:.2f}".format(value)
        
        for row in sheet.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
            for cell in row:
                cell.border = border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[current_row]:
            cell.border = thick_border

        for idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2   # 加上一些额外宽度
            column_letter = chr(ord('A') + idx - 1)  # 根据索引计算列字母表示
            if column_letter == 'A':
                adjusted_width = 12
            sheet.column_dimensions[column_letter].width = adjusted_width
        workbook.save(self.export_file + ".xlsx")

    def show_message_box(self,title,content):
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle(title)
        message_box.setText(content + "       ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )
    
        message_box.exec_()
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def expenses_data(self):
        self.expenses_checkboxes = []
        scroll_layout = self.expenses_scroll_widget.layout()

        query = QSqlQuery("SELECT description FROM Expenses")
        while query.next():
            name = query.value(0)
            new_checkbox = QCheckBox(name)
            new_checkbox.setStyleSheet("background-color:white;")
            self.expenses_checkboxes.append(new_checkbox)
            new_checkbox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
            scroll_layout.addWidget(new_checkbox)  # 将新的 QCheckBox 添加到布局中

    def expenses_company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Expenses_Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def get_file(self):
        return self.export_file
    
    def get_company(self):
        return self.export_company

    def get_exported(self):
        return self.exported

class ExportSalaryDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    month_list = {1:'Jan',2:'Feb',3:'Mac',4:'Apr',5:'May',6:'June',7:'July',8:'Aug',9:'Sept',10:'Oct',11:'Nov',12:'Dec'}
    
    def __init__(self,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\export_salary_dialog.ui', self)
        self.setWindowTitle("Export Excel")
        self.setMinimumSize(900, 800)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()
    
    def init_ui(self):
        self.title_label.resize(self.label_width,self.label_height)
        self.title_label.setFont(self.l_font)
        self.title_label.move(280,40)

        move_x = 280
        move_y = 60

        self.filename_label.resize(self.label_width,self.label_height)
        self.filename_label.setFont(self.s_font)
        move_y += self.label_height - 5
        self.filename_label.move(move_x,move_y)

        self.filename_line.resize(self.line_width,self.line_height)
        self.filename_line.setFont(self.s_font)
        move_y += self.label_height - 15
        self.filename_line.move(move_x,move_y)
        self.filename_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
 
        self.employee_label.resize(self.label_width + 50,self.label_height)
        self.employee_label.setFont(self.s_font)
        move_y += self.line_height + 10
        self.employee_label.move(move_x,move_y)

        self.employee_combobox.resize(self.line_width+80,self.line_height)
        move_y += self.label_height - 15
        self.employee_combobox.move(move_x,move_y)
        self.setup_combo_box(self.employee_combobox)
        self.employee_data(self.employee_combobox)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        move_y += self.line_height + 10
        self.date_label.move(move_x,move_y)

        self.month.resize(self.line_width,self.line_height)
        move_y += self.label_height - 15
        self.month.move(move_x,move_y)
        self.month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.month.addItem("Not Chosen")
        self.month.addItems(months)
        self.month.setCurrentIndex(0)


        self.year.resize(self.line_width,self.line_height)
        move_y += self.line_height + 10
        self.year.move(move_x,move_y)
        self.year.setMinimum(2022)
        self.year.setMaximum(2030)
        self.year.setValue(2023)
        self.year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))


        self.export_month_button.resize(320,70)
        self.export_month_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_month_button.move(70,move_y + 100)
        self.export_month_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_month_button.clicked.connect(self.export_month_clicked)

        self.export_year_button.resize(320,70)
        self.export_year_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.export_year_button.move(470,move_y + 100)
        self.export_year_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.export_year_button.clicked.connect(self.export_year_clicked)

    def check_duplicate(self,export_year=False):
        self.export_file = self.filename_line.text()
        self.employee = self.employee_combobox.currentText()
        self.export_month = self.month.currentIndex()
        self.export_year = self.year.value()
        
        if len(self.export_file) == 0:
            self.show_message_box("Warning","Please enter filename")
            return False

        if self.employee == 'Not Chosen':
            self.show_message_box("Warning","Please choose employee")
            return False
        
        salary_folder = os.path.join(output_folder, 'Salary')
        os.makedirs(salary_folder, exist_ok=True)
        output_file = os.path.join(salary_folder, self.export_file)

        if os.path.exists(output_file):
            self.show_message_box("Duplicate Name","File exist, please enter another filename")
            print("File already exists. Please choose a different filename.")
            return False
        
        if export_year:
            os.makedirs(output_file, exist_ok=True)
            self.export_file = os.path.join(output_file, self.export_file)
        else:
            self.export_file = output_file + ".xlsx"

        return True

    def export_month_clicked(self,export_month=None,filename=None):
        if export_month == False:
            if not self.check_duplicate():
                return
            
            if self.export_month == 0:
                self.show_message_box("Warning","Please choose month")
                return
            
        
        print("after check month")
        
        month_total = 0
        page_margins = PageMargins(
            left=0.5,  # 左边距
            right=0.5,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距 
        )
        
        normal_font = Font(name='Times New Roman', size=11, bold=False, color='000000')
        bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
        underline_font = Font(name='Times New Roman', size=11, underline="single",bold=True, color='000000')
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        thick_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        
        month = export_month


        if export_month == False:
            month = self.export_month

        if month < 10 :
            month = f'0{month}'
        else:
            month = f"{month}"

        print(month)

        employee_data = []
        query = QSqlQuery()
        query_str = f"SELECT * FROM Employee where name = '{self.employee}'"
        query.exec_(query_str)

        if query.next():
            for i in range(1,4):
                employee_data.append(query.value(i))

        
        query_str = f"SELECT * FROM Salary where employee = '{self.employee}' AND date like '%/{month}/{self.export_year}' ORDER BY date"
        print(query_str)
        query.exec_(query_str)

        
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.page_margins = page_margins

        # employee name
        sheet.merge_cells(f"A1:B1")
        merged_cell = sheet['A1']
        merged_cell.font = underline_font
        merged_cell.value = employee_data[0]

        for cell in sheet[1]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # employee ic
        sheet.merge_cells(f"A2:B2")
        merged_cell = sheet['A2']
        merged_cell.font = normal_font
        merged_cell.value = employee_data[1]
       

        # employee lorry
        sheet.merge_cells(f"A3:B3")
        merged_cell = sheet['A3']
        merged_cell.font = normal_font
        merged_cell.value = employee_data[2]

        print("here1")
        # month
        sheet.merge_cells(f"A4:B4")
        merged_cell = sheet['A4']
        merged_cell.font = normal_font
        
        if export_month == False:
            merged_cell.value = self.month_list[self.export_month]
        else:
            merged_cell.value = self.month_list[export_month]
        print("here1")

        header = ['Date','Amount']
        sheet.append(header)

        current_row = 6
        
        while query.next(): # 1 and 3
            row_data = [query.value(1),query.value(3)]
            month_total += query.value(3)
            sheet.append(row_data)
            current_row += 1

        
        sheet[f'A{current_row}'] = "Total"
        sheet[f'B{current_row}'] = month_total
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
        for cell in sheet[current_row]:
            cell.border = thick_border

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20

        print("here2")

        if filename != None:
            workbook.save(self.export_file + filename + ".xlsx")
        else:
            workbook.save(self.export_file)
        
            self.show_message_box("Export","Data Exported")
            self.close()
        
    def export_year_clicked(self):
        if not self.check_duplicate(export_year=True):
            return
        
        year_total = 0
        month_total = {}

        page_margins = PageMargins(
            left=0.5,  # 左边距
            right=0.5,  # 右边距
            top=0.75,  # 上边距
            bottom=0.75,  # 下边距
            header=0.3,  # 页眉边距
            footer=0.3  # 页脚边距 
        )
        
        normal_font = Font(name='Times New Roman', size=11, bold=False, color='000000')
        bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
        underline_font = Font(name='Times New Roman', size=11, underline="single",bold=True, color='000000')
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        thick_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))
        
        employee_data = []
        query = QSqlQuery()
        query_str = f"SELECT * FROM Employee where name = '{self.employee}'"
        query.exec_(query_str)

        if query.next():
            for i in range(1,4):
                employee_data.append(query.value(i))

        workbook = Workbook()
        sheet = workbook.active
        sheet.page_margins = page_margins
    
        # employee name
        sheet.merge_cells(f"A1:B1")
        merged_cell = sheet['A1']
        merged_cell.font = underline_font
        merged_cell.value = employee_data[0]

        for cell in sheet[1]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # employee ic
        sheet.merge_cells(f"A2:B2")
        merged_cell = sheet['A2']
        merged_cell.font = normal_font
        merged_cell.value = employee_data[1]

        # employee lorry
        sheet.merge_cells(f"A3:B3")
        merged_cell = sheet['A3']
        merged_cell.font = normal_font
        merged_cell.value = employee_data[2]

        # month
        sheet.merge_cells(f"A4:B4")
        merged_cell = sheet['A4']
        merged_cell.font = normal_font
        merged_cell.value = self.export_year

        header = ['Month','Amount']
        sheet.append(header)

        current_row = 6

        for i in range(1,13):
            if i < 10:
                month = f'0{i}'
            else:
                month = i

            query_str = f"SELECT SUM(amount) FROM Salary where employee = '{self.employee}' AND date like '%/{month}/{self.export_year}' ORDER BY date"

            query.exec_(query_str)

            if query.next():
                if query.value(0):
                    year_total += query.value(0)
                    if is_3_decimal_places(query.value(0)):
                        row_data = [self.month_list[i],"{:.3f}".format(query.value(0))]
                    else:
                        row_data = [self.month_list[i],"{:.2f}".format(query.value(0))]
                else:
                    row_data = [self.month_list[i],0]
                
                sheet.append(row_data)
                current_row += 1
        
        sheet[f'A{current_row}'] = "Total"
        sheet[f'B{current_row}'] = year_total

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = border
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
        for cell in sheet[current_row]:
            cell.border = thick_border

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20

        workbook.save(self.export_file + ".xlsx")

        for i in range(1,13):
            filename = "_" + self.month_list[i]
            self.export_month_clicked(export_month=i,filename=filename)
           
        
        self.show_message_box("Export","Data Exported")
        self.close()
    
    def show_message_box(self,title,content):
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle(title)
        message_box.setText(content + "       ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )
    
        message_box.exec_()
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def employee_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Employee")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
class UpdateRecordDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,record,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\update_record_dialog.ui', self)
        self.row_data = record
        self.setWindowTitle("Update Record")
        self.setMinimumSize(1000, 1250)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()

    def init_ui(self):
        print(self.row_data)
        self.workspace_frame.resize(400,1250)
        self.workspace_frame.move(300,40)
        
        self.add_record_label.resize(self.label_width + 200,self.label_height)
        self.add_record_label.setFont(self.l_font)
        self.add_record_label.move(0,0)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(0,80)
        move_y = 80


        date_time = QDate.fromString(self.row_data[1], "dd/MM/yyyy")
        self.date_line.setDate(date_time)
        self.date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.date_line.setDisplayFormat("dd/MM/yyyy")
        self.date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.date_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.date_line.move(0,move_y)
        self.date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:10px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 28px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white }"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:20px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 35px; height: 35px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 3px solid black; font-family: Constantia; font-size: 28px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:50px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:50px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_left.png); }"
        )
    
        self.lorry_label.resize(self.label_width,self.label_height)
        self.lorry_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.lorry_label.move(0,move_y)        

        self.lorry_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.lorry_line.move(0,move_y)
        self.setup_combo_box(self.lorry_line)
        self.lorry_data(self.lorry_line)
        self.lorry_line.setCurrentText(self.row_data[2])



        self.reference_label.resize(self.label_width,self.label_height)
        self.reference_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.reference_label.move(0,move_y)     

        self.reference_line.resize(self.line_width,self.line_height)
        self.reference_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.reference_line.move(0,move_y)
        self.reference_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.reference_line.setText(self.row_data[3])   
        
    
        self.company_label.resize(self.label_width,self.label_height)
        self.company_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.company_label.move(0,move_y)

        self.company_line.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.company_line.move(0,move_y)
        self.setup_combo_box(self.company_line)
        self.company_data(self.company_line)
        self.company_line.setCurrentText(self.row_data[4])

        self.description_label.resize(self.label_width + 80,self.label_height)
        self.description_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.description_label.move(0,move_y)

        self.description_line.resize(self.line_width + 80,self.line_height)
        self.description_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.description_line.move(0,move_y)
        self.description_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        self.description_line.setText(self.row_data[5])

        
        self.quantity_label.resize(self.label_width,self.label_height)
        self.quantity_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.quantity_label.move(0,move_y)

        self.quantity_line.resize(self.line_width,self.line_height)
        self.quantity_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.quantity_line.move(0,move_y)
        self.quantity_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        self.quantity_line.setText(self.row_data[6])

        
        self.unitprice_label.resize(self.label_width,self.label_height)
        self.unitprice_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.unitprice_label.move(0,move_y)

        self.unitprice_line.resize(self.line_width,self.line_height)
        self.unitprice_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.unitprice_line.move(0,move_y)
        self.unitprice_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")   
        self.unitprice_line.setText(self.row_data[7])
        
        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.amount_line.setText(str(self.row_data[8]))


        self.add_record_submit_button.resize(180,70)
        self.add_record_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_record_submit_button.move(50,move_y + 80)
        self.add_record_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.add_record_submit_button.clicked.connect(self.submit_button_clicked)
    
    def submit_button_clicked(self):
        print("here")
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
        
        print("here")
        query = QSqlQuery()
        query.prepare(f"UPDATE Record SET date = :date, lorry = :lorry, reference = :reference, company = :company,description = :description,quantity = :quantity, unit_price = :unit_price, amount = :amount WHERE id = {self.row_data[0]}")

        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        print("here2")
        query.bindValue(":lorry", set_not_chosen(self.lorry_line.currentText()))
        query.bindValue(":reference", self.reference_line.text())
        query.bindValue(":company", set_not_chosen(self.company_line.currentText()))
        print("here3")
        query.bindValue(":description", self.description_line.text())
        query.bindValue(":quantity", self.quantity_line.text())
        query.bindValue(":unit_price", self.unitprice_line.text())
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data updated successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Update")
        message_box.setText("Record Updated        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()
        self.updated = True
        self.close()

    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_updated(self):
        return self.updated

class UpdateInvoiceDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,record,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\update_invoice_dialog.ui', self)
        self.row_data = record
        self.setWindowTitle("Update Invoice")
        self.setMinimumSize(700, 1000)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()

    def init_ui(self):
        self.workspace_frame.resize(400,1000)
        self.workspace_frame.move(200,50)
        
        self.add_record_label.resize(self.label_width + 200,self.label_height)
        self.add_record_label.setFont(self.l_font)
        self.add_record_label.move(0,0)

        move_y = self.label_height + 30
        self.invoice_label.resize(self.label_width + 50,self.label_height)
        self.invoice_label.setFont(self.s_font)
        self.invoice_label.move(0,move_y)

        self.invoice_line.resize(self.line_width,self.line_height)
        self.invoice_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.invoice_line.move(0,move_y)
        self.invoice_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.invoice_line.setText(self.row_data[1])

        self.invoice_company_label.resize(self.label_width,self.label_height)
        self.invoice_company_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.invoice_company_label.move(0,move_y)

        self.invoice_company.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.invoice_company.move(0,move_y)
        self.setup_combo_box(self.invoice_company)
        self.company_data(self.invoice_company)
        self.invoice_company.setCurrentText(self.row_data[2])


        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.date_label.move(0,move_y)

        self.invoice_month.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.invoice_month.move(0,move_y)
        self.invoice_month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.invoice_month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.invoice_month.addItem("Not Chosen")
        self.invoice_month.addItems(months)
        self.invoice_month.setCurrentIndex(0)
        self.invoice_month.setCurrentText(self.row_data[3])


        self.invoice_year.resize(self.line_width,self.line_height)
        move_y = move_y + self.line_height + 10
        self.invoice_year.move(0,move_y)
        self.invoice_year.setMinimum(2022)
        self.invoice_year.setMaximum(2030)
        self.invoice_year.setValue(2023)
        self.invoice_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.invoice_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.invoice_year.setValue(int(self.row_data[4]))

        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.amount_line.setText(self.row_data[5])


        self.add_record_submit_button.resize(180,70)
        self.add_record_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_record_submit_button.move(50,move_y + 110)
        self.add_record_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.add_record_submit_button.clicked.connect(self.submit_button_clicked)
    
    def submit_button_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
        
        query = QSqlQuery()
        query.prepare(f"UPDATE Invoice SET no = :no, company = :company, month = :month, year = :year, amount = :amount WHERE id = {self.row_data[0]}")

        query.bindValue(":no", self.invoice_line.text())
        query.bindValue(":company", self.invoice_company.currentText())
        query.bindValue(":month", self.invoice_month.currentText())
        query.bindValue(":year", self.invoice_year.value())
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data updated successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Update")
        message_box.setText("Invoice Updated        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()
        self.updated = True
        self.close()
    
    def company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_updated(self):
        return self.updated

class UpdateSalaryDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,record,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\update_salary_dialog.ui', self)
        self.row_data = record
        self.setWindowTitle("Update Salary Record")
        self.setMinimumSize(700, 600)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()

    def init_ui(self):
        self.workspace_frame.resize(400,650)
        self.workspace_frame.move(150,50)
        
        self.update_salary_label.resize(self.label_width + 200,self.label_height)
        self.update_salary_label.setFont(self.l_font)
        self.update_salary_label.move(0,0)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(0,80)
        move_y = 80

        date_time = QDate.fromString(self.row_data[1], "dd/MM/yyyy")
        self.date_line.setDate(date_time)
        self.date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.date_line.setDisplayFormat("dd/MM/yyyy")
        self.date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.date_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.date_line.move(0,move_y)
        self.date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:10px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 28px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white }"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:20px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 35px; height: 35px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 3px solid black; font-family: Constantia; font-size: 28px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:50px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:50px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_left.png); }"
        )
    
        self.employee_label.resize(self.label_width,self.label_height)
        self.employee_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.employee_label.move(0,move_y)        

        self.employee_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.employee_line.move(0,move_y)
        self.setup_combo_box(self.employee_line)
        self.employee_data(self.employee_line)
        self.employee_line.setCurrentText(self.row_data[2])

        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.amount_line.setText(self.row_data[3])

        
        move_y = move_y + self.label_height - 15
        self.update_button.resize(180,70)
        self.update_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_button.move(50,move_y + 50)
        self.update_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_button.clicked.connect(self.update_button_clicked)

    def employee_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Employee")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def update_button_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
        
        query = QSqlQuery()

        query.prepare(f"UPDATE Salary SET date = :date, employee = :employee, amount = :amount WHERE id = {self.row_data[0]}")

        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":employee", self.employee_line.currentText())
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data updated successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Update")
        message_box.setText("Salary Record Updated        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()
        self.updated = True
        self.close()
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_updated(self):
        return self.updated

class UpdateEmployeeDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,employee,parent=None):
        super().__init__(parent)
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\update_employee_dialog.ui', self)
        self.data = self.get_employee(employee)
        self.setWindowTitle("Update Employee")
        self.setMinimumSize(700, 600)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()

    def init_ui(self):
        self.workspace_frame.resize(400,650)
        self.workspace_frame.move(150,50)
        
        self.update_employee_label.resize(self.label_width + 200,self.label_height)
        self.update_employee_label.setFont(self.l_font)
        self.update_employee_label.move(0,0)

        move_y = self.label_height + 30
        self.name_label.resize(self.label_width + 50,self.label_height)
        self.name_label.setFont(self.s_font)
        self.name_label.move(0,move_y)

        self.name_line.resize(self.line_width,self.line_height)
        self.name_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.name_line.move(0,move_y)
        self.name_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.name_line.setText(self.data[1])

        self.ic_label.resize(self.label_width,self.label_height)
        self.ic_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.ic_label.move(0,move_y)

        self.ic_line.resize(self.line_width,self.line_height)
        self.ic_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.ic_line.move(0,move_y)
        self.ic_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.ic_line.setText(self.data[2])

        self.lorry_label.resize(self.label_width,self.label_height)
        self.lorry_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.lorry_label.move(0,move_y)

        self.lorry_line.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.lorry_line.move(0,move_y)
        self.setup_combo_box(self.lorry_line)
        self.lorry_data(self.lorry_line)
        self.lorry_line.setCurrentText(self.data[3])

        self.update_button.resize(180,70)
        self.update_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_button.move(50,move_y + 110)
        self.update_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_button.clicked.connect(self.update_button_clicked)
    
    def update_button_clicked(self):
        query = QSqlQuery()
        query.prepare(f"UPDATE Employee SET name = :name, ic = :ic, lorry = :lorry WHERE id = {self.data[0]}")

        query.bindValue(":name", self.name_line.text())
        query.bindValue(":lorry", self.lorry_line.currentText())
        query.bindValue(":ic", self.ic_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data updated successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Update")
        message_box.setText("Employee Updated        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()
        self.updated = True
        self.close()
    
    def get_employee(self,name):
        query = QSqlQuery()
        if not query.exec(f"SELECT * FROM Employee where name = '{name}'"):
            print("Error executing query:", query.lastError().text())

        data = []
        if query.next():
            for i in range(4):
                data.append(query.value(i))
 
        return data
    
    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_updated(self):
        return self.updated

class AddInvoiceDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\add_invoice_dialog.ui', self)
        self.setWindowTitle("Add Invoice")
        self.setMinimumSize(700, 1000)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
    
        self.init_ui()

    def init_ui(self):
        self.workspace_frame.resize(400,1000)
        self.workspace_frame.move(200,50)
        
        self.add_record_label.resize(self.label_width + 200,self.label_height)
        self.add_record_label.setFont(self.l_font)
        self.add_record_label.move(0,0)

        move_y = self.label_height + 30
        self.invoice_label.resize(self.label_width + 50,self.label_height)
        self.invoice_label.setFont(self.s_font)
        self.invoice_label.move(0,move_y)

        self.invoice_line.resize(self.line_width,self.line_height)
        self.invoice_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.invoice_line.move(0,move_y)
        self.invoice_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")


        self.invoice_company_label.resize(self.label_width,self.label_height)
        self.invoice_company_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.invoice_company_label.move(0,move_y)

        self.invoice_company.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.invoice_company.move(0,move_y)
        self.setup_combo_box(self.invoice_company)
        self.company_data(self.invoice_company)

        
        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.date_label.move(0,move_y)

        self.invoice_month.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.invoice_month.move(0,move_y)
        self.invoice_month.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        self.invoice_month.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.invoice_month.addItem("Not Chosen")
        self.invoice_month.addItems(months)
        self.invoice_month.setCurrentIndex(0)


        self.invoice_year.resize(self.line_width,self.line_height)
        move_y = move_y + self.line_height + 10
        self.invoice_year.move(0,move_y)
        self.invoice_year.setMinimum(2022)
        self.invoice_year.setMaximum(2030)
        self.invoice_year.setValue(2023)
        self.invoice_year.setStyleSheet(
            "QSpinBox::up-button { width:50px; height:20px ;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { width:50px; height:20px ;image: url(photo/arrow_down.png)}"
            "QSpinBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"

            "QSpinBox:hover{ border: 3px solid black}"
        )
        self.invoice_year.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))

        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")


        self.add_record_submit_button.resize(180,70)
        self.add_record_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_record_submit_button.move(50,move_y + 110)
        self.add_record_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.add_record_submit_button.clicked.connect(self.submit_button_clicked)
    
    def submit_button_clicked(self):
        query = QSqlQuery()
        query.prepare("INSERT INTO Invoice(no,company,month,year,amount) VALUES (:no,:company,:month,:year,:amount)")
        
        query.bindValue(":no", self.invoice_line.text())
        query.bindValue(":company", self.invoice_company.currentText())
        query.bindValue(":month", self.invoice_month.currentText())
        query.bindValue(":year", self.invoice_year.value())
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data added successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Add Invoice")
        message_box.setText("Invoice Added        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()
        self.close()
    
    def company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_updated(self):
        return self.updated
    
class AddDebtorDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 600
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\add_debtor_dialog.ui', self)
        self.setWindowTitle("Add Debtor Company")
        self.setMinimumSize(800, 1100)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
    
        self.init_ui()

    def init_ui(self):
        self.workspace_frame.resize(800,1100)
        self.workspace_frame.move(100,10)
        
        self.add_record_label.resize(self.label_width + 200,self.label_height)
        self.add_record_label.setFont(self.l_font)
        self.add_record_label.move(0,0)

        move_y = self.label_height
        self.company_label.resize(self.label_width + 50,self.label_height)
        self.company_label.setFont(self.s_font)
        self.company_label.move(0,move_y)

        self.company_line.resize(self.line_width,self.line_height)
        self.company_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.company_line.move(0,move_y)
        self.company_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")


        self.address_label_1.resize(self.label_width,self.label_height)
        self.address_label_1.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_1.move(0,move_y)

        self.address_line_1.resize(self.line_width,self.line_height)
        self.address_line_1.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_1.move(0,move_y)
        self.address_line_1.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        
        self.address_label_2.resize(self.label_width,self.label_height)
        self.address_label_2.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_2.move(0,move_y)

        self.address_line_2.resize(self.line_width,self.line_height)
        self.address_line_2.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_2.move(0,move_y)
        self.address_line_2.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        
        self.address_label_3.resize(self.label_width,self.label_height)
        self.address_label_3.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_3.move(0,move_y)

        self.address_line_3.resize(self.line_width,self.line_height)
        self.address_line_3.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_3.move(0,move_y)
        self.address_line_3.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        
        self.address_label_4.resize(self.label_width,self.label_height)
        self.address_label_4.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_4.move(0,move_y)

        self.address_line_4.resize(self.line_width,self.line_height)
        self.address_line_4.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_4.move(0,move_y)
        self.address_line_4.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        
        self.address_label_5.resize(self.label_width,self.label_height)
        self.address_label_5.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_5.move(0,move_y)

        self.address_line_5.resize(self.line_width,self.line_height)
        self.address_line_5.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_5.move(0,move_y)
        self.address_line_5.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        self.tel_label.resize(self.label_width,self.label_height)
        self.tel_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.tel_label.move(0,move_y)

        self.tel_line.resize(self.line_width - 100,self.line_height)
        self.tel_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.tel_line.move(0,move_y)
        self.tel_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        self.submit_button.resize(180,70)
        self.submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.submit_button.move(200,move_y + 80)
        self.submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.submit_button.clicked.connect(self.submit_button_clicked)
    
    def submit_button_clicked(self):
        query = QSqlQuery()
        query.prepare("INSERT INTO Company(name,address_1,address_2,address_3,address_4,address_5,tel) VALUES (:name,:address_1,:address_2,:address_3,:address_4,:address_5,:tel)")
        
        query.bindValue(":name", self.company_line.text())
        query.bindValue(":address_1", self.address_line_1.text())
        query.bindValue(":address_2", self.address_line_2.text())
        query.bindValue(":address_3", self.address_line_3.text())
        query.bindValue(":address_4", self.address_line_4.text())
        query.bindValue(":address_5", self.address_line_5.text())
        query.bindValue(":tel", self.tel_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data added successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Add Debtor Company")
        message_box.setText("Debtor Company Added        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()
        self.close()
    
    def company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

class UpdateDebtorDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 600
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)

    def __init__(self,company,parent=None):
        super().__init__(parent)
        loadUi('ui\\update_debtor_dialog.ui', self)
        self.data = self.get_company(company)
        self.setWindowTitle("Update Debtor Company")
        self.setMinimumSize(800,1100)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()
    
    def init_ui(self):
        self.workspace_frame.resize(800,1100)
        self.workspace_frame.move(100,10)
        
        self.title_label.resize(self.label_width + 300,self.label_height)
        self.title_label.setFont(self.l_font)
        self.title_label.move(0,0)

        move_y = self.label_height
        self.company_label.resize(self.label_width + 50,self.label_height)
        self.company_label.setFont(self.s_font)
        self.company_label.move(0,move_y)

        self.company_line.resize(self.line_width,self.line_height)
        self.company_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.company_line.move(0,move_y)
        self.company_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.company_line.setText(self.data[1])
        self.ori_debtor = self.data[1]


        self.address_label_1.resize(self.label_width,self.label_height)
        self.address_label_1.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_1.move(0,move_y)

        self.address_line_1.resize(self.line_width,self.line_height)
        self.address_line_1.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_1.move(0,move_y)
        self.address_line_1.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.address_line_1.setText(self.data[2])
        
        self.address_label_2.resize(self.label_width,self.label_height)
        self.address_label_2.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_2.move(0,move_y)

        self.address_line_2.resize(self.line_width,self.line_height)
        self.address_line_2.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_2.move(0,move_y)
        self.address_line_2.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.address_line_2.setText(self.data[3])
        
        self.address_label_3.resize(self.label_width,self.label_height)
        self.address_label_3.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_3.move(0,move_y)

        self.address_line_3.resize(self.line_width,self.line_height)
        self.address_line_3.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_3.move(0,move_y)
        self.address_line_3.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.address_line_3.setText(self.data[4])   
        
        self.address_label_4.resize(self.label_width,self.label_height)
        self.address_label_4.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_4.move(0,move_y)

        self.address_line_4.resize(self.line_width,self.line_height)
        self.address_line_4.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_4.move(0,move_y)
        self.address_line_4.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.address_line_4.setText(self.data[5])
        
        self.address_label_5.resize(self.label_width,self.label_height)
        self.address_label_5.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.address_label_5.move(0,move_y)

        self.address_line_5.resize(self.line_width,self.line_height)
        self.address_line_5.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.address_line_5.move(0,move_y)
        self.address_line_5.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.address_line_5.setText(self.data[6])

        self.tel_label.resize(self.label_width,self.label_height)
        self.tel_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.tel_label.move(0,move_y)

        self.tel_line.resize(self.line_width - 100,self.line_height)
        self.tel_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.tel_line.move(0,move_y)
        self.tel_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.tel_line.setText(self.data[7])

        self.update_button.resize(180,70)
        self.update_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.update_button.move(200,move_y + 80)
        self.update_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.update_button.clicked.connect(self.update_button_clicked)
    
    def update_button_clicked(self):
        query = QSqlQuery()

        self.new_debtor = self.company_line.text()
        print(f"ori debtor:{self.ori_debtor}")
        print(f"new debtor:{self.new_debtor}")
        if self.ori_debtor != self.new_debtor:
            query.exec(f"SELECT * FROM Record WHERE company = '{self.ori_debtor}'")
            while query.next():
                id = query.value(0)
                new_query = QSqlQuery()
                str = f"UPDATE Record SET company = '{self.new_debtor}' WHERE id={id}"
                new_query.exec(str)
            
            query.exec(f"SELECT * FROM Invoice WHERE company = '{self.ori_debtor}'")
            while query.next():
                id = query.value(0)
                new_query = QSqlQuery()
                str = f"UPDATE Invoice SET company = '{self.new_debtor}' WHERE id={id}"
                new_query.exec(str)
               
        query.prepare(f"UPDATE Company SET name = :name, address_1 = :address_1,address_2 = :address_2,address_3 = :address_3 ,address_4 = :address_4,address_5 = :address_5,tel = :tel WHERE id = {self.data[0]}")

        query.bindValue(":name", self.company_line.text())
        query.bindValue(":address_1", self.address_line_1.text())
        query.bindValue(":address_2", self.address_line_2.text())
        query.bindValue(":address_3", self.address_line_3.text())
        query.bindValue(":address_4", self.address_line_4.text())
        query.bindValue(":address_5", self.address_line_5.text())
        query.bindValue(":tel", self.tel_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data added successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Update Debtor Company")
        message_box.setText("Debtor Company Updated        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()
        self.close()

    def company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_company(self,name):
        query = QSqlQuery()
        if not query.exec(f"SELECT * FROM Company where name = '{name}'"):
            print("Error executing query:", query.lastError().text())

        data = []
        if query.next():
            for i in range(8):
                data.append(query.value(i))
 
        return data

class UpdateCreditorExpensesDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,record,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\update_creditor_expenses_dialog.ui', self)
        self.row_data = record
        
        self.setWindowTitle("Update Expenses")
        self.setMinimumSize(1000, 1200)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()

    def init_ui(self):
        self.workspace_frame.resize(400,1200)
        self.workspace_frame.move(300,50)
        
        self.add_record_label.resize(self.label_width + 200,self.label_height)
        self.add_record_label.setFont(self.l_font)
        self.add_record_label.move(0,0)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(0,80)
        move_y = 80

        date_time = QDate.fromString(self.row_data[1], "dd/MM/yyyy")
        self.date_line.setDate(date_time)
        self.date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.date_line.setDisplayFormat("dd/MM/yyyy")
        self.date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.date_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.date_line.move(0,move_y)
        self.date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:10px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 28px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white }"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:20px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 35px; height: 35px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 3px solid black; font-family: Constantia; font-size: 28px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:50px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:50px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_left.png); }"
        )
    
        self.lorry_label.resize(self.label_width,self.label_height)
        self.lorry_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.lorry_label.move(0,move_y)        

        self.lorry_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.lorry_line.move(0,move_y)
        self.setup_combo_box(self.lorry_line)
        self.lorry_data(self.lorry_line)
        self.lorry_line.setCurrentText(self.row_data[2])


        self.reference_label.resize(self.label_width,self.label_height)
        self.reference_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.reference_label.move(0,move_y)     

        self.reference_line.resize(self.line_width,self.line_height)
        self.reference_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.reference_line.move(0,move_y)
        self.reference_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")
        self.reference_line.setText(self.row_data[3])  
        
        

        self.expenses_label.resize(self.label_width,self.label_height)
        self.expenses_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.expenses_label.move(0,move_y)        
        
        self.expenses_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.expenses_line.move(0,move_y)
        self.setup_combo_box(self.expenses_line)
        self.expenses_data(self.expenses_line)
        self.expenses_line.setCurrentText(self.row_data[4])
       
    
        self.company_label.resize(self.label_width,self.label_height)
        self.company_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.company_label.move(0,move_y)

        self.company_line.resize(self.line_width+80,self.line_height)
        move_y = move_y + self.label_height - 15
        self.company_line.move(0,move_y)
        self.setup_combo_box(self.company_line)
        self.expenses_company_data(self.company_line)
        self.company_line.setCurrentText(self.row_data[5])
       
        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)


        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)      
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        self.amount_line.setText(self.row_data[6])
        
       
        self.add_record_submit_button.resize(180,70)
        self.add_record_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_record_submit_button.move(50,move_y + 110)
        self.add_record_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.add_record_submit_button.clicked.connect(self.submit_button_clicked)
    
    def submit_button_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
        query = QSqlQuery()

        query.prepare(f"UPDATE Expenses_Record SET date = :date, lorry = :lorry, reference = :reference, expenses = :expenses,company = :company,amount = :amount WHERE id = {self.row_data[0]}")        
        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":lorry", set_not_chosen(self.lorry_line.currentText()))
        query.bindValue(":reference", self.reference_line.text())
        query.bindValue(":expenses", self.expenses_line.currentText())
        query.bindValue(":company", set_not_chosen(self.company_line.currentText()))
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data updated successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Update")
        message_box.setText("Expenses Record Updated        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()

        self.updated = True
        self.close()

    def expenses_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT description FROM Expenses")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
  
    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def expenses_company_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT name FROM Expenses_Company")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_updated(self):
        return self.updated

class UpdateLorryExpensesDialog(QDialog):
    label_width = 200
    label_height = 80
    line_width = 300
    line_height = 50
    l_font = QFont("Tw Cen MT Condensed Extra Bold",20)
    s_font = QFont("Tw Cen MT Condensed Extra Bold",15)
    row_data = None
    updated = False

    def __init__(self,record,parent=None):
        super().__init__(parent)
    
        # 加载并应用 UpdateRecordDialog 的 Ui 文件
        loadUi('ui\\update_lorry_expenses_dialog.ui', self)
        self.row_data = record
        
        self.setWindowTitle("Update Expenses")
        self.setMinimumSize(1000, 800)
        self.setStyleSheet("background-color: rgba(216,216,216,255);");
        self.init_ui()

    def init_ui(self):
        self.workspace_frame.resize(400,1200)
        self.workspace_frame.move(300,50)
        
        self.add_record_label.resize(self.label_width + 200,self.label_height)
        self.add_record_label.setFont(self.l_font)
        self.add_record_label.move(0,0)

        self.date_label.resize(self.label_width,self.label_height)
        self.date_label.setFont(self.s_font)
        self.date_label.move(0,80)
        move_y = 80

        date_time = QDate.fromString(self.row_data[1], "dd/MM/yyyy")
        self.date_line.setDate(date_time)
        self.date_line.setFont(QFont("Tw Cen MT Condensed Extra Bold",14))
        self.date_line.setDisplayFormat("dd/MM/yyyy")
        self.date_line.setCalendarPopup(True)  # 设置为True以展开日历
        self.date_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.date_line.move(0,move_y)
        self.date_line.setStyleSheet(
            "QDateEdit { background-color: white; border: 2px solid black; color: black; padding-left:10px}"
            "QDateEdit:hover{ border: 3px solid black}"
            "QCalendarWidget { font-family: Constantia; font-size: 28px; }"
            "QDateEdit QCalendarWidget QTableView { color: black; background-color:white }"
            "QDateEdit QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: white; color : black; height:20px}"
            "QDateEdit::drop-down { image: url(photo/arrow_down.png); width: 35px; height: 35px; subcontrol-position: right;}"
            "QCalendarWidget QToolButton{"
                    "color: black;background-color: white;}"
            "QCalendarWidget QMenu { background-color: white; border : 3px solid black; font-family: Constantia; font-size: 28px;}"
            "QCalendarWidget QMenu::item:selected { background-color: black; color: white; }"
            "QSpinBox::up-button { subcontrol-origin: border;""subcontrol-position: top right; width:50px;image: url(photo/arrow_up.png)}"
            "QSpinBox::down-button { subcontrol-origin: border;""subcontrol-position: bottom right; width:50px;image: url(photo/arrow_down.png)}"
            "QCalendarWidget QToolButton#qt_calendar_nextmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_right.png); }"
            "QCalendarWidget QToolButton#qt_calendar_prevmonth { background-color:transparent; width: 50px; height: 50px; icon-size: 40px, 40px; qproperty-icon: url(photo/arrow_left.png); }"
        )
    
        self.lorry_label.resize(self.label_width,self.label_height)
        self.lorry_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.lorry_label.move(0,move_y)        

        self.lorry_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.lorry_line.move(0,move_y)
        self.setup_combo_box(self.lorry_line)
        self.lorry_data(self.lorry_line)
        self.lorry_line.setCurrentText(self.row_data[2])

        self.expenses_label.resize(self.label_width,self.label_height)
        self.expenses_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.expenses_label.move(0,move_y)        
        
        self.expenses_line.resize(self.line_width,self.line_height)
        move_y = move_y + self.label_height - 15
        self.expenses_line.move(0,move_y)
        self.setup_combo_box(self.expenses_line)
        self.expenses_data(self.expenses_line)
        self.expenses_line.setCurrentText(self.row_data[4])
          
        self.amount_label.resize(self.label_width,self.label_height)
        self.amount_label.setFont(self.s_font)
        move_y = move_y + self.line_height + 10
        self.amount_label.move(0,move_y)

        self.amount_line.resize(self.line_width,self.line_height)
        self.amount_line.setFont(self.s_font)
        move_y = move_y + self.label_height - 15
        self.amount_line.move(0,move_y)      
        self.amount_line.setStyleSheet("QLineEdit { border: 2px solid black; color:black; background-color:white; }" 
                                          "QLineEdit:hover{ border: 3px solid black}")

        self.amount_line.setText(self.row_data[6])
        
       
        self.add_record_submit_button.resize(180,70)
        self.add_record_submit_button.setFont(QFont("Tw Cen MT Condensed Extra Bold",18))
        self.add_record_submit_button.move(50,move_y + 110)
        self.add_record_submit_button.setStyleSheet("background-color:black;color:white;border-radius:35px;")
        self.add_record_submit_button.clicked.connect(self.submit_button_clicked)
    
    def submit_button_clicked(self):
        if not is_numeric_string(self.amount_line.text()):
            self.show_message_box("Warning","Amount Format Wrong!")
            return
        query = QSqlQuery()

        query.prepare(f"UPDATE Expenses_Record SET date = :date, lorry = :lorry, reference = '-', expenses = :expenses,company = '-',amount = :amount WHERE id = {self.row_data[0]}")        
        query.bindValue(":date", self.date_line.date().toString("dd/MM/yyyy"))
        query.bindValue(":lorry", set_not_chosen(self.lorry_line.currentText()))
        query.bindValue(":expenses", set_not_chosen(self.expenses_line.currentText()))
        query.bindValue(":amount", self.amount_line.text())

        # Create a QSqlQuery to execute the SQL insert statement
        if query.exec():
            print("Data updated successfully.")
        else:
            print("Error while inserting data:", query.lastError().text())

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.NoIcon)
        message_box.setWindowTitle("Update")
        message_box.setText("Expenses Record Updated        ")
        message_box.setStandardButtons(QMessageBox.Ok)
        message_box.setStyleSheet(
            "QMessageBox { font-family:Times New Roman; font-size:26px;}"
            "QPushButton { background-color: white; color: black;}"
        )

        # 显示弹窗
        message_box.exec_()

        self.updated = True
        self.close()

    def expenses_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT description FROM Expenses")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)
        combobox.setCurrentIndex(0)
  
    def lorry_data(self,combobox):
        combobox.clear()
        combobox.addItem("Not Chosen")
        query = QSqlQuery("SELECT lorry_number FROM Lorry")
        while query.next():
            name = query.value(0)
            combobox.addItem(name)

        combobox.setCurrentIndex(0)
    
    def setup_combo_box(self,combobox):
        combobox.setFont(QFont("Tw Cen MT Condensed Extra Bold",13))
        combobox.setStyleSheet(
            "QComboBox {"
            "   background-color: white; "
            "   border: 2px solid black; "
            "   color: black;"
            "   padding : 5px; "
            "}"
            "QComboBox::drop-down { image: url(photo/arrow_down.png); width: 40px; height: 40px ;padding-top: 3px; }"
            "QComboBox:hover{ border: 3px solid black}"
        )

    def get_updated(self):
        return self.updated
